import asyncpg
import json
import time
from datetime import date, datetime
from typing import List, Dict, Any, FrozenSet, Optional
from zoneinfo import ZoneInfo
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.logger import AuditLogger
from core.exceptions import RoomNotFoundError, StudentNotFoundError, ForbiddenError, ValidationError
from core.rbac import require_permission, require_member
from core.config import settings
from core.name_utils import normalize_nfc, normalize_en, identity_pair
from services.action_service import ActionService
from models.student_schemas import StudentUpdateRequest

service_logger = AuditLogger(service_name="STUDENT")

STUDENT_PATCHABLE_COLUMNS: FrozenSet[str] = frozenset(StudentUpdateRequest.model_fields.keys())

GLOBAL_FIELDS: FrozenSet[str] = frozenset([
    'prefix', 'first_name', 'last_name', 'nickname',
    'first_name_en', 'last_name_en', 'nickname_en', 'birthday',
    'blood_group', 'shirt_size', 'food_allergy', 'congenital_disease',
    'phone_number', 'phone_number_parent', 'phone_number_parent_relation',
    'line_id', 'ig_username', 'email',
    'address_house_no', 'address_road', 'address_sub_district',
    'address_district', 'address_province', 'address_post_code'
])

# ฟิลด์ชื่อที่ต้อง NFC-normalize ก่อนเขียนลง users (แก้ อำ/อํา ฯลฯ; normalize_nfc = NFC + strip)
NAME_NFC_FIELDS: FrozenSet[str] = frozenset(['prefix', 'first_name', 'last_name', 'nickname', 'nickname_en'])

LOCAL_FIELDS: FrozenSet[str] = frozenset([
    'student_id', 'class_role', 'cleaning_duty', 'olympic_camp',
    'portfolio', 'target_faculty', 'is_admin', 'permissions', 'status'
])

THAI_TZ = ZoneInfo("Asia/Bangkok")

THAI_MONTH_NAMES: List[str] = [
    "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
    "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
]

# หัวคอลัมน์ภาษาไทยสำหรับ Export (คีย์ = field ใน BASE_STUDENT_SELECT)
EXPORT_HEADER_LABELS: Dict[str, str] = {
    "student_no": "เลขที่",
    "student_id": "รหัสนักเรียน",
    "prefix": "คำนำหน้า",
    "first_name": "ชื่อจริง",
    "last_name": "นามสกุล",
    "nickname": "ชื่อเล่น",
    "first_name_en": "ชื่อจริง (EN)",
    "last_name_en": "นามสกุล (EN)",
    "nickname_en": "ชื่อเล่น (EN)",
    "birthday": "วันเกิด",
    "class_role": "บทบาทในห้อง",
    "cleaning_duty": "เวรทำความสะอาด",
    "olympic_camp": "สอวน. / ค่าย",
    "target_faculty": "คณะเป้าหมาย",
    "portfolio": "ผลงาน",
    "blood_group": "กรุ๊ปเลือด",
    "shirt_size": "ไซส์เสื้อ",
    "food_allergy": "แพ้อาหาร",
    "congenital_disease": "โรคประจำตัว",
    "phone_number": "เบอร์โทรศัพท์",
    "phone_number_parent": "เบอร์ผู้ปกครอง",
    "phone_number_parent_relation": "ความสัมพันธ์",
    "line_id": "LINE ID",
    "ig_username": "IG Username",
    "email": "อีเมล",
    "address_house_no": "บ้านเลขที่/หมู่/ซอย",
    "address_road": "ถนน",
    "address_sub_district": "ตำบล/แขวง",
    "address_district": "อำเภอ/เขต",
    "address_province": "จังหวัด",
    "address_post_code": "รหัสไปรษณีย์",
    "status": "สถานะ",  # defensive: มีใน BASE_STUDENT_SELECT แม้ frontend ไม่ให้เลือก
}

ROLE_LABELS: Dict[str, str] = {
    "student": "นักเรียน",
    "president": "หัวหน้าห้อง",
    "vice_president": "รองหัวหน้าห้อง",
    "secretary": "เลขานุการ (เรขา)",
    "vice_academic": "รองวิชาการ",
    "vice_activity": "รองกิจกรรม",
    "vice_discipline": "รองระเบียบวินัย",
    "vice_reception": "รองปฏิคม",
    "vice_pr": "รองประชาสัมพันธ์",
    "vice_sanitation": "รองสุขาภิบาล",
    "staff_academic": "กรรมการวิชาการ",
    "staff_activity": "กรรมการกิจกรรม",
    "staff_discipline": "กรรมการระเบียบวินัย",
    "staff_reception": "กรรมการปฏิคม",
    "staff_pr": "กรรมการประชาสัมพันธ์",
    "staff_sanitation": "กรรมการสุขาภิบาล",
    "treasurer": "เหรัญญิก",
    "admin": "ผู้ดูแลระบบ",
}

STATUS_LABELS: Dict[str, str] = {
    "active": "กำลังเรียน",
    "pending": "รออนุมัติ",
    "inactive": "พ้นสภาพ",
}

DEFAULT_COL_WIDTH = 18
EXPORT_COLUMN_WIDTHS: Dict[str, int] = {
    "student_no": 8, "student_id": 14, "prefix": 10,
    "first_name": 16, "last_name": 16, "nickname": 14,
    "first_name_en": 16, "last_name_en": 16, "nickname_en": 14,
    "birthday": 22,
    "class_role": 18, "cleaning_duty": 20, "olympic_camp": 24,
    "target_faculty": 18, "portfolio": 30,
    "blood_group": 10, "shirt_size": 10, "food_allergy": 20, "congenital_disease": 20,
    "phone_number": 18, "phone_number_parent": 18, "phone_number_parent_relation": 14,
    "line_id": 16, "ig_username": 16, "email": 26,
    "address_house_no": 24, "address_road": 20, "address_sub_district": 20,
    "address_district": 20, "address_province": 14, "address_post_code": 14,
}

CENTER_FIELDS: FrozenSet[str] = frozenset({"student_no", "blood_group", "shirt_size", "address_post_code"})
WRAP_TEXT_FIELDS: FrozenSet[str] = frozenset({
    "portfolio", "olympic_camp", "cleaning_duty", "food_allergy",
    "congenital_disease", "address_house_no",
})

class StudentService:

    @staticmethod
    async def resolve_room_id(conn: asyncpg.Connection, server_id: Optional[int] = None, room_id: Optional[int] = None) -> int:
        if room_id:
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE id = $1 AND deleted_at IS NULL", room_id):
                raise RoomNotFoundError("ไม่พบห้องเรียนนี้")
            return room_id
        if server_id:
            r_id = await conn.fetchval("SELECT id FROM rooms WHERE server_id = $1 AND deleted_at IS NULL", server_id)
            if not r_id: 
                raise RoomNotFoundError(f"ไม่พบห้องสำหรับ server {server_id}")
            return r_id
        raise ValueError("ต้องระบุ server_id หรือ room_id")

    BASE_STUDENT_SELECT = """
        SELECT 
            s.id, s.room_id, u.id as user_id, u.discord_id, s.student_no, s.student_id,
            u.prefix, u.first_name, u.last_name, u.nickname,
            u.first_name_en, u.last_name_en, u.nickname_en, u.birthday,
            s.class_role, s.cleaning_duty, s.olympic_camp, s.portfolio, s.target_faculty,
            u.blood_group, u.shirt_size, u.food_allergy, u.congenital_disease,
            u.phone_number, u.phone_number_parent, u.phone_number_parent_relation,
            u.line_id, u.ig_username, u.email,
            u.address_house_no, u.address_road, u.address_sub_district, u.address_district, u.address_province, u.address_post_code,
            s.status, s.is_admin, s.permissions, s.created_at, s.updated_at
        FROM students s
        LEFT JOIN users u ON s.user_id = u.id
    """

    @staticmethod
    def _parse_permissions(perms: Any) -> List[str]:
        if not perms: return []
        if isinstance(perms, list): return perms
        if isinstance(perms, str):
            try: return json.loads(perms)
            except: return []
        return []

    @staticmethod
    def _calculate_completion(row: dict) -> dict:
        expected_fields = [
            'student_id', 'prefix', 'nickname', 'birthday',
            'cleaning_duty', 'olympic_camp', 'target_faculty',
            'blood_group', 'shirt_size', 'food_allergy', 'congenital_disease',
            'phone_number', 'phone_number_parent', 'phone_number_parent_relation',
            'line_id', 'ig_username', 'email',
            'address_house_no', 'address_road', 'address_sub_district',
            'address_district', 'address_province', 'address_post_code'
        ]
        missing = [f for f in expected_fields if not row.get(f) or str(row.get(f)).strip() == ""]
        total = len(expected_fields)
        filled = total - len(missing)
        percent = int((filled / total) * 100)
        return {"percentage": percent, "missing_fields": missing}

    @staticmethod
    async def _find_or_create_user(
        conn: asyncpg.Connection,
        first_name: str,
        last_name: str,
        first_name_en: str = "",
        last_name_en: str = "",
        nickname: str = "",
        nickname_en: str = "",
    ) -> int:
        """หา user ตาม "กุญแจตัวตน" — ชื่ออังกฤษก่อน (English-primary) ถ้าไม่มีอังกฤษ
        หรือหาไม่เจอ → fallback เป็นชื่อไทยแบบ NFC-normalized (แก้ อำ/อํา match กันไม่เจอ).

        ทั้งชื่อไทยและอังกฤษถูก normalize ก่อนเก็บ (ดู core/name_utils) เพื่อให้
        exact-match ตรงกันเสมอ. คืน user_id เดิมถ้ามีอยู่แล้ว ไม่งั้นสร้าง ghost user."""
        th_first = normalize_nfc(first_name)
        th_last = normalize_nfc(last_name)
        en_first = normalize_nfc(first_name_en)
        en_last = normalize_nfc(last_name_en)
        th_nickname = normalize_nfc(nickname)
        en_nickname = normalize_nfc(nickname_en)

        if en_first or en_last:
            # 1) ชื่ออังกฤษเป็นกุญแจหลัก — ค้นแบบไม่ไวตัวพิมพ์ (LOWER) สอดคล้องกับ identity_pair
            #    ที่ใช้ normalize_en (casefold) ใน bulk_add / join_room
            user_id = await conn.fetchval(
                "SELECT id FROM users WHERE LOWER(first_name_en) = $1 AND LOWER(last_name_en) = $2 AND deleted_at IS NULL",
                normalize_en(first_name_en), normalize_en(last_name_en),
            )
            if user_id:
                return user_id
            # 2) fallback: user เก่าที่มีแต่ชื่อไทย (อังกฤษยังว่าง) — กันสร้าง user ซ้ำตอนกรอกชื่ออังกฤษทีหลัง
            if th_first or th_last:
                user_id = await conn.fetchval(
                    "SELECT id FROM users WHERE first_name = $1 AND last_name = $2 AND deleted_at IS NULL",
                    th_first, th_last,
                )
                if user_id:
                    return user_id
        else:
            user_id = await conn.fetchval(
                "SELECT id FROM users WHERE first_name = $1 AND last_name = $2 AND deleted_at IS NULL",
                th_first, th_last,
            )
            if user_id:
                return user_id

        return await conn.fetchval(
            "INSERT INTO users (first_name, last_name, nickname, first_name_en, last_name_en, nickname_en) VALUES ($1, $2, $3, $4, $5, $6) RETURNING id",
            th_first or None, th_last or None, th_nickname or None,
            en_first or None, en_last or None, en_nickname or None,
        )

    @classmethod
    async def add_student(cls, pool: asyncpg.Pool, student_no: int, first_name: str, last_name: str, user_name: str, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, actor_user_id: Optional[int] = None, first_name_en: str = "", last_name_en: str = "", nickname: str = "", nickname_en: str = ""):
        start_time = time.time()
        target_room_id = None
        new_values = {"student_no": student_no, "first_name": first_name, "last_name": last_name, "first_name_en": first_name_en or None, "last_name_en": last_name_en or None, "nickname": nickname or None, "nickname_en": nickname_en or None, "user_name": user_name}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    # 🛡️ RBAC: ต้องมี MANAGE_STUDENTS ถึงจะเพิ่มนักเรียนได้ (กันนักเรียนธรรมดาเพิ่มเพื่อนเอง)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_STUDENTS")
                    user_id = await cls._find_or_create_user(conn, first_name, last_name, first_name_en, last_name_en, nickname, nickname_en)

                    res = await conn.execute("""
                        INSERT INTO students (room_id, student_no, user_id, status) 
                        SELECT $1, $2, $3, 'active' WHERE NOT EXISTS (
                            SELECT 1 FROM students WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL
                        )
                    """, target_room_id, student_no, user_id)
                    
                    if res == "INSERT 0 1":
                        exec_time = int((time.time() - start_time) * 1000)
                        await service_logger.log(
                            conn=conn, action="CREATE", actor_identifier=actor_identifier,
                            client_source=client_source, room_id=target_room_id, user_id=user_id,
                            entity_type="STUDENT", entity_id=str(student_no), status="success",
                            new_values=new_values, endpoint_or_command="add_student", execution_time_ms=exec_time
                        )
                        # 📢 แจ้งเตือน Discord: มีสมาชิกใหม่ (ไม่ @everyone)
                        room_server_id = await conn.fetchval(
                            "SELECT server_id FROM rooms WHERE id = $1 AND deleted_at IS NULL", target_room_id
                        )
                        if room_server_id:
                            await ActionService.notify_new_student(
                                server_id=room_server_id,
                                student_no=student_no,
                                first_name=first_name,
                                last_name=last_name,
                                first_name_en=first_name_en,
                                last_name_en=last_name_en,
                                user_name=user_name,
                            )
                    else:
                        raise ValueError(f"เลขที่ {student_no} มีรายชื่ออยู่ในห้องนี้แล้ว")
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, entity_type="STUDENT",
                    entity_id=str(student_no), status="failed", error_detail=str(e),
                    new_values=new_values, endpoint_or_command="add_student", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def bulk_add_students(cls, pool: asyncpg.Pool, students: List[dict], user_name: str, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, actor_user_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        new_values = {"students": students, "user_name": user_name}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    # 🛡️ RBAC: ต้องมี MANAGE_STUDENTS ถึงจะเพิ่มนักเรียน bulk ได้
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_STUDENTS")
                    # 🌟 identity: ชื่ออังกฤษเป็นกุญแจหลัก (English-primary), fallback ไทย NFC
                    # — ใช้ _find_or_create_user จุดเดียวกับ add_student กัน dedupe พัง
                    user_map = {}
                    user_by_index = {}
                    for idx, s in enumerate(students):
                        key = identity_pair(
                            s.get('first_name_en'), s.get('last_name_en'),
                            s['first_name'], s['last_name'],
                        )
                        if key not in user_map:
                            user_map[key] = await cls._find_or_create_user(
                                conn, s['first_name'], s['last_name'],
                                s.get('first_name_en') or '', s.get('last_name_en') or '',
                                s.get('nickname') or '', s.get('nickname_en') or '',
                            )
                        user_by_index[idx] = user_map[key]

                    student_tuples = [(target_room_id, s['student_no'], user_by_index[i]) for i, s in enumerate(students)]
                    await conn.executemany("""
                        INSERT INTO students (room_id, student_no, user_id, status) 
                        SELECT $1, $2, $3, 'active' WHERE NOT EXISTS (
                            SELECT 1 FROM students WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL
                        )
                    """, student_tuples)
                    
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, entity_type="STUDENT_BULK",
                        status="success", new_values=new_values, endpoint_or_command="bulk_add_students", execution_time_ms=exec_time
                    )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, entity_type="STUDENT_BULK",
                    status="failed", error_detail=str(e), new_values=new_values, 
                    endpoint_or_command="bulk_add_students", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def update_student(cls, pool: asyncpg.Pool, student_no: int, update_data: dict, updater_user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        new_values = update_data.copy()
        old_values = None
        
        new_student_no = update_data.pop('new_student_no', None)
        
        if 'permissions' in update_data:
            update_data['permissions'] = json.dumps(update_data['permissions']) if update_data['permissions'] else '[]'

        clean_data = {k: v for k, v in update_data.items() if v is not None and k in STUDENT_PATCHABLE_COLUMNS}
        if not clean_data and new_student_no is None: return

        # 🌟 NFC-normalize ชื่อไทยก่อนเขียนลง users (แก้ อำ/อํา ฯลฯ ให้ exact-match ตรงกัน)
        for k in list(clean_data):
            if k in NAME_NFC_FIELDS:
                clean_data[k] = normalize_nfc(clean_data[k])

        global_updates = {k: v for k, v in clean_data.items() if k in GLOBAL_FIELDS}
        local_updates = {k: v for k, v in clean_data.items() if k in LOCAL_FIELDS}

        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    
                    old_row = await conn.fetchrow(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND s.student_no = $2 AND s.deleted_at IS NULL", target_room_id, student_no)
                    if old_row: old_values = dict(old_row)

                    target_info = await conn.fetchrow(
                        "SELECT user_id FROM students WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL", 
                        target_room_id, student_no
                    )
                    if not target_info: raise StudentNotFoundError("ไม่พบเลขที่นี้")
                    target_user_id = target_info['user_id']

                    actor_row = await conn.fetchrow("SELECT is_admin, permissions FROM students WHERE room_id = $1 AND user_id = $2", target_room_id, updater_user_id)
                    actor_is_god = actor_row['is_admin'] if actor_row else False
                    
                    from core.config import settings
                    is_super_admin = settings.SUPER_ADMIN_ID and int(updater_user_id) == int(settings.SUPER_ADMIN_ID)
                    if is_super_admin: actor_is_god = True

                    is_editing_self = (target_user_id == updater_user_id)
                    has_manage_permission = False
                    
                    try:
                        await require_permission(conn, target_room_id, updater_user_id, "MANAGE_STUDENTS")
                        has_manage_permission = True
                    except ForbiddenError:
                        pass

                    if not is_editing_self and not has_manage_permission:
                        raise ForbiddenError("คุณไม่มีสิทธิ์แก้ไขข้อมูลของผู้อื่น")

                    if not actor_is_god:
                        local_updates.pop('is_admin', None)
                        local_updates.pop('permissions', None)
                        if not has_manage_permission:
                            local_updates.pop('class_role', None)
                            local_updates.pop('status', None)
                            if new_student_no and new_student_no != student_no:
                                raise ForbiddenError("คุณไม่มีสิทธิ์แก้ไขเลขที่ของตนเอง")

                    if new_student_no and new_student_no != student_no:
                        exists = await conn.fetchval("SELECT 1 FROM students WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL", target_room_id, new_student_no)
                        if exists: raise ValidationError(f"เลขที่ {new_student_no} มีคนใช้ไปแล้วในห้องนี้")
                        
                        await conn.execute("UPDATE students SET student_no = $1, updated_at = CURRENT_TIMESTAMP WHERE room_id = $2 AND student_no = $3", new_student_no, target_room_id, student_no)
                        student_no = new_student_no 

                    if global_updates and target_user_id:
                        keys = sorted(global_updates.keys())
                        set_clauses = [f"{key} = ${i+2}" for i, key in enumerate(keys)]
                        values = [target_user_id] + [global_updates[k] for k in keys]
                        await conn.execute(f"UPDATE users SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP WHERE id = $1", *values)

                    if local_updates:
                        keys = sorted(local_updates.keys())
                        set_clauses = [f"{key} = ${i+3}" if key != 'permissions' else f"{key} = ${i+3}::jsonb" for i, key in enumerate(keys)]
                        values = [target_room_id, student_no] + [local_updates[k] for k in keys]
                        await conn.execute(f"UPDATE students SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP WHERE room_id = $1 AND student_no = $2", *values)

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=updater_user_id,
                        entity_type="STUDENT", entity_id=str(new_student_no or student_no), status="success",
                        old_values=old_values, new_values=new_values, endpoint_or_command="update_student", execution_time_ms=exec_time
                    )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=updater_user_id,
                    entity_type="STUDENT", entity_id=str(student_no), status="failed", error_detail=str(e),
                    old_values=old_values, new_values=new_values, endpoint_or_command="update_student", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def get_student_by_user_id(cls, pool: asyncpg.Pool, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                row = await conn.fetchrow(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND u.id = $2 AND s.deleted_at IS NULL", target_room_id, user_id)
                if not row: raise StudentNotFoundError("คุณยังไม่มีรายชื่อนักเรียนในห้องนี้")
                data = dict(row)
                data['permissions'] = cls._parse_permissions(data['permissions'])
                data['data_completion'] = cls._calculate_completion(data)
                
                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT", status="success", endpoint_or_command="get_student_by_user_id", execution_time_ms=exec_time
                )
                return data
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT", status="failed", error_detail=str(e), 
                    endpoint_or_command="get_student_by_user_id", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def get_all_students(cls, pool: asyncpg.Pool, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> List[dict]:
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                is_member = await conn.fetchval("SELECT 1 FROM students WHERE room_id = $1 AND user_id = $2 AND deleted_at IS NULL", target_room_id, user_id)
                
                from core.config import settings
                is_super_admin = settings.SUPER_ADMIN_ID and int(user_id) == int(settings.SUPER_ADMIN_ID)
                
                if not is_member and not is_super_admin: raise ForbiddenError("คุณไม่มีสิทธิ์ดูรายชื่อ เพราะไม่ได้อยู่ในห้องเรียนนี้")

                rows = await conn.fetch(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND s.deleted_at IS NULL ORDER BY s.student_no ASC", target_room_id)
                
                results = []
                for row in rows:
                    full_data = dict(row)
                    results.append({
                        "id": full_data["id"],
                        "student_no": full_data["student_no"],
                        "student_id": full_data.get("student_id"),
                        "first_name": full_data["first_name"],
                        "last_name": full_data["last_name"],
                        "nickname": full_data.get("nickname"),
                        "first_name_en": full_data.get("first_name_en"),
                        "last_name_en": full_data.get("last_name_en"),
                        "nickname_en": full_data.get("nickname_en"),
                        "class_role": full_data["class_role"],
                        "status": full_data["status"],
                        "is_admin": full_data.get("is_admin", False),
                        "discord_id_str": str(full_data['discord_id']) if full_data.get('discord_id') else None,
                        "data_completion": cls._calculate_completion(full_data)
                    })
                
                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT_LIST", status="success", endpoint_or_command="get_all_students", execution_time_ms=exec_time
                )
                return results
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT_LIST", status="failed", error_detail=str(e), 
                    endpoint_or_command="get_all_students", execution_time_ms=exec_time
                )
            raise e

    @staticmethod
    def _format_buddhist_birthday(value: Any) -> str:
        """วันเกิดแบบไทย: '25 กรกฎาคม 2553' (ปี พ.ศ. = ค.ศ. + 543)."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            value = value.date()
        if not isinstance(value, date):
            return str(value)
        return f"{value.day} {THAI_MONTH_NAMES[value.month - 1]} {value.year + 543}"

    @staticmethod
    def _translate_value(field: str, value: Any) -> Any:
        """แปลงค่าตามชนิดคอลัมน์ (birthday → พ.ศ., class_role/status → ไทย, None → '')."""
        if value is None:
            return ""
        if field == "birthday":
            return StudentService._format_buddhist_birthday(value)
        if field == "class_role":
            return ROLE_LABELS.get(str(value), value)
        if field == "status":
            return STATUS_LABELS.get(str(value), value)
        return value

    @classmethod
    def _build_student_workbook(
        cls,
        room_name: str,
        room_code: Optional[str],
        fields: List[str],
        rows: List[dict],
        generated_at: Optional[datetime] = None,
    ) -> io.BytesIO:
        """สร้าง Workbook 2 แผ่น: 'สรุป' (ภาพรวมห้อง) + 'รายชื่อ' (คอลัมน์ตามที่ผู้ใช้เลือก)."""
        if generated_at is None:
            generated_at = datetime.now(THAI_TZ)

        active_count = sum(1 for r in rows if r.get("_status") == "active")
        pending_count = sum(1 for r in rows if r.get("_status") == "pending")
        inactive_count = max(0, len(rows) - active_count - pending_count)
        avg_completion = int(sum(r["_completion_percent"] for r in rows) / len(rows))
        full_count = sum(1 for r in rows if r["_completion_percent"] == 100)

        HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")   # น้ำเงินเข้ม (เหมือน finance)
        TOTAL_FILL = PatternFill("solid", fgColor="D1D5DB")    # เทาอ่อน
        SECTION_FILL = PatternFill("solid", fgColor="EFF6FF")  # ฟ้าอ่อน
        white_bold = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="0F172A")

        wb = Workbook()

        # ---- Sheet 1: สรุป ----
        ws_summary = wb.active
        ws_summary.title = "สรุป"
        ws_summary.sheet_view.showGridLines = False
        ws_summary.column_dimensions["A"].width = 34
        ws_summary.column_dimensions["B"].width = 26

        ws_summary["A1"] = f"สรุปข้อมูลนักเรียน — {room_name}"
        ws_summary["A1"].font = title_font
        ws_summary["A2"] = f"สร้างเมื่อ {generated_at.strftime('%d/%m/%Y %H:%M')} น. (เวลาไทย)"
        ws_summary["A2"].font = Font(color="64748B", size=10)

        def _summary_section(row: int, label: str):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
            for col in (1, 2):
                ws_summary.cell(row=row, column=col).fill = SECTION_FILL

        def _summary_row(row: int, label: str, value: Any):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=value)

        _summary_section(4, "ข้อมูลพื้นฐาน")
        _summary_row(5, "ชื่อห้องเรียน", room_name)
        _summary_row(6, "รหัสห้อง", room_code or "—")
        ws_summary.cell(row=7, column=1, value="จำนวนนักเรียน").font = Font(bold=True)
        count_cell = ws_summary.cell(row=7, column=2, value=len(rows))
        count_cell.font = Font(bold=True, size=12)
        count_cell.fill = TOTAL_FILL
        _summary_row(8, "กำลังเรียน (Active)", active_count)
        _summary_row(9, "รออนุมัติ (Pending)", pending_count)
        _summary_row(10, "พ้นสภาพ (Inactive)", inactive_count)

        _summary_section(12, "ความครบถ้วนของข้อมูล")
        _summary_row(13, "ข้อมูลครบถ้วนเฉลี่ย", avg_completion)
        ws_summary.cell(row=13, column=2).number_format = '0"%"'
        _summary_row(14, "มีข้อมูลครบ 100%", full_count)

        # ---- Sheet 2: รายชื่อ ----
        ws_data = wb.create_sheet("รายชื่อ")
        ws_data.sheet_view.showGridLines = False

        ws_data.append([EXPORT_HEADER_LABELS.get(f, f) for f in fields])
        for idx, field in enumerate(fields, start=1):
            ws_data.column_dimensions[get_column_letter(idx)].width = EXPORT_COLUMN_WIDTHS.get(field, DEFAULT_COL_WIDTH)
            cell = ws_data.cell(row=1, column=idx)
            cell.fill = HEADER_FILL
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, row in enumerate(rows, start=2):
            ws_data.append([row.get(f, "") for f in fields])
            for col_idx, field in enumerate(fields, start=1):
                cell = ws_data.cell(row=i, column=col_idx)
                if field in CENTER_FIELDS:
                    cell.alignment = Alignment(horizontal="center")
                elif field in WRAP_TEXT_FIELDS:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F8FAFC")
        ws_data.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    async def export_students_excel(cls, pool, fields: List[str], user_name: str, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "EXPORT_STUDENTS")

                    if not fields:
                        raise ValidationError("กรุณาเลือกอย่างน้อย 1 คอลัมน์")

                    rows = await conn.fetch(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND s.deleted_at IS NULL ORDER BY s.student_no ASC", target_room_id)
                    if not rows: raise StudentNotFoundError("ไม่พบข้อมูลนักเรียนในห้องนี้")

                    room = await conn.fetchrow("SELECT room_name, room_code FROM rooms WHERE id = $1 AND deleted_at IS NULL", target_room_id)
                    room_name = room["room_name"] if room else f"ห้อง #{target_room_id}"
                    room_code = room["room_code"] if room else None

                    # ป้องกันคอลัมน์ซ้ำ (ถ้า frontend ส่งซ้ำ) รักษาลำดับแรกที่เจอ
                    seen = set()
                    fields = [f for f in fields if not (f in seen or seen.add(f))]

                    data = []
                    for r in rows:
                        row_dict = dict(r)
                        processed = {f: cls._translate_value(f, row_dict.get(f)) for f in fields}
                        # คีย์ภายในสำหรับ Sheet สรุป (ไม่ถูกเขียนลง Sheet รายชื่อ)
                        processed["_status"] = row_dict.get("status")
                        processed["_completion_percent"] = cls._calculate_completion(row_dict)["percentage"]
                        data.append(processed)

                    output = cls._build_student_workbook(
                        room_name=room_name,
                        room_code=room_code,
                        fields=fields,
                        rows=data,
                        generated_at=datetime.now(THAI_TZ),
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="EXPORT", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="STUDENT_LIST", status="success", new_values={"fields": fields},
                        endpoint_or_command="export_students_excel", execution_time_ms=exec_time
                    )

                    return output
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="EXPORT", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT_LIST", status="failed", error_detail=str(e),
                    new_values={"fields": fields}, endpoint_or_command="export_students_excel", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def search_students(cls, pool, query: str, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันการค้นหาข้ามห้อง / คนนอก)
                if user_id is not None:
                    await require_member(conn, target_room_id, user_id)
                sql_query = f"""
                    {cls.BASE_STUDENT_SELECT}
                    WHERE s.room_id = $1
                    AND (
                        u.first_name ILIKE $2 OR
                        u.last_name ILIKE $2 OR
                        u.first_name_en ILIKE $2 OR
                        u.last_name_en ILIKE $2 OR
                        u.nickname ILIKE $2 OR
                        u.nickname_en ILIKE $2 OR
                        CAST(s.student_no AS TEXT) = $3
                    )
                    AND s.status = 'active'
                    AND s.deleted_at IS NULL
                    AND u.deleted_at IS NULL
                    LIMIT 5
                """
                search_pattern = f"%{query}%"
                rows = await conn.fetch(sql_query, target_room_id, search_pattern, query)
                
                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, entity_type="STUDENT_SEARCH",
                    status="success", new_values={"query": query}, endpoint_or_command="search_students", execution_time_ms=exec_time
                )
                return [dict(r) for r in rows]
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, entity_type="STUDENT_SEARCH",
                    status="failed", error_detail=str(e), new_values={"query": query},
                    endpoint_or_command="search_students", execution_time_ms=exec_time
                )
            raise e
    
    @classmethod
    async def get_student_profile(cls, pool: asyncpg.Pool, student_no: int, requester_user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                target_row = await conn.fetchrow(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND s.student_no = $2 AND s.deleted_at IS NULL", target_room_id, student_no)
                if not target_row: raise StudentNotFoundError("ไม่พบข้อมูลนักเรียน")

                target_data = dict(target_row)
                target_user_id = target_data.get('user_id')
                target_data['permissions'] = cls._parse_permissions(target_data['permissions'])

                is_super_admin = settings.SUPER_ADMIN_ID and int(requester_user_id) == int(settings.SUPER_ADMIN_ID)

                is_self = (target_user_id == requester_user_id)
                has_permission = False

                if not is_super_admin:
                    # 🛡️ ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันการอ่านโปรไฟล์ข้ามห้อง)
                    await require_member(conn, target_room_id, requester_user_id)
                    try:
                        await require_permission(conn, target_room_id, requester_user_id, "VIEW_ALL_STUDENTS")
                        has_permission = True
                    except ForbiddenError:
                        has_permission = False

                if not (is_super_admin or is_self or has_permission):
                    private_fields = [
                        'phone_number', 'phone_number_parent', 'phone_number_parent_relation',
                        'email', 'line_id', 'ig_username', 'birthday',
                        'address_house_no', 'address_road', 'address_sub_district',
                        'address_district', 'address_province', 'address_post_code',
                        'blood_group', 'shirt_size', 'food_allergy', 'congenital_disease',
                    ]
                    for field in private_fields:
                        if field in target_data: target_data[field] = "🔒 ไม่มีสิทธิ์เข้าถึง"

                target_data['data_completion'] = cls._calculate_completion(dict(target_row))

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=requester_user_id,
                    entity_type="STUDENT", entity_id=str(student_no), status="success",
                    endpoint_or_command="get_student_profile", execution_time_ms=exec_time
                )
                return target_data
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=requester_user_id,
                    entity_type="STUDENT", entity_id=str(student_no), status="failed", error_detail=str(e),
                    endpoint_or_command="get_student_profile", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def get_user_rooms(cls, pool, user_id: int, client_source: str, actor_identifier: str):
        start_time = time.time()
        try:
            async with pool.acquire() as conn:
                query = """
                    SELECT 
                        r.id as room_id, r.server_id, r.room_code, r.room_name, 
                        s.class_role as role, s.status, s.is_admin, s.permissions
                    FROM students s
                    JOIN rooms r ON s.room_id = r.id
                    WHERE s.user_id = $1 AND s.deleted_at IS NULL AND r.deleted_at IS NULL
                """
                rows = await conn.fetch(query, user_id)
                res = []
                for row in rows:
                    d = dict(row)
                    d['permissions'] = cls._parse_permissions(d['permissions'])
                    res.append(d)
                
                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, user_id=user_id, entity_type="ROOM_LIST",
                    status="success", endpoint_or_command="get_user_rooms", execution_time_ms=exec_time
                )
                return res
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, user_id=user_id, entity_type="ROOM_LIST",
                    status="failed", error_detail=str(e), endpoint_or_command="get_user_rooms", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def update_status(cls, pool, student_no: int, status: str, user_name: str, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        old_values = None
        new_values = {"status": status}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    # 🛡️ RBAC: มีแค่ผู้ดูแล (MANAGE_STUDENTS) ถึงจะเปลี่ยนสถานะสมาชิกได้
                    if user_id is not None:
                        await require_permission(conn, target_room_id, user_id, "MANAGE_STUDENTS")

                    old_row = await conn.fetchrow(
                        "SELECT user_id, status, is_admin FROM students WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL",
                        target_room_id, student_no
                    )
                    if old_row: old_values = dict(old_row)

                    # 🛡️ กันการปลดตัวเอง / ปลด admin อีกคน / ปลด owner (เลข 0)
                    if old_row and user_id is not None:
                        is_super_admin = settings.SUPER_ADMIN_ID and int(user_id) == int(settings.SUPER_ADMIN_ID)
                        if not is_super_admin:
                            if int(old_row['user_id']) == int(user_id):
                                raise ForbiddenError("ไม่สามารถเปลี่ยนสถานะของตนเองได้")
                            if old_row['is_admin'] or student_no == 0:
                                raise ForbiddenError("ไม่สามารถเปลี่ยนสถานะผู้ดูแลห้องได้")

                    res = await conn.execute("UPDATE students SET status = $1 WHERE room_id = $2 AND student_no = $3 AND deleted_at IS NULL", status, target_room_id, student_no)
                    if res == "UPDATE 0": raise StudentNotFoundError("ไม่พบเลขที่นี้")

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, entity_type="STUDENT",
                        entity_id=str(student_no), status="success", old_values=old_values,
                        new_values=new_values, endpoint_or_command="update_status", execution_time_ms=exec_time
                    )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, entity_type="STUDENT",
                    entity_id=str(student_no), status="failed", error_detail=str(e),
                    old_values=old_values, new_values=new_values, endpoint_or_command="update_status", execution_time_ms=exec_time
                )
            raise e
    
    @classmethod
    async def delete_student(cls, pool: asyncpg.Pool, student_no: int, user_name: str, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        old_values = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_STUDENTS")

                    old_row = await conn.fetchrow(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND s.student_no = $2 AND s.deleted_at IS NULL", target_room_id, student_no)
                    if old_row: old_values = dict(old_row)

                    # 🛡️ กันการลบตัวเอง / ลบ admin อีกคน / ลบ owner (เลข 0) — ห้องต้องเหลือคนดูแลเสมอ
                    if old_row:
                        is_super_admin = settings.SUPER_ADMIN_ID and int(user_id) == int(settings.SUPER_ADMIN_ID)
                        if not is_super_admin:
                            if int(old_row['user_id']) == int(user_id):
                                raise ForbiddenError("ไม่สามารถลบตนเองออกจากห้องได้")
                            if old_row['is_admin'] or student_no == 0:
                                raise ForbiddenError("ไม่สามารถลบผู้ดูแลห้องได้")

                    res = await conn.execute("UPDATE students SET deleted_at = NOW() WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL", target_room_id, student_no)
                    if res == "UPDATE 0": raise StudentNotFoundError("ไม่พบข้อมูล หรือถูกลบไปแล้ว")

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="STUDENT", entity_id=str(student_no), status="success",
                        old_values=old_values, endpoint_or_command="delete_student", execution_time_ms=exec_time
                    )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="DELETE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT", entity_id=str(student_no), status="failed", error_detail=str(e),
                    old_values=old_values, endpoint_or_command="delete_student", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def delete_student_permanent(cls, pool: asyncpg.Pool, student_no: int, user_name: str, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None):
        start_time = time.time()
        target_room_id = None
        old_values = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    # 🛡️ กันการลบตัวเองถาวร (owner/เลข 0) — ห้องต้องเหลือคนดูแลเสมอ
                    old_row = await conn.fetchrow(f"{cls.BASE_STUDENT_SELECT} WHERE s.room_id = $1 AND s.student_no = $2", target_room_id, student_no)
                    if old_row: old_values = dict(old_row)
                    if old_row:
                        is_super_admin = settings.SUPER_ADMIN_ID and int(user_id) == int(settings.SUPER_ADMIN_ID)
                        if not is_super_admin:
                            if int(old_row['user_id']) == int(user_id):
                                raise ForbiddenError("ไม่สามารถลบตนเองออกจากห้องได้")

                    # 🛡️ RBAC: ตรวจสิทธิ์หลังเช็คว่ามีแถวจริง (กัน idempotency ทำ fail-log ซ้ำตอนลบซ้ำ)
                    await require_permission(conn, target_room_id, user_id, "HARD_DELETE_STUDENTS")

                    has_payments = await conn.fetchval(
                        "SELECT 1 FROM student_payments WHERE student_id = (SELECT id FROM students WHERE room_id = $1 AND student_no = $2) LIMIT 1",
                        target_room_id, student_no
                    )
                    if has_payments: raise ValidationError("ไม่สามารถลบข้อมูลถาวรได้ เนื่องจากมีประวัติการเงิน ให้ใช้ Soft Delete แทน")

                    res = await conn.execute("DELETE FROM students WHERE room_id = $1 AND student_no = $2", target_room_id, student_no)
                    if res == "DELETE 0": raise StudentNotFoundError("ไม่พบข้อมูลนักเรียนเลขที่นี้")
                    
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="STUDENT", entity_id=str(student_no), status="success",
                        old_values=old_values, endpoint_or_command="delete_student_permanent", execution_time_ms=exec_time
                    )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="DELETE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="STUDENT", entity_id=str(student_no), status="failed", error_detail=str(e),
                    old_values=old_values, endpoint_or_command="delete_student_permanent", execution_time_ms=exec_time
                )
            raise e

    @classmethod
    async def sync_discord_account(
        cls,
        pool: asyncpg.Pool,
        room_code: str,
        student_no: int,
        discord_id: str,
        discord_username: str,
        client_source: str,
        actor_identifier: str,
        actor_user_id: Optional[int] = None,
    ) -> None:
        """ผูก Discord ID เข้ากับ student ที่ระบุ (room_code + student_no).

        🛡️ IDOR guard: ถ้ามี actor_user_id (ผู้ยิง request จาก get_current_user) ต้องเป็น
        เจ้าของ student เองเสมอ — กันคนอื่นส่ง X-Discord-Id/room_code ของคนอื่น
        แล้วมา "จี้" ผูกบัญชี Discord ของตัวเองทับบัญชีเพื่อน (privilege escalation).
        """
        start_time = time.time()
        target_room_id = None
        new_values = {"room_code": room_code, "student_no": student_no, "discord_id": discord_id, "discord_username": discord_username}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    # find room by room_code
                    row = await conn.fetchrow(
                        "SELECT id FROM rooms WHERE room_code = $1 AND deleted_at IS NULL", room_code
                    )
                    if not row:
                        raise RoomNotFoundError("ไม่พบรหัสห้องนี้")
                    target_room_id = row["id"]

                    # find student in this room
                    student_row = await conn.fetchrow(
                        "SELECT user_id FROM students WHERE room_id = $1 AND student_no = $2 AND deleted_at IS NULL",
                        target_room_id, student_no
                    )
                    if not student_row:
                        raise StudentNotFoundError("ไม่พบเลขที่นักเรียนในห้องนี้")

                    user_id = student_row["user_id"]

                    # 🛡️ IDOR guard: ผู้ยิงต้องเป็นเจ้าของ student นี้ (หรือ Super Admin)
                    if actor_user_id is not None:
                        is_super_admin = settings.SUPER_ADMIN_ID and int(actor_user_id) == int(settings.SUPER_ADMIN_ID)
                        if not is_super_admin and int(actor_user_id) != int(user_id):
                            raise ForbiddenError("ไม่สามารถผูก Discord ให้กับเลขที่ของผู้อื่นได้")

                    # check if discord_id is already used by another user
                    existing = await conn.fetchval(
                        "SELECT id FROM users WHERE discord_id = $1 AND id != $2 AND deleted_at IS NULL",
                        discord_id, user_id
                    )
                    if existing:
                        raise ValidationError("Discord ID นี้ถูกผูกไว้กับบัญชีอื่นแล้ว")

                    # update user's discord_id and discord_username
                    await conn.execute(
                        "UPDATE users SET discord_id = $1, discord_username = $2, updated_at = NOW() WHERE id = $3",
                        discord_id, discord_username, user_id
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="SYNC_DISCORD", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="STUDENT", entity_id=str(student_no), status="success",
                        new_values=new_values, endpoint_or_command="sync_discord_account", execution_time_ms=exec_time
                    )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="SYNC_DISCORD", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=None,
                    entity_type="STUDENT", entity_id=str(student_no) if student_no else None,
                    status="failed", error_detail=str(e), new_values=new_values,
                    endpoint_or_command="sync_discord_account", execution_time_ms=exec_time
                )
            raise e
