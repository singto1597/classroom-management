import json
import time
import io
from datetime import date, datetime
from typing import Any, Dict, FrozenSet, List, Optional
from zoneinfo import ZoneInfo

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.config import settings
from core.exceptions import (
    ActivityNotFoundError,
    CheckinSheetNotFoundError,
    ForbiddenError,
    ParticipantNotFoundError,
    RoomNotFoundError,
    StudentNotFoundError,
    ValidationError,
)
from core.logger import AuditLogger
from core.rbac import require_member, require_permission
from services.action_service import ActionService

THAI_TZ = ZoneInfo("Asia/Bangkok")

service_logger = AuditLogger(service_name="ACTIVITY")

THAI_MONTH_NAMES: List[str] = [
    "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
    "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
]

# 🌟 คอลัมน์มาตรฐานใน Excel export — แผนที่ชื่อคอลัมน์ → ภาษาไทย
# ครอบคลุมทั้ง base fields, Type A (profile), และ Type B (metadata เฉพาะกิจกรรม)
EXPORT_HEADER_LABELS: Dict[str, str] = {
    "student_no": "เลขที่",
    "first_name": "ชื่อจริง",
    "last_name": "นามสกุล",
    "nickname": "ชื่อเล่น",
    "first_name_en": "ชื่อจริง (EN)",
    "last_name_en": "นามสกุล (EN)",
    "nickname_en": "ชื่อเล่น (EN)",
    "role_type": "หน้าที่ (ประเภท)",
    "role_detail": "รายละเอียดหน้าที่",
    "earned_hours": "ชั่วโมงจิตอาสา",
    "status": "สถานะเข้าร่วม",
    # Type A — จากโปรไฟล์ users
    "blood_group": "กรุ๊ปเลือด",
    "shirt_size": "ไซส์เสื้อ",
    "food_allergy": "อาหารที่แพ้",
    "congenital_disease": "โรคประจำตัว",
    "phone_number": "เบอร์โทรศัพท์นักเรียน",
    "phone_number_parent": "เบอร์โทรศัพท์ผู้ปกครอง",
    # Type B — metadata เฉพาะกิจกรรม (หมวดการเดินทาง)
    "bus_number": "หมายเลขรถบัส",
    "van_number": "หมายเลขรถตู้",
    "seat_number": "เลขที่นั่ง",
    "travel_method": "วิธีการเดินทาง",
    # Type B — ที่พักและการจัดกลุ่ม
    "room_number": "หมายเลขห้องพัก",
    "building_name": "ชื่ออาคาร/ตึกพัก",
    "group_name": "ชื่อกลุ่ม/สี/ค่าย/บ้าน",
    "team_role": "บทบาทในทีม",
    # Type B — การจัดการหน้างาน
    "consent_status": "ใบขออนุญาตผู้ปกครอง",
    "is_paid": "สถานะจ่ายเงินค่าค่าย",
    "check_in_time": "เวลาเช็คอิน",
    # 🌟 ข้อมูลเพิ่มเติมต่อคน (custom_fields) — คอลัมน์ที่ได้จาก participant.metadata.custom_fields
    "custom_fields": "ข้อมูลเพิ่มเติม",
}

# 🌟 ป้ายภาษาไทยสำหรับคีย์ metadata ของกิจกรรม (ใช้ตอน export สรุป) — กันแสดงคีย์ดิบ
# ครอบคลุมคีย์เก่า (dual-write) + คีย์ภายใน (positions/required_fields)
ACTIVITY_META_LABELS: Dict[str, str] = {
    "location_name": "สถานที่",
    "location_url": "ลิงก์แผนที่",
    "agenda": "กำหนดการ",
    "tags": "หมวดหมู่",
    "positions": "หน้าที่/ตำแหน่ง",
    "required_fields": "ข้อมูลที่เก็บต่อคน",
    "dynamic_fields": "ฟิลด์เพิ่มเติมต่อคน",
}

# 🌟 Type A — Profile Fields: ดึงจากตาราง users (READ ONLY ในบริบทกิจกรรม) — ห้ามเก็บซ้ำลง JSONB
# ตอน GET participants จะ JOIN กลับมาพร้อมเสมอ และตอน Export จะอ่านจาก record ตรง ๆ
PROFILE_FIELDS: FrozenSet[str] = frozenset({
    "blood_group", "shirt_size", "food_allergy", "congenital_disease",
    "phone_number", "phone_number_parent",
})
PROFILE_FIELD_LABELS: Dict[str, str] = {
    "blood_group": "กรุ๊ปเลือด",
    "shirt_size": "ไซส์เสื้อ",
    "food_allergy": "อาหารที่แพ้",
    "congenital_disease": "โรคประจำตัว",
    "phone_number": "เบอร์โทรศัพท์นักเรียน",
    "phone_number_parent": "เบอร์โทรศัพท์ผู้ปกครอง",
}

ROLE_TYPE_LABELS: Dict[str, str] = {
    "participant": "ผู้เข้าร่วม",
    "staff": "ทีมงาน",
    "leader": "หัวหน้ากลุ่ม",
}

PARTICIPANT_STATUS_LABELS: Dict[str, str] = {
    "confirmed": "ยืนยันแล้ว",
    "cancelled": "ยกเลิก",
    "attended": "มาแล้ว",
}


class ActivityService:

    # ================================================================
    # 🌟 JSONB Parsing Helper (ดึงจาก docs/skills.md)
    # asyncpg คืน JSONB เป็น str หรือ dict ขึ้นกับเวอร์ชัน → normalize เสมอ
    # ================================================================
    @staticmethod
    def _parse_metadata(raw: Any) -> dict:
        if not raw:
            return {}
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, str):
            try:
                parsed = json.loads(raw)
                return parsed if isinstance(parsed, dict) else {}
            except (json.JSONDecodeError, TypeError):
                return {}
        return {}

    @staticmethod
    async def _resolve_room_id(conn: asyncpg.Connection, server_id: Optional[int] = None, room_id: Optional[int] = None) -> int:
        if room_id:
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE id = $1 AND deleted_at IS NULL", room_id):
                raise RoomNotFoundError("ไม่พบห้องเรียนนี้")
            return room_id
        if server_id:
            r_id = await conn.fetchval("SELECT id FROM rooms WHERE server_id = $1 AND deleted_at IS NULL", server_id)
            if not r_id:
                raise RoomNotFoundError(f"ไม่พบห้องสำหรับ server {server_id}")
            return r_id
        raise ValueError("ต้องระบุ server_id หรือ room_id")

    @staticmethod
    def _serializable(obj: Any) -> Any:
        """แปลง object ที่ Pydantic/JSON รับไม่ได้ (Decimal, date) ให้เป็นค่าเซฟสำหรับ audit log"""
        if isinstance(obj, (date, datetime)):
            return str(obj)
        if hasattr(obj, "item"):  # numpy-ish fallback (Decimal → float ได้ผ่าน float())
            return float(obj)
        return obj

    # 🌟 Dynamic Fields — ฟิลด์ที่ผู้จัดการกิจกรรมสร้างเอง (activities.metadata.dynamic_fields)
    # โครงสร้าง: [{ key: 'df_<n>', label: str, type: input|dropdown|boolean|datetime, options?: [{value,label}] }]
    # ต่อมา participant เก็บค่าเป็น activity_participants.metadata['df_<n>']
    @staticmethod
    def _validate_dynamic_fields(dynamic_fields: Any) -> None:
        """ตรวจโครงสร้างของ dynamic_fields (ต้องเป็น list ของ def ที่ key ต่างกัน ไม่มี label ว่าง)"""
        if dynamic_fields is None:
            return
        if not isinstance(dynamic_fields, list):
            raise ValidationError("dynamic_fields ต้องเป็น array")
        allowed_types = {"input", "dropdown", "boolean", "datetime"}
        seen = set()
        for item in dynamic_fields:
            if not isinstance(item, dict):
                raise ValidationError("dynamic_fields แต่ละรายการต้องเป็น object")
            key = item.get("key")
            label = item.get("label")
            ftype = item.get("type")
            if not isinstance(key, str) or not key.startswith("df_") or not key[3:].isdigit():
                raise ValidationError("dynamic_fields key ต้องอยู่ในรูปแบบ df_<number> (เช่น df_1)")
            if key in seen:
                raise ValidationError(f"dynamic_fields key '{key}' ซ้ำกันในรายการ")
            seen.add(key)
            if not isinstance(label, str) or not label.strip():
                raise ValidationError(f"dynamic_fields '{key}' ต้องมี label (หัวข้อ) ไม่เว้นว่าง")
            if ftype not in allowed_types:
                raise ValidationError(f"dynamic_fields '{key}' type ต้องเป็นหนึ่งใน {sorted(allowed_types)}")
            if ftype == "dropdown":
                options = item.get("options")
                if not isinstance(options, list) or not options:
                    raise ValidationError(f"dynamic_fields '{key}' type=dropdown ต้องมี options")
                for opt in options:
                    if not isinstance(opt, dict) or not str(opt.get("value", "")).strip() or not str(opt.get("label", "")).strip():
                        raise ValidationError(f"dynamic_fields '{key}' option ต้องมี value และ label")

    # ================================================================
    # 📖 Read helpers
    # ================================================================
    PARTICIPANT_SELECT = """
        SELECT
            ap.id, ap.activity_id, ap.student_id, ap.role_type, ap.role_detail,
            ap.earned_hours, ap.status, ap.metadata, ap.recorded_by,
            s.student_no,
            u.first_name, u.last_name, u.nickname,
            u.first_name_en, u.last_name_en, u.nickname_en,
            -- 🌟 Type A Profile Fields (READ ONLY จาก users) — JOIN มาพร้อมเสมอ ห้ามบันทึกซ้ำลง metadata
            u.blood_group, u.shirt_size, u.food_allergy, u.congenital_disease,
            u.phone_number, u.phone_number_parent
        FROM activity_participants ap
        JOIN students s ON ap.student_id = s.id
        LEFT JOIN users u ON s.user_id = u.id
    """

    @classmethod
    async def _fetch_activity_row(cls, conn: asyncpg.Connection, room_id: int, activity_id: int) -> Optional[dict]:
        row = await conn.fetchrow(
            "SELECT id, room_id, title, description, activity_date, base_hours, status, metadata, created_by, created_at, updated_at "
            "FROM activities WHERE id = $1 AND room_id = $2 AND deleted_at IS NULL",
            activity_id, room_id,
        )
        if not row:
            return None
        data = dict(row)
        data["metadata"] = cls._parse_metadata(data["metadata"])
        data["base_hours"] = float(data["base_hours"] or 0)
        return data

    @classmethod
    async def _fetch_participants(cls, conn: asyncpg.Connection, activity_id: int) -> List[dict]:
        rows = await conn.fetch(
            f"{cls.PARTICIPANT_SELECT} WHERE ap.activity_id = $1 AND ap.deleted_at IS NULL ORDER BY s.student_no ASC",
            activity_id,
        )
        result = []
        for row in rows:
            d = dict(row)
            d["metadata"] = cls._parse_metadata(d["metadata"])
            d["earned_hours"] = float(d["earned_hours"] or 0)
            result.append(d)
        return result

    @classmethod
    def _build_activity_response(cls, activity: dict, participants: Optional[List[dict]] = None) -> dict:
        resp = dict(activity)
        if participants is None:
            participants = activity.get("_participants") or []
        resp["participant_count"] = len(participants)
        resp["participants"] = participants
        resp.pop("_participants", None)
        return resp

    # ================================================================
    # 📥 ตรวจผู้เข้าร่วมก่อน insert (ต้องเป็นสมาชิก active ของห้อง)
    # ================================================================
    @classmethod
    async def _validate_participants(cls, conn: asyncpg.Connection, room_id: int, participants: List[dict]) -> List[tuple]:
        """คืน list ของ (student_id, role_type, role_detail, earned_hours, status, metadata) ตามลำดับ participants"""
        validated = []
        seen = set()
        for p in participants:
            student_no = p["student_no"]
            if student_no in seen:
                raise ValidationError(f"เลขที่ {student_no} ถูกเลือกซ้ำในรายชื่อผู้เข้าร่วม")
            seen.add(student_no)

            student_id = await conn.fetchval(
                "SELECT id FROM students WHERE room_id = $1 AND student_no = $2 AND status = 'active' AND deleted_at IS NULL",
                room_id, student_no,
            )
            if not student_id:
                raise StudentNotFoundError(f"ไม่พบเลขที่ {student_no} ในห้องนี้ (หรือยังไม่ active)")

            # 🌟 metadata ต้องเป็น dict เสมอ (กันยัด list/str เข้า JSONB)
            metadata = p.get("metadata") or {}
            if not isinstance(metadata, dict):
                raise ValidationError(f"metadata ของเลขที่ {student_no} ต้องเป็น object")

            validated.append((
                student_id,
                p.get("role_type", "participant"),
                p.get("role_detail"),
                p.get("earned_hours", 0.0),
                p.get("status", "confirmed"),
                metadata,
            ))
        return validated

    @classmethod
    async def _reconcile_participants(
        cls,
        conn: asyncpg.Connection,
        room_id: int,
        activity_id: int,
        participants: List[dict],
        user_name: str,
        actor_identifier: str,
        client_source: str,
        start_time: float,
    ) -> None:
        """
        🎯 แทนที่ผู้เข้าร่วมทั้งชุด (ใช้ในหน้าแก้ไขกิจกรรม — "แก้ได้ทุกอย่างเหมือนตอนสร้าง")

        รับ participant dicts ที่ถูกส่งมาจาก form (student_no + role_type + role_detail +
        earned_hours + status + metadata) แล้ว reconcile กับชุดปัจจุบันภายใน transaction เดียว:
        - มีอยู่แล้ว (student_id ตรง) → UPDATE (แทนที่ metadata เต็ม เพราะ form round-trip ทั้งชุด)
        - เคยถูก soft-delete → กู้คืน (revive) ให้ `deleted_at = NULL` (กันชน partial unique index)
        - เป็นสมาชิกใหม่ → INSERT
        - คนที่ไม่ถูกส่งมาอีกต่อไป → soft delete (deleted_at = NOW())
        เขียน audit log 1 รายการต่อ mutation (CREATE/UPDATE/DELETE) เหมือน batch_update_participants
        """
        validated = await cls._validate_participants(conn, room_id, participants)

        # อ่านชุดปัจจุบัน (id + student_id) เพื่อหาเป้าหมายการ UPDATE / DELETE
        current = await conn.fetch(
            "SELECT id, student_id FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL",
            activity_id,
        )
        current_by_student: Dict[int, int] = {row["student_id"]: row["id"] for row in current}
        incoming_students: set = set()

        for (student_id, role_type, role_detail, earned_hours, p_status, p_meta) in validated:
            incoming_students.add(student_id)
            pid = current_by_student.get(student_id)

            if pid is not None:
                # 1) มีอยู่แล้ว → UPDATE (แทนที่ metadata เต็ม)
                old = await conn.fetchrow(
                    f"{cls.PARTICIPANT_SELECT} WHERE ap.id = $1 AND ap.activity_id = $2 AND ap.deleted_at IS NULL",
                    pid, activity_id,
                )
                old_meta = cls._parse_metadata(old["metadata"]) if old else {}
                await conn.execute(
                    """
                    UPDATE activity_participants
                    SET role_type = $1, role_detail = $2, earned_hours = $3, status = $4,
                        metadata = $5::jsonb, recorded_by = $6, updated_at = CURRENT_TIMESTAMP
                    WHERE id = $7 AND activity_id = $8 AND deleted_at IS NULL
                    """,
                    role_type, role_detail, earned_hours, p_status, json.dumps(p_meta), user_name,
                    pid, activity_id,
                )
                new_vals = {"student_no": None, "role_type": role_type, "role_detail": role_detail,
                            "earned_hours": float(earned_hours), "status": p_status, "metadata": p_meta}
                if old:
                    new_vals["student_no"] = old["student_no"]
                await service_logger.log(
                    conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=room_id,
                    entity_type="ACTIVITY_PARTICIPANT", entity_id=str(pid), status="success",
                    old_values={"metadata": old_meta, "role_detail": (old["role_detail"] if old else None)},
                    new_values={k: cls._serializable(v) for k, v in new_vals.items()},
                    endpoint_or_command="update_activity/reconcile", execution_time_ms=int((time.time() - start_time) * 1000),
                )
                continue

            # 2) ไม่มีในชุด active → ลองกู้คืน soft-deleted (กันชน partial unique index)
            revived = await conn.fetchrow(
                """
                UPDATE activity_participants
                SET deleted_at = NULL, role_type = $1, role_detail = $2, earned_hours = $3,
                    status = $4, metadata = $5::jsonb, recorded_by = $6, updated_at = CURRENT_TIMESTAMP
                WHERE activity_id = $7 AND student_id = $8 AND deleted_at IS NOT NULL
                RETURNING id
                """,
                role_type, role_detail, earned_hours, p_status, json.dumps(p_meta), user_name,
                activity_id, student_id,
            )
            if revived:
                pid = revived["id"]
                await service_logger.log(
                    conn=conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=room_id,
                    entity_type="ACTIVITY_PARTICIPANT", entity_id=str(pid), status="success",
                    new_values={"action": "revived_soft_deleted", "role_type": role_type,
                                "role_detail": role_detail, "metadata": p_meta},
                    endpoint_or_command="update_activity/reconcile", execution_time_ms=int((time.time() - start_time) * 1000),
                )
                continue

            # 3) เป็นสมาชิกใหม่ → INSERT
            pid = await conn.fetchval(
                """
                INSERT INTO activity_participants
                    (activity_id, student_id, role_type, role_detail, earned_hours, status, metadata, recorded_by)
                VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb, $8)
                RETURNING id
                """,
                activity_id, student_id, role_type, role_detail, earned_hours, p_status,
                json.dumps(p_meta), user_name,
            )
            await service_logger.log(
                conn=conn, action="CREATE", actor_identifier=actor_identifier,
                client_source=client_source, room_id=room_id,
                entity_type="ACTIVITY_PARTICIPANT", entity_id=str(pid), status="success",
                new_values={"role_type": role_type, "role_detail": role_detail,
                            "earned_hours": float(earned_hours), "status": p_status, "metadata": p_meta},
                endpoint_or_command="update_activity/reconcile", execution_time_ms=int((time.time() - start_time) * 1000),
            )

        # 4) ผู้เข้าร่วมที่ถูกถอดออกจากชุด → soft delete
        for student_id, pid in current_by_student.items():
            if student_id in incoming_students:
                continue
            old = await conn.fetchrow(
                f"{cls.PARTICIPANT_SELECT} WHERE ap.id = $1 AND ap.activity_id = $2 AND ap.deleted_at IS NULL",
                pid, activity_id,
            )
            await conn.execute(
                "UPDATE activity_participants SET deleted_at = NOW() WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
                pid, activity_id,
            )
            old_meta = cls._parse_metadata(old["metadata"]) if old else {}
            await service_logger.log(
                conn=conn, action="DELETE", actor_identifier=actor_identifier,
                client_source=client_source, room_id=room_id,
                entity_type="ACTIVITY_PARTICIPANT", entity_id=str(pid), status="success",
                old_values={"metadata": old_meta, "role_detail": (old["role_detail"] if old else None)},
                new_values={"deleted_at": "soft-deleted"},
                endpoint_or_command="update_activity/reconcile", execution_time_ms=int((time.time() - start_time) * 1000),
            )

    # ================================================================
    # ✏️ CRUD: Activities
    # ================================================================
    @classmethod
    async def create_activity(
        cls,
        pool: asyncpg.Pool,
        title: str,
        activity_date: date,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
        description: Optional[str] = None,
        base_hours: float = 0.0,
        status: str = "upcoming",
        metadata: Optional[dict] = None,
        participants: Optional[List[dict]] = None,
    ) -> dict:
        """
        สร้างกิจกรรม + ผู้เข้าร่วมหลายคนพร้อมกัน (executemany ภายใน transaction เดียว)
        - RBAC: ต้องมี MANAGE_ACTIVITIES
        - Audit: CREATE กิจกรรม + CREATE participant (ใน transaction เดียวกัน)
        - Notification: publish NEW_ACTIVITY หลัง commit (ถ้าห้องผูก Discord แล้ว)
        """
        start_time = time.time()
        target_room_id = None
        new_values: dict = {}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")

                    meta = metadata or {}
                    if not isinstance(meta, dict):
                        raise ValidationError("metadata ต้องเป็น object")
                    # 🌟 validate dynamic_fields (ถ้ามีใน metadata) — กัน def ผิดโครงสร้างตอนสร้าง
                    if "dynamic_fields" in meta:
                        cls._validate_dynamic_fields(meta.get("dynamic_fields"))

                    activity_id = await conn.fetchval(
                        """
                        INSERT INTO activities (room_id, title, description, activity_date, base_hours, status, metadata, created_by)
                        VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb, $8)
                        RETURNING id
                        """,
                        target_room_id, title, description, activity_date, base_hours, status,
                        json.dumps(meta), user_name,
                    )

                    new_values = {
                        "title": title,
                        "description": description,
                        "activity_date": str(activity_date),
                        "base_hours": float(base_hours),
                        "status": status,
                        "metadata": meta,
                        "created_by": user_name,
                    }

                    # 👥 แทรกผู้เข้าร่วมแบบ executemany (atomic กับ activity)
                    participant_records = participants or []
                    if participant_records:
                        validated = await cls._validate_participants(conn, target_room_id, participant_records)
                        await conn.executemany(
                            """
                            INSERT INTO activity_participants
                                (activity_id, student_id, role_type, role_detail, earned_hours, status, metadata, recorded_by)
                            VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb, $8)
                            """,
                            [
                                (
                                    activity_id, student_id, role_type, role_detail,
                                    earned_hours, p_status, json.dumps(p_meta), user_name,
                                )
                                for (student_id, role_type, role_detail, earned_hours, p_status, p_meta) in validated
                            ],
                        )
                        new_values["participants"] = [
                            {"student_no": p["student_no"], "role_type": p.get("role_type", "participant"),
                             "role_detail": p.get("role_detail"), "earned_hours": float(p.get("earned_hours", 0))}
                            for p in participant_records
                        ]

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY", entity_id=str(activity_id), status="success",
                        new_values=new_values, endpoint_or_command="create_activity", execution_time_ms=exec_time,
                    )

            # 📢 แจ้งเตือน Discord (หลัง commit transaction สำเร็จ)
            room_server_id = None
            async with pool.acquire() as conn:
                room_server_id = await conn.fetchval(
                    "SELECT server_id FROM rooms WHERE id = $1 AND deleted_at IS NULL", target_room_id
                )
            if room_server_id:
                await ActionService.notify_new_activity(
                    server_id=room_server_id,
                    title=title,
                    activity_date=activity_date,
                    base_hours=float(base_hours),
                    metadata=meta,
                    participant_count=len(participant_records),
                    user_name=user_name,
                    activity_id=activity_id,
                )

            return {"activity_id": activity_id, "status": "success"}

        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY", status="failed", error_detail=str(e),
                    new_values=new_values or {"title": title}, endpoint_or_command="create_activity", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def list_activities(
        cls,
        pool: asyncpg.Pool,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        user_id: Optional[int] = None,
        status: Optional[str] = None,
        include_participants: bool = False,
    ) -> List[dict]:
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                if user_id is not None:
                    await require_member(conn, target_room_id, user_id)

                base_sql = (
                    "SELECT id, room_id, title, description, activity_date, base_hours, status, metadata, created_by, created_at, updated_at "
                    "FROM activities WHERE room_id = $1 AND deleted_at IS NULL"
                )
                params: List[Any] = [target_room_id]
                if status:
                    base_sql += " AND status = $" + str(len(params) + 1)
                    params.append(status)
                base_sql += " ORDER BY activity_date ASC, id DESC"

                rows = await conn.fetch(base_sql, *params)
                activities = []
                for row in rows:
                    data = dict(row)
                    data["metadata"] = cls._parse_metadata(data["metadata"])
                    data["base_hours"] = float(data["base_hours"] or 0)
                    data["_participants"] = []
                    if include_participants:
                        data["_participants"] = await cls._fetch_participants(conn, data["id"])
                    activities.append(cls._build_activity_response(data))

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id,
                    entity_type="ACTIVITY_LIST", status="success",
                    new_values={"status": status}, endpoint_or_command="list_activities", execution_time_ms=exec_time,
                )
                return activities
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_LIST", status="failed", error_detail=str(e),
                    endpoint_or_command="list_activities", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def get_activity(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        user_id: Optional[int] = None,
    ) -> dict:
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                if user_id is not None:
                    await require_member(conn, target_room_id, user_id)

                activity = await cls._fetch_activity_row(conn, target_room_id, activity_id)
                if not activity:
                    raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")
                participants = await cls._fetch_participants(conn, activity_id)

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id,
                    entity_type="ACTIVITY", entity_id=str(activity_id), status="success",
                    endpoint_or_command="get_activity", execution_time_ms=exec_time,
                )
                return cls._build_activity_response(activity, participants)
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY", entity_id=str(activity_id), status="failed",
                    error_detail=str(e), endpoint_or_command="get_activity", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def update_activity(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        update_data: dict,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
        participants: Optional[List[dict]] = None,
    ) -> dict:
        """
        PATCH กิจกรรม — อัปเดตเฉพาะฟิลด์ที่ส่งมา (exclude_unset=True จาก router)
        metadata ถ้าส่งมา จะ merge กับของเดิม (deep-merge 1 ระดับ) — กันทำหายคีย์เก่า
        participants ถ้าส่งมา (ไม่ใช่ None) → reconcile ผู้เข้าร่วมทั้งชุดใน transaction เดียว
        """
        start_time = time.time()
        target_room_id = None
        old_values: dict = {}
        try:
            clean = {k: v for k, v in update_data.items() if v is not None}
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")

                    old = await cls._fetch_activity_row(conn, target_room_id, activity_id)
                    if not old:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")
                    old_values = old

                    # 🌟 merge metadata (deep 1 ระดับ): คีย์ที่ผู้ใช้ไม่ส่งยังอยู่ครบ
                    # 📌 delete-on-null: คีย์ที่ส่งค่า None มา → ลบออก (ใช้ตอน dual-write คีย์เก่า
                    # location_name/url, agenda, tags ถูกถอดออกจากข้อมูลเพิ่มเติม) — กัน ghost key ค้าง
                    if "metadata" in clean and isinstance(clean["metadata"], dict):
                        merged = dict(old["metadata"])
                        for k, v in clean["metadata"].items():
                            if v is None:
                                merged.pop(k, None)
                            else:
                                merged[k] = v
                        clean["metadata"] = merged
                        # 🌟 validate dynamic_fields หลัง merge (เฉพาะเมื่อมีและไม่ใช่ None) — ลบ (null) ไม่ต้อง validate
                        if merged.get("dynamic_fields") is not None:
                            cls._validate_dynamic_fields(merged.get("dynamic_fields"))

                    allowed = {"title", "description", "activity_date", "base_hours", "status", "metadata"}
                    fields = {k: v for k, v in clean.items() if k in allowed}
                    if not fields and participants is None:
                        raise ValidationError("ไม่มีฟิลด์ที่แก้ไขได้ถูกส่งมา")

                    keys = sorted(fields.keys())
                    set_clauses = []
                    values: List[Any] = []
                    for i, key in enumerate(keys, start=1):
                        if key == "metadata":
                            set_clauses.append(f"{key} = ${i}::jsonb")
                            values.append(json.dumps(fields[key]))
                        elif key == "base_hours":
                            set_clauses.append(f"{key} = ${i}")
                            values.append(fields[key])
                        else:
                            set_clauses.append(f"{key} = ${i}")
                            values.append(fields[key])
                    set_clauses.append("updated_at = CURRENT_TIMESTAMP")
                    values.extend([activity_id, target_room_id])
                    sql = (
                        f"UPDATE activities SET {', '.join(set_clauses)} "
                        f"WHERE id = ${len(values)-1} AND room_id = ${len(values)} AND deleted_at IS NULL"
                    )
                    res = await conn.execute(sql, *values)
                    if res == "UPDATE 0":
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    new_values = {k: cls._serializable(v) for k, v in fields.items()}

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY", entity_id=str(activity_id), status="success",
                        old_values=old_values, new_values=new_values,
                        endpoint_or_command="update_activity", execution_time_ms=exec_time,
                    )

                    # 👥 ถ้าส่ง participants มาด้วย → reconcile ทั้งชุด (เพิ่ม/แก้/ลบ/กู้คืน) ใน transaction เดียว
                    if participants is not None:
                        await cls._reconcile_participants(
                            conn=conn,
                            room_id=target_room_id,
                            activity_id=activity_id,
                            participants=participants,
                            user_name=user_name,
                            actor_identifier=actor_identifier,
                            client_source=client_source,
                            start_time=start_time,
                        )

                    activity = await cls._fetch_activity_row(conn, target_room_id, activity_id)
                    participants = await cls._fetch_participants(conn, activity_id)
                    return cls._build_activity_response(activity, participants)
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY", entity_id=str(activity_id), status="failed",
                    error_detail=str(e), old_values=old_values,
                    endpoint_or_command="update_activity", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def delete_activity(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        user_name: str,
        user_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
    ) -> dict:
        """Soft delete กิจกรรม — ผู้เข้าร่วมยังอยู่ (soft delete ด้วย) เผื่อกู้คืนได้"""
        start_time = time.time()
        target_room_id = None
        old_values: dict = {}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_ACTIVITIES")

                    old = await cls._fetch_activity_row(conn, target_room_id, activity_id)
                    if not old:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")
                    old_values = old

                    # soft-delete กิจกรรม + ผู้เข้าร่วมทั้งหมด (กันข้อมูลฟ้องว่ายังมีคน)
                    await conn.execute(
                        "UPDATE activities SET deleted_at = NOW() WHERE id = $1 AND room_id = $2 AND deleted_at IS NULL",
                        activity_id, target_room_id,
                    )
                    await conn.execute(
                        "UPDATE activity_participants SET deleted_at = NOW() WHERE activity_id = $1 AND deleted_at IS NULL",
                        activity_id,
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="ACTIVITY", entity_id=str(activity_id), status="success",
                        old_values=old_values, new_values={"deleted_at": "soft-deleted"},
                        endpoint_or_command="delete_activity", execution_time_ms=exec_time,
                    )
                    return {"activity_id": activity_id, "deleted_at": "soft-deleted"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="DELETE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id, user_id=user_id,
                    entity_type="ACTIVITY", entity_id=str(activity_id), status="failed",
                    error_detail=str(e), old_values=old_values,
                    endpoint_or_command="delete_activity", execution_time_ms=exec_time,
                )
            raise e

    # ================================================================
    # 👥 CRUD: Participants (standalone)
    # ================================================================
    @classmethod
    async def _get_activity_room(cls, conn: asyncpg.Connection, activity_id: int) -> Optional[int]:
        """คืน room_id ของกิจกรรม (เพื่อเช็คว่าอยู่ใน target room เดียวกัน)"""
        return await conn.fetchval(
            "SELECT room_id FROM activities WHERE id = $1 AND deleted_at IS NULL", activity_id
        )

    @classmethod
    async def add_participant(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        student_no: int,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
        role_type: str = "participant",
        role_detail: Optional[str] = None,
        earned_hours: float = 0.0,
        status: str = "confirmed",
        metadata: Optional[dict] = None,
    ) -> dict:
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")

                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    student_id = await conn.fetchval(
                        "SELECT id FROM students WHERE room_id = $1 AND student_no = $2 AND status = 'active' AND deleted_at IS NULL",
                        target_room_id, student_no,
                    )
                    if not student_id:
                        raise StudentNotFoundError(f"ไม่พบเลขที่ {student_no} ในห้องนี้")

                    meta = metadata or {}
                    if not isinstance(meta, dict):
                        raise ValidationError("metadata ต้องเป็น object")

                    # ถ้าเคย soft-delete ไว้ → กู้กลับมา (กันชน UNIQUE partial index)
                    revived = await conn.fetchrow(
                        """
                        UPDATE activity_participants
                        SET deleted_at = NULL, role_type = $1, role_detail = $2, earned_hours = $3,
                            status = $4, metadata = $5::jsonb, recorded_by = $6, updated_at = CURRENT_TIMESTAMP
                        WHERE activity_id = $7 AND student_id = $8 AND deleted_at IS NOT NULL
                        RETURNING id
                        """,
                        role_type, role_detail, earned_hours, status, json.dumps(meta), user_name,
                        activity_id, student_id,
                    )
                    if revived:
                        participant_id = revived["id"]
                    else:
                        participant_id = await conn.fetchval(
                            """
                            INSERT INTO activity_participants
                                (activity_id, student_id, role_type, role_detail, earned_hours, status, metadata, recorded_by)
                            VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb, $8)
                            RETURNING id
                            """,
                            activity_id, student_id, role_type, role_detail, earned_hours, status,
                            json.dumps(meta), user_name,
                        )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY_PARTICIPANT", entity_id=str(participant_id), status="success",
                        new_values={
                            "activity_id": activity_id, "student_no": student_no,
                            "role_type": role_type, "role_detail": role_detail,
                            "earned_hours": float(earned_hours), "status": status, "metadata": meta,
                        },
                        endpoint_or_command="add_participant", execution_time_ms=exec_time,
                    )
                    return {"participant_id": participant_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_PARTICIPANT", status="failed", error_detail=str(e),
                    endpoint_or_command="add_participant", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def update_participant(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        participant_id: int,
        update_data: dict,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """PATCH participant — อัปเดตเฉพาะฟิลด์ที่ส่งมา; metadata merge กับของเดิม"""
        start_time = time.time()
        target_room_id = None
        old_values: dict = {}
        try:
            clean = {k: v for k, v in update_data.items() if v is not None}
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")

                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    old = await conn.fetchrow(
                        f"{cls.PARTICIPANT_SELECT} WHERE ap.id = $1 AND ap.activity_id = $2 AND ap.deleted_at IS NULL",
                        participant_id, activity_id,
                    )
                    if not old:
                        raise ParticipantNotFoundError(f"ไม่พบผู้เข้าร่วม ID: {participant_id}")
                    old_values = dict(old)
                    old_values["metadata"] = cls._parse_metadata(old_values["metadata"])

                    if "metadata" in clean and isinstance(clean["metadata"], dict):
                        merged = dict(old_values["metadata"])
                        merged.update(clean["metadata"])
                        clean["metadata"] = merged

                    allowed = {"role_type", "role_detail", "earned_hours", "status", "metadata"}
                    fields = {k: v for k, v in clean.items() if k in allowed}
                    if not fields:
                        raise ValidationError("ไม่มีฟิลด์ที่แก้ไขได้ถูกส่งมา")

                    keys = sorted(fields.keys())
                    set_clauses = []
                    values: List[Any] = []
                    for i, key in enumerate(keys, start=1):
                        if key == "metadata":
                            set_clauses.append(f"{key} = ${i}::jsonb")
                            values.append(json.dumps(fields[key]))
                        else:
                            set_clauses.append(f"{key} = ${i}")
                            values.append(fields[key])
                    set_clauses.append("updated_at = CURRENT_TIMESTAMP")
                    values.extend([participant_id, activity_id])
                    sql = (
                        f"UPDATE activity_participants SET {', '.join(set_clauses)} "
                        f"WHERE id = ${len(values)-1} AND activity_id = ${len(values)} AND deleted_at IS NULL"
                    )
                    res = await conn.execute(sql, *values)
                    if res == "UPDATE 0":
                        raise ParticipantNotFoundError(f"ไม่พบผู้เข้าร่วม ID: {participant_id}")

                    new_values = {k: cls._serializable(v) for k, v in fields.items()}

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY_PARTICIPANT", entity_id=str(participant_id), status="success",
                        old_values=old_values, new_values=new_values,
                        endpoint_or_command="update_participant", execution_time_ms=exec_time,
                    )
                    return {"participant_id": participant_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_PARTICIPANT", entity_id=str(participant_id), status="failed",
                    error_detail=str(e), old_values=old_values,
                    endpoint_or_command="update_participant", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def remove_participant(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        participant_id: int,
        user_name: str,
        user_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
    ) -> dict:
        """Soft delete participant"""
        start_time = time.time()
        target_room_id = None
        old_values: dict = {}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_ACTIVITIES")

                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    old = await conn.fetchrow(
                        f"{cls.PARTICIPANT_SELECT} WHERE ap.id = $1 AND ap.activity_id = $2 AND ap.deleted_at IS NULL",
                        participant_id, activity_id,
                    )
                    if not old:
                        raise ParticipantNotFoundError(f"ไม่พบผู้เข้าร่วม ID: {participant_id}")
                    old_values = dict(old)
                    old_values["metadata"] = cls._parse_metadata(old_values["metadata"])

                    await conn.execute(
                        "UPDATE activity_participants SET deleted_at = NOW() WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
                        participant_id, activity_id,
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="ACTIVITY_PARTICIPANT", entity_id=str(participant_id), status="success",
                        old_values=old_values, new_values={"deleted_at": "soft-deleted"},
                        endpoint_or_command="remove_participant", execution_time_ms=exec_time,
                    )
                    return {"participant_id": participant_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="DELETE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id, user_id=user_id,
                    entity_type="ACTIVITY_PARTICIPANT", entity_id=str(participant_id), status="failed",
                    error_detail=str(e), old_values=old_values,
                    endpoint_or_command="remove_participant", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def update_participant_status(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        participant_id: int,
        status: str,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """เปลี่ยนสถานะผู้เข้าร่วม (confirmed/cancelled/attended) — ใช้ตอนเช็คอิน / ยกเลิก"""
        return await cls.update_participant(
            pool=pool,
            activity_id=activity_id,
            participant_id=participant_id,
            update_data={"status": status},
            user_name=user_name,
            client_source=client_source,
            actor_identifier=actor_identifier,
            server_id=server_id,
            room_id=room_id,
            actor_user_id=actor_user_id,
        )

    @classmethod
    async def batch_update_participants(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        items: List[dict],
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """
        🎯 Batch Apply (คลุมดำตั้งค่า) — อัปเดต metadata ของผู้เข้าร่วมหลายคนภายใน transaction เดียว
        - รายการไหน error (participant ไม่ใช่ของกิจกรรม / soft-delete) → rollback ทั้งก้อน (atomic)
        - metadata ถูก merge กับของเดิม (ไม่ทับคีย์ที่ไม่ได้ส่ง)
        - เขียน audit log 1 รายการต่อ participant ที่ถูกอัปเดต
        """
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")

                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    # 🧹 Dedupe participant_id (กันอัปเดตซ้ำในชุดเดียว) — ตาม lesson batch finance
                    seen = set()
                    unique_items: List[dict] = []
                    for item in items:
                        pid = int(item["participant_id"])
                        if pid in seen:
                            raise ValidationError(f"participant_id {pid} ถูกส่งซ้ำใน batch")
                        seen.add(pid)
                        meta = item.get("metadata") or {}
                        if not isinstance(meta, dict):
                            raise ValidationError(f"metadata ของ participant {pid} ต้องเป็น object")
                        # 🎖️ พก role_detail ต่อ (batch ตั้งหน้าที่) — ไม่ส่ง = None = ไม่แตะของเดิม
                        role_detail = item.get("role_detail")
                        if role_detail is not None and not isinstance(role_detail, str):
                            raise ValidationError(f"role_detail ของ participant {pid} ต้องเป็น string")
                        # 🌟 role_type / status / earned_hours — optional; ไม่ส่ง = None = ไม่แตะของเดิม
                        role_type = item.get("role_type")
                        if role_type is not None and role_type not in ("participant", "staff", "leader"):
                            raise ValidationError(f"role_type ของ participant {pid} ต้องเป็น participant/staff/leader")
                        p_status = item.get("status")
                        if p_status is not None and p_status not in ("confirmed", "cancelled", "attended"):
                            raise ValidationError(f"status ของ participant {pid} ต้องเป็น confirmed/cancelled/attended")
                        earned_hours = item.get("earned_hours")
                        if earned_hours is not None:
                            try:
                                earned_hours = float(earned_hours)
                            except (TypeError, ValueError):
                                raise ValidationError(f"earned_hours ของ participant {pid} ต้องเป็นตัวเลข")
                            if earned_hours < 0:
                                raise ValidationError(f"earned_hours ของ participant {pid} ต้องไม่ติดลบ")
                        unique_items.append({
                            "participant_id": pid,
                            "role_detail": role_detail,
                            "role_type": role_type,
                            "status": p_status,
                            "earned_hours": earned_hours,
                            "metadata": meta,
                        })

                    updated = []
                    for item in unique_items:
                        pid = item["participant_id"]
                        new_meta = item["metadata"]
                        new_role_detail = item["role_detail"]
                        new_role_type = item["role_type"]
                        new_status = item["status"]
                        new_earned_hours = item["earned_hours"]
                        old = await conn.fetchrow(
                            f"{cls.PARTICIPANT_SELECT} WHERE ap.id = $1 AND ap.activity_id = $2 AND ap.deleted_at IS NULL",
                            pid, activity_id,
                        )
                        if not old:
                            raise ParticipantNotFoundError(f"ไม่พบผู้เข้าร่วม ID: {pid}")
                        old_meta = cls._parse_metadata(old["metadata"])
                        merged = dict(old_meta)
                        merged.update(new_meta)

                        # Dynamic SET — ถ้าส่ง field ไหนมา (ไม่ใช่ None) → ตั้งค่าให้
                        # role_detail: ("" = เคลียร์หน้าที่) — pattern เดียวกับ update_participant
                        set_clauses = ["metadata = $1::jsonb"]
                        values: List[Any] = [json.dumps(merged)]
                        if new_role_detail is not None:
                            set_clauses.append(f"role_detail = ${len(values) + 1}")
                            values.append(new_role_detail)
                        if new_role_type is not None:
                            set_clauses.append(f"role_type = ${len(values) + 1}")
                            values.append(new_role_type)
                        if new_status is not None:
                            set_clauses.append(f"status = ${len(values) + 1}")
                            values.append(new_status)
                        if new_earned_hours is not None:
                            set_clauses.append(f"earned_hours = ${len(values) + 1}")
                            values.append(new_earned_hours)
                        set_clauses.append("updated_at = CURRENT_TIMESTAMP")
                        values.extend([pid, activity_id])
                        sql = (
                            f"UPDATE activity_participants SET {', '.join(set_clauses)} "
                            f"WHERE id = ${len(values)-1} AND activity_id = ${len(values)} AND deleted_at IS NULL"
                        )
                        await conn.execute(sql, *values)

                        new_vals: dict = {"metadata": merged}
                        if new_role_detail is not None:
                            new_vals["role_detail"] = new_role_detail
                        if new_role_type is not None:
                            new_vals["role_type"] = new_role_type
                        if new_status is not None:
                            new_vals["status"] = new_status
                        if new_earned_hours is not None:
                            new_vals["earned_hours"] = float(new_earned_hours)
                        updated.append({
                            "participant_id": pid,
                            "student_no": old["student_no"],
                            "role_detail": new_role_detail if new_role_detail is not None else old["role_detail"],
                            "role_type": new_role_type if new_role_type is not None else old["role_type"],
                            "status": new_status if new_status is not None else old["status"],
                            "earned_hours": float(new_earned_hours) if new_earned_hours is not None else float(old["earned_hours"] or 0),
                            "metadata": merged,
                        })

                        exec_time = int((time.time() - start_time) * 1000)
                        await service_logger.log(
                            conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                            client_source=client_source, room_id=target_room_id,
                            entity_type="ACTIVITY_PARTICIPANT", entity_id=str(pid), status="success",
                            old_values={"metadata": old_meta, "role_detail": old["role_detail"]},
                            new_values={k: cls._serializable(v) for k, v in new_vals.items()},
                            endpoint_or_command="batch_update_participants", execution_time_ms=exec_time,
                        )

                    return {"status": "success", "updated_count": len(updated), "updated": updated}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_PARTICIPANT", status="failed", error_detail=str(e),
                    endpoint_or_command="batch_update_participants", execution_time_ms=exec_time,
                )
            raise e

    # ================================================================
    # ✅ ระบบเช็คชื่อแยกแผ่น (Multiple Attendance Sheets)
    # หนึ่งแผ่น = จุดเช็คหนึ่งจุด เช่น 'เช็คขึ้นรถ' 'เช็คเข้าฐาน'
    # แยกจากสถานะ overall (activity_participants.status) — additive
    # ================================================================
    @classmethod
    async def list_checkin_sheets(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        user_id: Optional[int] = None,
    ) -> List[dict]:
        """รายการแผ่นเช็คชื่อของกิจกรรม + สรุป checked/total"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                if user_id is not None:
                    await require_member(conn, target_room_id, user_id)
                if await cls._get_activity_room(conn, activity_id) != target_room_id:
                    raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                total_count = await conn.fetchval(
                    "SELECT COUNT(*) FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL",
                    activity_id,
                ) or 0

                rows = await conn.fetch(
                    """
                    SELECT s.id, s.activity_id, s.title, s.event_date, s.created_by, s.created_at, s.updated_at,
                           COUNT(r.id) FILTER (WHERE r.is_present = TRUE AND r.deleted_at IS NULL) AS checked_count
                    FROM activity_checkin_sheets s
                    LEFT JOIN activity_checkin_records r ON r.sheet_id = s.id
                    WHERE s.activity_id = $1 AND s.deleted_at IS NULL
                    GROUP BY s.id
                    ORDER BY s.created_at ASC, s.id ASC
                    """,
                    activity_id,
                )
                sheets = []
                for row in rows:
                    d = dict(row)
                    d["checked_count"] = int(d["checked_count"] or 0)
                    d["total_count"] = int(total_count)
                    sheets.append(d)

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET_LIST", entity_id=str(activity_id), status="success",
                    endpoint_or_command="list_checkin_sheets", execution_time_ms=exec_time,
                )
                return sheets
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET_LIST", entity_id=str(activity_id), status="failed",
                    error_detail=str(e), endpoint_or_command="list_checkin_sheets", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def create_checkin_sheet(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        title: str,
        event_date: Optional[date],
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """สร้างแผ่นเช็คชื่อใหม่ เช่น 'เช็คขึ้นรถ'"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")
                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    sheet_id = await conn.fetchval(
                        """
                        INSERT INTO activity_checkin_sheets (activity_id, title, event_date, created_by)
                        VALUES ($1, $2, $3, $4)
                        RETURNING id
                        """,
                        activity_id, title, event_date, user_name,
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="success",
                        new_values={"activity_id": activity_id, "title": title,
                                    "event_date": str(event_date) if event_date else None},
                        endpoint_or_command="create_checkin_sheet", execution_time_ms=exec_time,
                    )
                    return {"sheet_id": sheet_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET", status="failed", error_detail=str(e),
                    new_values={"activity_id": activity_id, "title": title},
                    endpoint_or_command="create_checkin_sheet", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def update_checkin_sheet(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        sheet_id: int,
        update_data: dict,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """PATCH แผ่นเช็คชื่อ — event_date: null = เคลียร์วันที่"""
        start_time = time.time()
        target_room_id = None
        old_values: dict = {}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")
                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    old = await conn.fetchrow(
                        "SELECT id, activity_id, title, event_date, created_by FROM activity_checkin_sheets "
                        "WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
                        sheet_id, activity_id,
                    )
                    if not old:
                        raise CheckinSheetNotFoundError(f"ไม่พบแผ่นเช็คชื่อ ID: {sheet_id}")
                    old_values = dict(old)

                    # title: None ข้าม; event_date: key มีอยู่ (แม้ None) → เซ็ต (null = เคลียร์)
                    fields: dict = {}
                    if "title" in update_data and update_data["title"] is not None:
                        fields["title"] = update_data["title"]
                    if "event_date" in update_data:
                        fields["event_date"] = update_data["event_date"]
                    if not fields:
                        raise ValidationError("ไม่มีฟิลด์ที่แก้ไขได้ถูกส่งมา")

                    keys = sorted(fields.keys())
                    set_clauses = []
                    values: List[Any] = []
                    for i, key in enumerate(keys, start=1):
                        set_clauses.append(f"{key} = ${i}")
                        values.append(fields[key])
                    set_clauses.append("updated_at = CURRENT_TIMESTAMP")
                    values.extend([sheet_id, activity_id])
                    sql = (
                        f"UPDATE activity_checkin_sheets SET {', '.join(set_clauses)} "
                        f"WHERE id = ${len(values)-1} AND activity_id = ${len(values)} AND deleted_at IS NULL"
                    )
                    res = await conn.execute(sql, *values)
                    if res == "UPDATE 0":
                        raise CheckinSheetNotFoundError(f"ไม่พบแผ่นเช็คชื่อ ID: {sheet_id}")

                    new_values = {k: cls._serializable(v) for k, v in fields.items()}
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="success",
                        old_values=old_values, new_values=new_values,
                        endpoint_or_command="update_checkin_sheet", execution_time_ms=exec_time,
                    )
                    return {"sheet_id": sheet_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="failed",
                    error_detail=str(e), old_values=old_values,
                    endpoint_or_command="update_checkin_sheet", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def delete_checkin_sheet(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        sheet_id: int,
        user_name: str,
        user_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
    ) -> dict:
        """Soft delete แผ่นเช็คชื่อ + บันทึกเช็คทั้งหมดในแผ่น"""
        start_time = time.time()
        target_room_id = None
        old_values: dict = {}
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_ACTIVITIES")
                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    old = await conn.fetchrow(
                        "SELECT id, activity_id, title, event_date, created_by FROM activity_checkin_sheets "
                        "WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
                        sheet_id, activity_id,
                    )
                    if not old:
                        raise CheckinSheetNotFoundError(f"ไม่พบแผ่นเช็คชื่อ ID: {sheet_id}")
                    old_values = dict(old)

                    await conn.execute(
                        "UPDATE activity_checkin_sheets SET deleted_at = NOW() WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
                        sheet_id, activity_id,
                    )
                    await conn.execute(
                        "UPDATE activity_checkin_records SET deleted_at = NOW() WHERE sheet_id = $1 AND deleted_at IS NULL",
                        sheet_id,
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="success",
                        old_values=old_values, new_values={"deleted_at": "soft-deleted"},
                        endpoint_or_command="delete_checkin_sheet", execution_time_ms=exec_time,
                    )
                    return {"sheet_id": sheet_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="DELETE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id, user_id=user_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="failed",
                    error_detail=str(e), old_values=old_values,
                    endpoint_or_command="delete_checkin_sheet", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def get_checkin_sheet(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        sheet_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        user_id: Optional[int] = None,
    ) -> dict:
        """ดูแผ่นเช็คชื่อ + ผู้เข้าร่วมทุกคนพร้อมเครื่องหมาย (is_present/checked_at/recorded_by)"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                if user_id is not None:
                    await require_member(conn, target_room_id, user_id)
                if await cls._get_activity_room(conn, activity_id) != target_room_id:
                    raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                sheet = await conn.fetchrow(
                    "SELECT id, activity_id, title, event_date, created_by, created_at, updated_at "
                    "FROM activity_checkin_sheets WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
                    sheet_id, activity_id,
                )
                if not sheet:
                    raise CheckinSheetNotFoundError(f"ไม่พบแผ่นเช็คชื่อ ID: {sheet_id}")

                participants = await cls._fetch_participants(conn, activity_id)
                records = await conn.fetch(
                    "SELECT participant_id, is_present, checked_at, recorded_by "
                    "FROM activity_checkin_records WHERE sheet_id = $1 AND deleted_at IS NULL",
                    sheet_id,
                )
                marks = {r["participant_id"]: r for r in records}
                for p in participants:
                    mark = marks.get(p["id"])
                    p["is_present"] = bool(mark["is_present"]) if mark else False
                    p["checked_at"] = mark["checked_at"] if mark else None
                    p["recorded_by"] = mark["recorded_by"] if mark else None

                sheet_dict = dict(sheet)
                sheet_dict["checked_count"] = sum(1 for p in participants if p["is_present"])
                sheet_dict["total_count"] = len(participants)

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="success",
                    endpoint_or_command="get_checkin_sheet", execution_time_ms=exec_time,
                )
                return {"sheet": sheet_dict, "participants": participants}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_CHECKIN_SHEET", entity_id=str(sheet_id), status="failed",
                    error_detail=str(e), endpoint_or_command="get_checkin_sheet", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def _checkin_sheet_exists(cls, conn: asyncpg.Connection, activity_id: int, sheet_id: int) -> bool:
        return await conn.fetchval(
            "SELECT 1 FROM activity_checkin_sheets WHERE id = $1 AND activity_id = $2 AND deleted_at IS NULL",
            sheet_id, activity_id,
        )

    @classmethod
    async def upsert_checkin_record(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        sheet_id: int,
        participant_id: int,
        is_present: bool,
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """เช็คชื่อ/แก้การเช็คของ participant 1 คนในแผ่น (upsert — 1 active row ต่อ (sheet, participant))"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")
                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")
                    if not await cls._checkin_sheet_exists(conn, activity_id, sheet_id):
                        raise CheckinSheetNotFoundError(f"ไม่พบแผ่นเช็คชื่อ ID: {sheet_id}")

                    participant = await conn.fetchrow(
                        f"{cls.PARTICIPANT_SELECT} WHERE ap.id = $1 AND ap.activity_id = $2 AND ap.deleted_at IS NULL",
                        participant_id, activity_id,
                    )
                    if not participant:
                        raise ParticipantNotFoundError(f"ไม่พบผู้เข้าร่วม ID: {participant_id}")

                    existing = await conn.fetchval(
                        "SELECT id FROM activity_checkin_records WHERE sheet_id = $1 AND participant_id = $2 AND deleted_at IS NULL",
                        sheet_id, participant_id,
                    )
                    action = "UPDATE" if existing else "CREATE"

                    record_id = await conn.fetchval(
                        """
                        INSERT INTO activity_checkin_records (sheet_id, participant_id, is_present, checked_at, recorded_by)
                        VALUES ($1, $2, $3, CURRENT_TIMESTAMP, $4)
                        ON CONFLICT (sheet_id, participant_id) WHERE deleted_at IS NULL
                        DO UPDATE SET is_present = EXCLUDED.is_present, checked_at = CURRENT_TIMESTAMP,
                                      recorded_by = EXCLUDED.recorded_by, updated_at = CURRENT_TIMESTAMP
                        RETURNING id
                        """,
                        sheet_id, participant_id, is_present, user_name,
                    )

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action=action, actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id,
                        entity_type="ACTIVITY_CHECKIN_RECORD", entity_id=str(record_id), status="success",
                        new_values={"sheet_id": sheet_id, "participant_id": participant_id,
                                    "student_no": participant["student_no"], "is_present": is_present},
                        endpoint_or_command="upsert_checkin_record", execution_time_ms=exec_time,
                    )
                    return {"record_id": record_id, "status": "success"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPSERT", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_CHECKIN_RECORD", status="failed", error_detail=str(e),
                    new_values={"sheet_id": sheet_id, "participant_id": participant_id, "is_present": is_present},
                    endpoint_or_command="upsert_checkin_record", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def batch_update_checkin_records(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        sheet_id: int,
        records: List[dict],
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """เช็คชื่อหลายคนในแผ่นเดียวพร้อมกัน (atomic — ตัวไหน error rollback ทั้งก้อน)"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")
                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")
                    if not await cls._checkin_sheet_exists(conn, activity_id, sheet_id):
                        raise CheckinSheetNotFoundError(f"ไม่พบแผ่นเช็คชื่อ ID: {sheet_id}")

                    seen = set()
                    unique_records: List[dict] = []
                    for r in records:
                        pid = int(r["participant_id"])
                        if pid in seen:
                            raise ValidationError(f"participant_id {pid} ถูกส่งซ้ำในชุดเช็คชื่อ")
                        seen.add(pid)
                        unique_records.append({"participant_id": pid, "is_present": bool(r["is_present"])})

                    active_ids = {
                        row["id"] for row in await conn.fetch(
                            "SELECT id FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL",
                            activity_id,
                        )
                    }
                    for r in unique_records:
                        if r["participant_id"] not in active_ids:
                            raise ParticipantNotFoundError(f"ไม่พบผู้เข้าร่วม ID: {r['participant_id']} ในกิจกรรมนี้")

                    updated = 0
                    for r in unique_records:
                        await conn.fetchval(
                            """
                            INSERT INTO activity_checkin_records (sheet_id, participant_id, is_present, checked_at, recorded_by)
                            VALUES ($1, $2, $3, CURRENT_TIMESTAMP, $4)
                            ON CONFLICT (sheet_id, participant_id) WHERE deleted_at IS NULL
                            DO UPDATE SET is_present = EXCLUDED.is_present, checked_at = CURRENT_TIMESTAMP,
                                          recorded_by = EXCLUDED.recorded_by, updated_at = CURRENT_TIMESTAMP
                            RETURNING id
                            """,
                            sheet_id, r["participant_id"], r["is_present"], user_name,
                        )
                        updated += 1
                        exec_time = int((time.time() - start_time) * 1000)
                        await service_logger.log(
                            conn=conn, action="UPDATE", actor_identifier=actor_identifier,
                            client_source=client_source, room_id=target_room_id,
                            entity_type="ACTIVITY_CHECKIN_RECORD", entity_id=str(r["participant_id"]), status="success",
                            new_values={"sheet_id": sheet_id, "participant_id": r["participant_id"],
                                        "is_present": r["is_present"]},
                            endpoint_or_command="batch_update_checkin_records", execution_time_ms=exec_time,
                        )

                    return {"status": "success", "updated_count": updated}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="UPDATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_CHECKIN_RECORD", status="failed", error_detail=str(e),
                    new_values={"sheet_id": sheet_id},
                    endpoint_or_command="batch_update_checkin_records", execution_time_ms=exec_time,
                )
            raise e

    # ================================================================
    # ➕ เพิ่มนักเรียน — รายชื่อที่ยังไม่เข้าร่วม + batch add (atomic, revive-or-insert)
    # ================================================================
    @classmethod
    async def list_available_students(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        user_id: Optional[int] = None,
    ) -> List[dict]:
        """นักเรียน active ในห้องที่ยังไม่ได้เป็นผู้เข้าร่วม active ของกิจกรรมนี้
        (soft-deleted participant = re-addable → รวมด้วย เพราะ batch-add จะ revive คืน)"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                if user_id is not None:
                    await require_member(conn, target_room_id, user_id)
                if await cls._get_activity_room(conn, activity_id) != target_room_id:
                    raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                rows = await conn.fetch(
                    """
                    SELECT s.id AS student_id, s.student_no,
                           u.first_name, u.last_name, u.nickname,
                           u.first_name_en, u.last_name_en, u.nickname_en,
                           u.blood_group, u.shirt_size, u.food_allergy, u.congenital_disease,
                           u.phone_number, u.phone_number_parent
                    FROM students s
                    LEFT JOIN users u ON s.user_id = u.id
                    WHERE s.room_id = $1 AND s.status = 'active' AND s.deleted_at IS NULL
                      AND NOT EXISTS (
                          SELECT 1 FROM activity_participants ap
                          WHERE ap.activity_id = $2 AND ap.student_id = s.id AND ap.deleted_at IS NULL
                      )
                    ORDER BY s.student_no ASC
                    """,
                    target_room_id, activity_id,
                )
                result = [dict(r) for r in rows]

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id,
                    entity_type="ACTIVITY_AVAILABLE_STUDENTS", entity_id=str(activity_id), status="success",
                    new_values={"available_count": len(result)},
                    endpoint_or_command="list_available_students", execution_time_ms=exec_time,
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_AVAILABLE_STUDENTS", entity_id=str(activity_id), status="failed",
                    error_detail=str(e), endpoint_or_command="list_available_students", execution_time_ms=exec_time,
                )
            raise e

    @classmethod
    async def batch_add_participants(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        items: List[dict],
        user_name: str,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
        actor_user_id: Optional[int] = None,
    ) -> dict:
        """เพิ่มผู้เข้าร่วมหลายคนพร้อมกัน (atomic) — revive-or-insert ต่อคน (กันชน full UNIQUE)
        เขียน audit log 1 รายการต่อคนที่เพิ่ม"""
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    if actor_user_id is not None:
                        await require_permission(conn, target_room_id, actor_user_id, "MANAGE_ACTIVITIES")
                    if await cls._get_activity_room(conn, activity_id) != target_room_id:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")

                    seen = set()
                    unique_items: List[dict] = []
                    for item in items:
                        no = int(item["student_no"])
                        if no in seen:
                            raise ValidationError(f"เลขที่ {no} ถูกส่งซ้ำในชุดเพิ่มผู้เข้าร่วม")
                        seen.add(no)
                        meta = item.get("metadata") or {}
                        if not isinstance(meta, dict):
                            raise ValidationError(f"metadata ของเลขที่ {no} ต้องเป็น object")
                        unique_items.append(item)

                    added: List[dict] = []
                    for item in unique_items:
                        no = int(item["student_no"])
                        role_type = item.get("role_type", "participant")
                        role_detail = item.get("role_detail")
                        earned_hours = float(item.get("earned_hours", 0.0) or 0.0)
                        status = item.get("status", "confirmed")
                        meta = item.get("metadata") or {}

                        student_id = await conn.fetchval(
                            "SELECT id FROM students WHERE room_id = $1 AND student_no = $2 AND status = 'active' AND deleted_at IS NULL",
                            target_room_id, no,
                        )
                        if not student_id:
                            raise StudentNotFoundError(f"ไม่พบเลขที่ {no} ในห้องนี้ (หรือยังไม่ active)")

                        # revive-or-insert (pattern เดียวกับ add_participant) — กันชน UNIQUE(activity_id, student_id)
                        revived = await conn.fetchrow(
                            """
                            UPDATE activity_participants
                            SET deleted_at = NULL, role_type = $1, role_detail = $2, earned_hours = $3,
                                status = $4, metadata = $5::jsonb, recorded_by = $6, updated_at = CURRENT_TIMESTAMP
                            WHERE activity_id = $7 AND student_id = $8 AND deleted_at IS NOT NULL
                            RETURNING id
                            """,
                            role_type, role_detail, earned_hours, status, json.dumps(meta), user_name,
                            activity_id, student_id,
                        )
                        if revived:
                            participant_id = revived["id"]
                            log_action = "CREATE"
                            audit_extra = {"action": "revived_soft_deleted"}
                        else:
                            participant_id = await conn.fetchval(
                                """
                                INSERT INTO activity_participants
                                    (activity_id, student_id, role_type, role_detail, earned_hours, status, metadata, recorded_by)
                                VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb, $8)
                                RETURNING id
                                """,
                                activity_id, student_id, role_type, role_detail, earned_hours, status,
                                json.dumps(meta), user_name,
                            )
                            log_action = "CREATE"
                            audit_extra = {}

                        exec_time = int((time.time() - start_time) * 1000)
                        await service_logger.log(
                            conn=conn, action=log_action, actor_identifier=actor_identifier,
                            client_source=client_source, room_id=target_room_id,
                            entity_type="ACTIVITY_PARTICIPANT", entity_id=str(participant_id), status="success",
                            new_values={"activity_id": activity_id, "student_no": no,
                                        "role_type": role_type, "role_detail": role_detail,
                                        "earned_hours": earned_hours, "status": status, "metadata": meta,
                                        **audit_extra},
                            endpoint_or_command="batch_add_participants", execution_time_ms=exec_time,
                        )
                        added.append({"participant_id": participant_id, "student_no": no})

                    return {"status": "success", "added": added, "updated_count": len(added)}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="CREATE", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id,
                    entity_type="ACTIVITY_PARTICIPANT", status="failed", error_detail=str(e),
                    new_values={"activity_id": activity_id, "items": items},
                    endpoint_or_command="batch_add_participants", execution_time_ms=exec_time,
                )
            raise e

    # ================================================================
    # 👤 ดูสิทธิ์/กิจกรรมของนักเรียนคนเดียว (สำหรับบอท /my_roles และหน้าโปรไฟล์)
    # ================================================================
    @classmethod
    async def get_student_activity_roles(
        cls,
        pool: asyncpg.Pool,
        user_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
    ) -> List[dict]:
        """
        คืนรายการกิจกรรม + หน้าที่ของ student คนนี้ (ผูกผ่าน users.user_id → students.student_id)
        ใช้กับบอท /my_roles: ดึง role_detail, bus_number (metadata), earned_hours ของกิจกรรมที่กำลังจะมา
        """
        start_time = time.time()
        target_room_id = None
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                student_id = await conn.fetchval(
                    "SELECT id FROM students WHERE room_id = $1 AND user_id = $2 AND deleted_at IS NULL",
                    target_room_id, user_id,
                )
                if not student_id:
                    return []

                rows = await conn.fetch(
                    """
                    SELECT
                        a.id AS activity_id, a.title, a.activity_date, a.base_hours, a.status,
                        a.metadata AS activity_metadata,
                        ap.role_type, ap.role_detail, ap.earned_hours, ap.status AS participant_status,
                        ap.metadata AS participant_metadata,
                        u.blood_group, u.shirt_size, u.food_allergy, u.congenital_disease,
                        u.phone_number, u.phone_number_parent
                    FROM activity_participants ap
                    JOIN activities a ON ap.activity_id = a.id
                    JOIN students s ON ap.student_id = s.id
                    LEFT JOIN users u ON s.user_id = u.id
                    WHERE ap.student_id = $1 AND ap.deleted_at IS NULL
                      AND a.deleted_at IS NULL AND a.room_id = $2
                    ORDER BY a.activity_date ASC
                    """,
                    student_id, target_room_id,
                )
                result = []
                for row in rows:
                    d = dict(row)
                    d["activity_metadata"] = cls._parse_metadata(d["activity_metadata"])
                    d["participant_metadata"] = cls._parse_metadata(d["participant_metadata"])
                    d["base_hours"] = float(d["base_hours"] or 0)
                    d["earned_hours"] = float(d["earned_hours"] or 0)
                    result.append(d)

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=target_room_id, user_id=user_id,
                    entity_type="ACTIVITY_MY_ROLES", status="success",
                    endpoint_or_command="get_student_activity_roles", execution_time_ms=exec_time,
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="VIEW", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id, user_id=user_id,
                    entity_type="ACTIVITY_MY_ROLES", status="failed", error_detail=str(e),
                    endpoint_or_command="get_student_activity_roles", execution_time_ms=exec_time,
                )
            raise e

    # ================================================================
    # 📄 Excel Export (openpyxl) — ดึง metadata คีย์สำคัญเป็นคอลัมน์
    # ================================================================
    @staticmethod
    def _format_buddhist_date(value: Any) -> str:
        """'15 ตุลาคม 2569' (พ.ศ. = ค.ศ. + 543)"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            value = value.date()
        if not isinstance(value, date):
            return str(value)
        return f"{value.day} {THAI_MONTH_NAMES[value.month - 1]} {value.year + 543}"

    @staticmethod
    def _translate_label(field: str, value: Any) -> Any:
        if value is None:
            return ""
        if field == "role_type":
            return ROLE_TYPE_LABELS.get(str(value), value)
        if field == "status":
            return PARTICIPANT_STATUS_LABELS.get(str(value), value)
        if field == "activity_date":
            return ActivityService._format_buddhist_date(value)
        if field == "is_paid":
            # Boolean Checkbox → ไทย
            if isinstance(value, bool):
                return "✅ จ่ายแล้ว" if value else "⏳ ยังไม่จ่าย"
            if isinstance(value, str):
                return "✅ จ่ายแล้ว" if value.lower() in ("true", "1", "yes", "y") else "⏳ ยังไม่จ่าย"
            return str(value)
        return value

    @staticmethod
    def _format_custom_fields(custom_fields: Any) -> str:
        """participant.metadata.custom_fields = [{label, value}] → 'หัวข้อ: ค่า' ต่อบรรทัด
        (กัน value ที่เป็น dict/list หลุดเป็น raw)"""
        if not isinstance(custom_fields, list):
            return ""
        lines = []
        for entry in custom_fields:
            if not isinstance(entry, dict):
                continue
            label = str(entry.get("label", "")).strip()
            value = entry.get("value")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False, default=str)
            value_str = str(value).strip() if value is not None else ""
            if label and value_str:
                lines.append(f"{label}: {value_str}")
        return "\n".join(lines)

    @staticmethod
    def _format_list_value(value: Any) -> str:
        """list → คั่นด้วย ' / ' (ใช้กับ positions, tags, required_fields)"""
        if isinstance(value, (list, tuple)):
            items = [str(v).strip() for v in value if str(v).strip()]
            return " / ".join(items) if items else ""
        return str(value) if value is not None else ""

    @classmethod
    def _format_activity_meta_lines(cls, meta: dict) -> str:
        """
        สรุป 'ข้อมูลเพิ่มเติม' ของกิจกรรมแบบ readable (ไม่แสดงคีย์ดิบ)
        - custom_fields → 'หัวข้อ: ค่า' ต่อบรรทัด
        - คีย์เก่าที่รู้จัก (location_name/url, agenda, tags) + positions/required_fields → label ไทย
        - คีย์อื่น ๆ → เก็บไว้ในกลุ่ม 'อื่นๆ' (กันข้อมูลเก่าหาย แต่ไม่ dump คีย์ดิบเดี่ยว ๆ)
        """
        if not meta:
            return "—"

        lines: List[str] = []

        # 1) custom_fields (ข้อมูลเพิ่มเติมแบบ friendly) — มาก่อนเสมอ
        custom_lines = cls._format_custom_fields(meta.get("custom_fields"))
        if custom_lines:
            lines.append(custom_lines)

        # 2) คีย์เก่า/คีย์ภายในที่รู้จัก → label ไทย (กันซ้ำกับ custom_fields ที่ mapping แล้ว)
        known_keys = ["location_name", "location_url", "agenda", "tags", "positions", "required_fields", "dynamic_fields"]
        seen_labels = set()
        for key in known_keys:
            if key not in meta:
                continue
            value = meta.get(key)
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            # ถ้า key นี้ถูก dual-write จาก custom_fields แล้ว ให้ข้าม (ป้องกันซ้ำ)
            label = ACTIVITY_META_LABELS.get(key, key)
            if label in seen_labels:
                continue
            seen_labels.add(label)
            if key in ("tags", "positions", "required_fields"):
                formatted = cls._format_list_value(value)
                if key == "required_fields":
                    # แปลงเป็นชื่อไทยของฟิลด์ที่เก็บต่อคน
                    labels = []
                    for item in value if isinstance(value, (list, tuple)) else []:
                        item_str = str(item).strip()
                        if not item_str:
                            continue
                        labels.append(EXPORT_HEADER_LABELS.get(item_str, item_str))
                    formatted = " / ".join(labels)
                if formatted:
                    lines.append(f"{label}: {formatted}")
            elif key == "dynamic_fields":
                # 🌟 Dynamic Fields (ฟิลด์ที่ผู้จัดการกิจกรรมสร้างเอง) → แสดง label ไทย ไม่ใช่คีย์ df_<n>
                labels = []
                for d in value if isinstance(value, list) else []:
                    if not isinstance(d, dict):
                        continue
                    lbl = str(d.get("label", "")).strip()
                    if lbl:
                        labels.append(lbl)
                formatted = " / ".join(labels)
                if formatted:
                    lines.append(f"{label}: {formatted}")
            elif key == "location_url":
                lines.append(f"{label}: {value}")
            else:
                lines.append(f"{label}: {value}")

        # 3) คีย์อื่น ๆ ที่เหลือ (กิจกรรมเก่า) → รวมเป็น 'อื่นๆ' ไม่ dump คีย์ดิบทีละตัว
        remaining = {}
        for k, v in meta.items():
            if k == "custom_fields" or k in ACTIVITY_META_LABELS:
                continue
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            remaining[k] = v
        if remaining:
            parts = []
            for k, v in remaining.items():
                if isinstance(v, (dict, list)):
                    v_str = json.dumps(v, ensure_ascii=False, default=str)
                else:
                    v_str = str(v)
                parts.append(f"{k}: {v_str}")
            lines.append(f"อื่นๆ: {' ; '.join(parts)}")

        return "\n".join(lines) if lines else "—"

    # ตัวอ่านค่าของแต่ละคอลัมน์ — Type A อ่านจาก profile record, Type B อ่านจาก participant metadata
    # DRY: export + (อนาคต) GET ใช้ตัวนี้อ่านค่าเดียวกัน ไม่ต้องแก้สองที่
    @staticmethod
    def _field_reader(field: str) -> Any:
        if field in PROFILE_FIELDS:
            # Type A — จาก JOIN users (record), ไม่มี value → ""
            return lambda p: p.get(field)
        return lambda p: (p.get("metadata") or {}).get(field)  # Type B — จาก metadata

    @classmethod
    async def export_activity_excel(
        cls,
        pool: asyncpg.Pool,
        activity_id: int,
        metadata_keys: List[str],
        user_name: str,
        user_id: int,
        client_source: str,
        actor_identifier: str,
        server_id: Optional[int] = None,
        room_id: Optional[int] = None,
    ) -> io.BytesIO:
        """Export ผู้เข้าร่วมกิจกรรมเป็น .xlsx — Dynamic Smart Columns

        คอลัมน์มาจาก activities.metadata.required_fields (Array คีย์ที่ผู้สร้างกิจกรรมเลือกไว้ใน Field Selector)
        - คีย์ในกลุ่ม Type A (profile) → อ่านจาก record (JOIN users) โดยตรง
        - คีย์ในกลุ่ม Type B (metadata) → อ่านจาก participant.metadata
        ถ้า metadata_keys ส่งมา (backward compat) จะใช้ตัวนั้นแทน — ลดคอลัมน์ขยะว่างเปล่า
        """
        start_time = time.time()
        target_room_id = None
        # 🌟 export_keys = required_fields (จาก Field Selector) หรือ fallback metadata_keys — init ก่อน try
        # กัน audit fallback เจอ unbound variable ตอน error เกิดก่อนกำหนดค่า (ตาม lesson "AuditLogger Fallback")
        export_keys: List[str] = [str(k).strip() for k in metadata_keys if str(k).strip()]
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls._resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_ACTIVITIES")

                    activity = await cls._fetch_activity_row(conn, target_room_id, activity_id)
                    if not activity:
                        raise ActivityNotFoundError(f"ไม่พบกิจกรรม ID: {activity_id}")
                    participants = await cls._fetch_participants(conn, activity_id)
                    if not participants:
                        raise ValidationError("ยังไม่มีผู้เข้าร่วมในกิจกรรมนี้")

                    # 🌟 Dynamic: เอา required_fields จาก metadata กิจกรรม (Field Selector)
                    # ถ้าไม่มี (กิจกรรมเก่า) → ใช้ metadata_keys ที่ frontend ส่งมา (backward compat)
                    required_fields = activity.get("metadata", {}).get("required_fields") or []
                    if not isinstance(required_fields, list):
                        required_fields = []
                    if required_fields:
                        export_keys = [str(k).strip() for k in required_fields if str(k).strip()]
                    else:
                        export_keys = [str(k).strip() for k in metadata_keys if str(k).strip()]

                    # คอลัมน์พื้นฐาน (English keys) + คอลัมน์ที่ผู้สร้างเลือก (ลำดับตามที่เลือก)
                    base_fields = ["student_no", "first_name", "last_name", "role_type", "role_detail", "earned_hours", "status"]
                    seen = set()
                    fields = [f for f in base_fields if not (f in seen or seen.add(f))]
                    for key in export_keys:
                        if key not in fields:
                            fields.append(key)
                    # 🌟 เพิ่มคอลัมน์ 'ข้อมูลเพิ่มเติม' (custom_fields ต่อคน) เฉพาะเมื่อมีใครสักคนกรอกแล้ว
                    # (กันคอลัมน์ว่างเปล่าเข้าไปขยาย header — เทสเดิมไม่มี custom_fields → ไม่กระทบ)
                    has_custom_fields = any(
                        cls._format_custom_fields((p.get("metadata") or {}).get("custom_fields"))
                        for p in participants
                    )
                    if has_custom_fields and "custom_fields" not in fields:
                        fields.append("custom_fields")

                    # 🌟 คอลัมน์ Dynamic Fields — ฟิลด์ที่ผู้จัดการกิจกรรมสร้างเอง (activities.metadata.dynamic_fields)
                    # header = label ที่ผู้ใช้ตั้ง (ไม่ใช่คีย์ df_<n>); ค่าอ่านจาก participant.metadata[df_<n>]
                    # เฉพาะเมื่อ activity มี defs — activity เก่าไม่มี → ไม่เพิ่มคอลัมน์ ไม่แตกเทสเดิม
                    dynamic_labels: Dict[str, str] = {}
                    dynamic_fields_defs = activity.get("metadata", {}).get("dynamic_fields") or []
                    if isinstance(dynamic_fields_defs, list):
                        for d in dynamic_fields_defs:
                            if not isinstance(d, dict):
                                continue
                            k = str(d.get("key", "")).strip()
                            lbl = str(d.get("label", "")).strip()
                            if k and lbl:
                                dynamic_labels[k] = lbl
                        for k in dynamic_labels:
                            if k not in fields:
                                fields.append(k)

                    room = await conn.fetchrow("SELECT room_name FROM rooms WHERE id = $1", target_room_id)
                    room_name = room["room_name"] if room else f"ห้อง #{target_room_id}"

                    wb = Workbook()
                    # ---- Sheet 1: สรุป ----
                    ws_summary = wb.active
                    ws_summary.title = "สรุป"
                    ws_summary.sheet_view.showGridLines = False
                    ws_summary.column_dimensions["A"].width = 30
                    ws_summary.column_dimensions["B"].width = 26

                    HEADER_FILL = PatternFill("solid", fgColor="8B5CF6")   # ม่วง (ธีมกิจกรรม)
                    TOTAL_FILL = PatternFill("solid", fgColor="D1D5DB")
                    SECTION_FILL = PatternFill("solid", fgColor="F5F3FF")
                    white_bold = Font(bold=True, color="FFFFFF")
                    title_font = Font(bold=True, size=16, color="0F172A")

                    ws_summary["A1"] = f"สรุปกิจกรรม — {room_name}"
                    ws_summary["A1"].font = title_font
                    ws_summary["A2"] = f"สร้างเมื่อ {datetime.now(THAI_TZ).strftime('%d/%m/%Y %H:%M')} น. (เวลาไทย)"
                    ws_summary["A2"].font = Font(color="64748B", size=10)

                    def _section(row, label):
                        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
                        for col in (1, 2):
                            ws_summary.cell(row=row, column=col).fill = SECTION_FILL

                    def _row(row, label, value):
                        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
                        ws_summary.cell(row=row, column=2, value=value)

                    _section(4, "ข้อมูลกิจกรรม")
                    _row(5, "ชื่อกิจกรรม", activity["title"])
                    _row(6, "วันที่", cls._format_buddhist_date(activity["activity_date"]))
                    _row(7, "ชั่วโมงฐาน", float(activity["base_hours"] or 0))
                    _row(8, "สถานะ", PARTICIPANT_STATUS_LABELS.get(activity["status"], activity["status"]))
                    _row(9, "จำนวนผู้เข้าร่วม", len(participants))

                    _section(11, "รายละเอียด")
                    _row(12, "คำอธิบาย", activity.get("description") or "—")

                    _section(14, "ข้อมูลเพิ่มเติมของกิจกรรม")
                    # 🌟 แสดงเป็น readable label ไทย (ไม่ dump คีย์ดิบ) — custom_fields / คีย์เก่า / positions / required_fields
                    meta_lines = cls._format_activity_meta_lines(activity.get("metadata") or {})
                    ws_summary.cell(row=15, column=1, value=meta_lines)
                    ws_summary.merge_cells(start_row=15, start_column=1, end_row=15, end_column=2)

                    # ---- Sheet 2: รายชื่อผู้เข้าร่วม ----
                    ws_data = wb.create_sheet("รายชื่อผู้เข้าร่วม")
                    ws_data.sheet_view.showGridLines = False
                    # header: dynamic field ใช้ label ที่ผู้ใช้ตั้ง (dynamic_labels) — ไม่ใช่คีย์ df_<n>
                    ws_data.append([EXPORT_HEADER_LABELS.get(f, dynamic_labels.get(f, f)) for f in fields])
                    for idx, field in enumerate(fields, start=1):
                        ws_data.column_dimensions[get_column_letter(idx)].width = 20
                        cell = ws_data.cell(row=1, column=idx)
                        cell.fill = HEADER_FILL
                        cell.font = white_bold
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # แหล่งข้อมูลของแต่ละคอลัมน์ (DRY ผ่าน _field_reader):
                    # - base fields (student_no/ชื่อ/role) → อ่านจาก record + แปลง label
                    # - Type A (profile) → อ่านจาก record ตรง ๆ (JOIN users)
                    # - Type B (metadata) → อ่านจาก participant.metadata
                    BASE_READERS: Dict[str, Any] = {
                        "student_no": lambda p: p["student_no"],
                        "first_name": lambda p: p["first_name"],
                        "last_name": lambda p: p["last_name"],
                        "first_name_en": lambda p: p.get("first_name_en"),
                        "last_name_en": lambda p: p.get("last_name_en"),
                        "nickname_en": lambda p: p.get("nickname_en"),
                        "role_type": lambda p: cls._translate_label("role_type", p["role_type"]),
                        "role_detail": lambda p: p["role_detail"],
                        "earned_hours": lambda p: float(p["earned_hours"] or 0),
                        "status": lambda p: cls._translate_label("status", p["status"]),
                    }
                    for i, p in enumerate(participants, start=2):
                        final = []
                        for field in fields:
                            if field in BASE_READERS:
                                final.append(BASE_READERS[field](p))
                            elif field == "custom_fields":
                                # 🌟 ข้อมูลเพิ่มเติมต่อคน → 'หัวข้อ: ค่า' (ไม่ใช่ raw array)
                                final.append(cls._format_custom_fields((p.get("metadata") or {}).get("custom_fields")))
                            else:
                                reader = cls._field_reader(field)
                                val = reader(p)
                                if val is None or (isinstance(val, str) and not val.strip()):
                                    final.append("")
                                else:
                                    final.append(cls._translate_label(field, val))
                        ws_data.append(final)
                        if i % 2 == 0:
                            for col_idx in range(1, len(fields) + 1):
                                ws_data.cell(row=i, column=col_idx).fill = PatternFill("solid", fgColor="F5F3FF")
                    ws_data.freeze_panes = "A2"

                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="EXPORT", actor_identifier=actor_identifier,
                        client_source=client_source, room_id=target_room_id, user_id=user_id,
                        entity_type="ACTIVITY", entity_id=str(activity_id), status="success",
                        new_values={"required_fields": export_keys, "metadata_keys": metadata_keys, "columns": fields},
                        endpoint_or_command="export_activity_excel", execution_time_ms=exec_time,
                    )
                    return output
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            safe_room_id = None if isinstance(e, RoomNotFoundError) else target_room_id
            async with pool.acquire() as error_conn:
                await service_logger.log(
                    conn=error_conn, action="EXPORT", actor_identifier=actor_identifier,
                    client_source=client_source, room_id=safe_room_id, user_id=user_id,
                    entity_type="ACTIVITY", entity_id=str(activity_id), status="failed",
                    error_detail=str(e), new_values={"required_fields": export_keys, "metadata_keys": metadata_keys},
                    endpoint_or_command="export_activity_excel", execution_time_ms=exec_time,
                )
            raise e
