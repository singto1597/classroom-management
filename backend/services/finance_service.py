import asyncpg
import io
import json
import re
import time
from datetime import date, datetime, time as dtime, timedelta
from typing import List, Optional, Dict, Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.logger import AuditLogger
from core.exceptions import RoomNotFoundError, PaymentNotFoundError, TransactionNotFoundError
from core.rbac import require_permission, require_member
from services.action_service import ActionService

THAI_TZ = ZoneInfo("Asia/Bangkok")

# [ROUTER] 🗓️ จุดแบ่งยุค (Cut-off) ระหว่างระบบ Single-Entry (legacy) กับ Double-Entry
# =================================================================================
# วันที่ 2026-09-01 เป็นวันที่ยอดยกมา (Opening Balance, Phase 2.5) เข้า journal_entries
# → ข้อมูลก่อนวันนี้ อ่านจากตารางเก่า (finance_transactions) เท่านั้น
# → ข้อมูลตั้งแต่วันนี้ขึ้นไป อ่านจากระบบบัญชีคู่ (journal_entries/journal_lines)
#   เพื่อไม่ให้รายการเก่า/ยอดยกมา เบิ้ลหรือตกหล่น
CUTOFF_DATE = date(2026, 9, 1)

service_logger = AuditLogger(service_name="FINANCE")


# [ROUTER] view ขนาดเล็ก ใช้ส่ง month/year/start_date/end_date แบบสลับกันไปมา
# ระหว่าง router กับ _resolve_export_period (เดิมรับ req object เดียว)
def _legacy_id_from_journal(metadata: dict, journal_uuid: str) -> int:
    """[DOUBLE-ENTRY] สังเคราะห์ TransactionResponse.id จาก journal.

    ระบบเดิม (frontend + revert_transaction) อ้างอิงธุรกรรมด้วย finance_transactions.id
    ซึ่ง journal เก็บไว้ใน metadata ('legacy_transaction_id' หรือ 'transfer_group_id').
    - ถ้ามี → คืนค่า int นั้น (revert ได้จริง)
    - ถ้าไม่มี (เช่น opening_balance) → คืนค่าลบที่ derived จาก UUID (คอลัมน์ไม่ซ้ำกัน,
      แต่ไม่สามารถใช้ revert ได้ — ตรงกับธรรมชาติของยอดยกมาที่ไม่ใช่ธุรกรรมรายการ)
    """
    if not isinstance(metadata, dict):
        metadata = {}
    for key in ("legacy_transaction_id", "transfer_group_id"):
        val = metadata.get(key)
        if val is not None:
            try:
                return int(val)
            except (TypeError, ValueError):
                continue
    # fallback: ใช้ hash ของ UUID มาสร้าง id ลบ (กันหน้าจอ key ซ้ำ)
    try:
        return -(abs(hash(str(journal_uuid))) % (2**31 - 1) + 1)
    except Exception:
        return -1


# [ROUTER] view ขนาดเล็ก ใช้ส่ง month/year/start_date/end_date แบบสลับกันไปมา
# ระหว่าง router กับ _resolve_export_period (เดิมรับ req object เดียว)
class _ExportPeriodView:
    def __init__(
        self,
        month: Optional[int] = None,
        year: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ):
        self.month = month
        self.year = year
        self.start_date = start_date
        self.end_date = end_date

# 🎯 หมวดหมู่รายรับ/รายจ่ายค่าเริ่มต้น — seed ให้ทุกห้องทันทีที่สร้างห้อง
# (RoomManagementService.create_room นำไป INSERT ลง finance_categories)
DEFAULT_INCOME_CATEGORIES = [
    "📥 เก็บเงินห้องปกติ",
    "💸 เงินตกหล่น/เก็บได้",
    "🎉 รายได้จากกิจกรรม",
    "⚖️ ค่าปรับ",
    "♻️ เงินทอน/เงินคืน",
    "💰 เงินสนับสนุน",
    "♻️ ขายขยะขวดพลาสติก/กระดาษ",
    "📈 ดอกเบี้ยธนาคาร",
    "💖 ผู้ใหญ่ใจดี/สปอนเซอร์",
    "🛒 กำไรจากการขายของ",
    "📈 ปรับปรุงยอด (เงินเกิน)",
]

DEFAULT_EXPENSE_CATEGORIES = [
    "✏️ เครื่องเขียน/อุปกรณ์การเรียน",
    "🧹 อุปกรณ์ทำความสะอาด",
    "📄 ชีทเรียน/ถ่ายเอกสาร",
    "🔬 อุปกรณ์ทำโครงงาน",
    "🏆 กีฬาสี",
    "🙏 วันไหว้ครู",
    "🎄 กิจกรรมอื่นๆ",
    "💊 สวัสดิการเพื่อน/พยาบาล",
    "🎂 ของขวัญ/รางวัล",
    "💸 ค่าธรรมเนียม/อื่นๆ",
    "💻 เซิร์ฟเวอร์/โดเมน/ไอที",
    "⚙️ อุปกรณ์ IoT/อิเล็กทรอนิกส์",
    "🧪 สารเคมี/อุปกรณ์ทดลอง",
    "🖨️ ค่ารูปเล่ม/พอร์ตโฟลิโอ",
    "🧻 ของใช้สิ้นเปลือง",
    "💡 ซ่อมแซม/บำรุงรักษา",
    "🪴 ตกแต่งห้องเรียน",
    "⛺ ค่ายวิชาการ/ทัศนศึกษา",
    "☕ เลี้ยงรับรอง/ซัพพอร์ตครู",
    "📸 อัดรูป/ถ่ายภาพ",
    "📉 ปรับปรุงยอด (เงินขาด)",
]

# 🎯 บัญชีเงินสดค่าเริ่มต้น — seed ให้ทุกห้องทันทีที่สร้างห้อง
# (RoomManagementService.create_room นำไป INSERT ลง finance_accounts)
DEFAULT_FINANCE_ACCOUNTS = [
    "🪙 กระเป๋าเงินสด",
    "🏦 บัญชีธนาคารห้อง",
]

class FinanceService:
    @staticmethod
    def _extract_req_data(req) -> dict:
        if isinstance(req, dict):
            return req
        # 🚀 Pydantic v2 ต้องใช้ model_dump() (dict() ถูกลบใน v3) — ตรงตามกฎ CLAUDE.md
        if hasattr(req, 'model_dump') and callable(req.model_dump):
            return req.model_dump()
        if hasattr(req, 'dict') and callable(req.dict):
            return req.dict()
        try:
            return vars(req)
        except Exception:
            return {"raw_data": str(req)}

    @staticmethod
    def _finance_actor(req) -> str:
        n = getattr(req, "user_name", None)
        if n is not None and str(n).strip():
            return str(n).strip()[:200]
        return "—"

    @staticmethod
    async def _get_room_server_id(conn: asyncpg.Connection, room_id: int) -> Optional[int]:
        """ดึง server_id ของห้อง (ใช้ publish ไป Discord) — คืน None ถ้าห้องยังไม่ผูก Discord"""
        return await conn.fetchval(
            "SELECT server_id FROM rooms WHERE id = $1 AND deleted_at IS NULL",
            room_id,
        )

    # =====================================================================
    # [ROUTER] ตัวช่วยตัดสินใจว่าจะอ่านจากระบบเก่า (legacy) หรือบัญชีคู่ (v2)
    # =====================================================================
    @staticmethod
    def _period_start(
        month: Optional[int] = None, year: Optional[int] = None,
        start_date: Optional[date] = None, end_date: Optional[date] = None,
    ) -> Optional[date]:
        """จุดเริ่มต้นของช่วงเวลาที่ถูกขอ → ใช้ตัดสินใจ Route ไป legacy หรือ v2.

        - month/year: วันที่ 1 ของเดือน
        - start_date: ตัวมันเอง (หรือจุดเริ่มต้นของช่วง)
        - ไม่ระบุเลย: ใช้จุดเริ่มต้นของเดือนปัจจุบัน (ทำนองเดียวกับ legacy fallback)
        คืน None ไม่มีทางเกิดขึ้นจริง (เดือนปัจจุบันมีจุดเริ่มเสมอ) แต่ใส่เผื่อ type safety.
        """
        if month is not None and year is not None:
            return date(year, month, 1)
        if start_date is not None:
            return start_date
        # ไม่มีตัวกรองช่วงเวลา → default เป็นเดือนปัจจุบัน (ตรงกับ logic เดิมของ get_summary)
        today = datetime.now(THAI_TZ).date()
        return date(today.year, today.month, 1)

    # =====================================================================
    # [DUAL-WRITE] Helpers — Double-Entry (Strangler Fig Phase 3)
    # =====================================================================
    # หลักการ: ระหว่างการเปลี่ยนผ่าน ข้อมูลเก่า (insert ตรงจาก test หรือ seed เก่า)
    # อาจยังไม่มีแถวใน accounting_ledgers ดังนั้น helper นี้จะ "provision เอง"
    # จาก legacy row (finance_accounts / finance_categories) ถ้ายังไม่มี — แบบเดียวกับ
    # Phase 2 migration script (migrate_phase2_ledgers.py) แต่เป็น per-row อัตโนมัติ
    # เพื่อให้ระบบ Double-Entry กับ Legacy sync กันเสมอ โดยไม่พังตอน migration ยังไม่รัน
    @classmethod
    async def _resolve_asset_ledger(cls, conn: asyncpg.Connection, room_id: int, legacy_account_id: int) -> int:
        """คืน ledger_id ของบัญชีสินทรัพย์ที่ map กับ finance_accounts.
        ถ้ายังไม่มีแถว → provision เอง (รหัสบัญชี '1' || LPAD(id,4,'0') เหมือน Phase 2)
        และถ้า legacy row ไม่อยู่จริง → raise ValueError (กันข้อมูลไม่ตรงกัน)."""
        ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE legacy_account_id = $1", legacy_account_id
        )
        if ledger_id:
            return ledger_id
        acc = await conn.fetchrow(
            "SELECT account_name FROM finance_accounts WHERE id = $1 AND room_id = $2",
            legacy_account_id, room_id,
        )
        if not acc:
            raise ValueError(f"[DUAL-WRITE] ไม่พบ ledger mapping สำหรับบัญชีสินทรัพย์ legacy_account_id={legacy_account_id}")
        return await conn.fetchval(
            """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, legacy_account_id, description)
               VALUES ($1, $2, $3, 'asset', $4, 'Auto-provisioned by dual-write')
               RETURNING id""",
            room_id, f"1{legacy_account_id:04d}", acc["account_name"], legacy_account_id,
        )

    @classmethod
    async def _resolve_category_ledger(cls, conn: asyncpg.Connection, room_id: int, legacy_category_id: int, category_type: Optional[str] = None) -> int:
        """คืน ledger_id ของหมวดหมู่ revenue/expense ที่ map กับ finance_categories.
        category_type (income/expense) ใช้ตอน provision ถ้า caller ไม่รู้ (income → 'revenue' 4xxxx, expense → 5xxxx)."""
        ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE legacy_category_id = $1", legacy_category_id
        )
        if ledger_id:
            return ledger_id
        cat = await conn.fetchrow(
            "SELECT category_name, category_type FROM finance_categories WHERE id = $1 AND room_id = $2",
            legacy_category_id, room_id,
        )
        if not cat:
            raise ValueError(f"[DUAL-WRITE] ไม่พบ ledger mapping สำหรับหมวดหมู่ legacy_category_id={legacy_category_id}")
        effective_type = category_type or cat["category_type"]
        if effective_type == "income":
            account_type, code = "revenue", f"4{legacy_category_id:04d}"
        else:
            account_type, code = "expense", f"5{legacy_category_id:04d}"
        return await conn.fetchval(
            """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, legacy_category_id, description)
               VALUES ($1, $2, $3, $4, $5, 'Auto-provisioned by dual-write')
               RETURNING id""",
            room_id, code, cat["category_name"], account_type, legacy_category_id,
        )

    @classmethod
    async def _find_revenue_ledger_by_name(cls, conn: asyncpg.Connection, room_id: int, legacy_category_id: Optional[int] = None, account_name: Optional[str] = None) -> Optional[int]:
        """ค้นหา revenue ledger สำหรับเครดิตขาของ confirm_payment.
        ลำดับการค้นหา: (1) legacy_category_id ที่ mapping ตรง ๆ, (2) ชื่อบัญชี (เช่น '📥 เก็บเงินห้องปกติ'),
        (3) revenue ตัวแรกสุดของห้อง (fallback ยืดหยุ่น). คืน None ถ้าไม่มี revenue เลย (เช่น ห้องที่ seed ยังไม่ครบ)."""
        if legacy_category_id is not None:
            ledger_id = await conn.fetchval("SELECT id FROM accounting_ledgers WHERE legacy_category_id = $1", legacy_category_id)
            if ledger_id:
                return ledger_id
        if account_name:
            ledger_id = await conn.fetchval(
                "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND account_name = $2 AND account_type = 'revenue'",
                room_id, account_name,
            )
            if ledger_id:
                return ledger_id
        return await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND account_type = 'revenue' ORDER BY id LIMIT 1",
            room_id,
        )

    @classmethod
    async def _insert_journal_entry(
        cls, conn: asyncpg.Connection, room_id: int,
        *,
        reference_type: str, reference_id: Optional[str] = None,
        description: str, recorded_by: Optional[str] = None, slip_image_url: Optional[str] = None,
        metadata: Optional[dict] = None,
        lines: List[dict],  # [{"ledger_id": int, "debit": float, "credit": float}, ...]
    ) -> str:
        """[DUAL-WRITE] สร้าง journal_entries (หัวบิล) + journal_lines (เดบิต/เครดิต) ใน transaction เดียวกับ legacy.
        คืน UUID ของ journal entry (สำหรับ revert ตาม reference ภายหลัง).
        💡 เงินทุกจำนวน cast float() ก่อน (กฎ CLAUDE.md: NUMERIC ต้อง cast ก่อน arithmetic)"""
        entry_id = await conn.fetchval(
            """INSERT INTO journal_entries (room_id, reference_type, reference_id, description, slip_image_url, recorded_by, metadata)
               VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb)
               RETURNING id""",
            room_id, reference_type, reference_id, description, slip_image_url, recorded_by,
            json.dumps(metadata or {}, ensure_ascii=False),
        )
        for line in lines:
            debit = line.get("debit", 0) or 0
            credit = line.get("credit", 0) or 0
            await conn.execute(
                """INSERT INTO journal_lines (journal_entry_id, ledger_id, debit, credit, line_description)
                   VALUES ($1, $2, $3, $4, $5)""",
                entry_id, line["ledger_id"], float(debit), float(credit), line.get("line_description"),
            )
        return entry_id

    @staticmethod
    async def resolve_room_id(conn: asyncpg.Connection, server_id: Optional[int] = None, room_id: Optional[int] = None) -> int:
        if room_id:
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE id = $1 AND deleted_at IS NULL", room_id):
                raise RoomNotFoundError("ไม่พบห้องเรียนนี้")
            return room_id
        if server_id:
            r_id = await conn.fetchval("SELECT id FROM rooms WHERE server_id = $1 AND deleted_at IS NULL", server_id)
            if not r_id: 
                raise RoomNotFoundError(f"ไม่พบห้องสำหรับ server {server_id}")
            return r_id
        raise ValueError("ต้องระบุ server_id หรือ room_id")

    @classmethod
    async def get_active_students(cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> List[dict]:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                rows = await conn.fetch("""
                    SELECT S.id, S.student_no, U.first_name, U.last_name, U.nickname
                    FROM students S
                    LEFT JOIN users U ON S.user_id = U.id
                    WHERE S.room_id = $1 AND S.status = 'active' AND S.deleted_at IS NULL
                    ORDER BY S.student_no ASC
                """, target_room_id)
                result = [dict(row) for row in rows]

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="STUDENT_LIST", status="success",
                    endpoint_or_command="FinanceService.get_active_students", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="STUDENT_LIST", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_active_students", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def create_account(cls, pool: asyncpg.Pool, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    # [DUAL-WRITE] ดึง id ของแถว legacy เพื่อ map ลง accounting_ledgers
                    new_account_id = await conn.fetchval(
                        "INSERT INTO finance_accounts (room_id, account_name, balance) VALUES ($1, $2, $3) RETURNING id",
                        target_room_id, req.account_name, req.initial_balance
                    )

                    # [DUAL-WRITE] สร้าง ledger ฝั่ง Double-Entry (asset 1xxxx) ภายใน transaction เดียวกัน
                    # รหัสบัญชี '1' || LPAD(id, 4, '0') — ตรงกับ migrate_phase2_ledgers.py
                    await conn.execute(
                        """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, legacy_account_id, description)
                           VALUES ($1, $2, $3, 'asset', $4, 'Created via dual-write (create_account)')""",
                        target_room_id, f"1{new_account_id:04d}", req.account_name, new_account_id
                    )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_ACCOUNT", status="success",
                        new_values=new_values, endpoint_or_command="FinanceService.create_account", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": f"สร้างบัญชี {req.account_name} สำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_ACCOUNT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.create_account", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def get_accounts(cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> List[dict]:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                rows = await conn.fetch("SELECT id, account_name, balance FROM finance_accounts WHERE room_id = $1 ORDER BY id", target_room_id)
                result = [dict(row) for row in rows]

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="FINANCE_ACCOUNT", status="success",
                    endpoint_or_command="FinanceService.get_accounts", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="FINANCE_ACCOUNT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_accounts", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def add_transaction(cls, pool: asyncpg.Pool, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    current_balance = await conn.fetchval(
                        "SELECT balance FROM finance_accounts WHERE id = $1 AND room_id = $2 AND deleted_at IS NULL FOR UPDATE",
                        req.account_id, target_room_id
                    )
                    if current_balance is None: raise ValueError("ไม่พบบัญชีนี้ในห้องของคุณ")
                    # 💡 balance กลับมาจาก DECIMAL เป็น Decimal (มี binary noise จาก float ที่เก็บเข้า)
                    # ต้อง cast float() ทั้งสองฝั่งก่อนเปรียบเทียบ (ตาม CLAUDE.md: cast float ก่อนเสมอ)
                    if req.transaction_type == 'expense' and float(current_balance) < float(req.amount):
                        raise ValueError(f"เงินไม่พอ! ยอดคงเหลือคือ {current_balance} บาท")

                    cat = await conn.fetchrow("SELECT room_id, category_type FROM finance_categories WHERE id = $1 AND deleted_at IS NULL", req.category_id)
                    if not cat or cat['room_id'] != target_room_id: raise ValueError("หมวดหมู่นี้ไม่มีอยู่ หรือไม่ใช่ของห้องคุณ!")
                    if cat['category_type'] != req.transaction_type:
                        raise ValueError(f"ประเภทหมวดหมู่ ({cat['category_type']}) ไม่ตรงกับประเภทการบันทึก ({req.transaction_type})!")

                    # [DUAL-WRITE] ดึง id ของ legacy transaction เพื่อเก็บลง journal metadata (สำหรับ revert)
                    new_tx_id = await conn.fetchval(
                        """INSERT INTO finance_transactions
                           (room_id, account_id, category_id, amount, description, transaction_type, slip_image_url, recorded_by)
                           VALUES ($1, $2, $3, $4, $5, $6, $7, $8) RETURNING id""",
                        target_room_id, req.account_id, req.category_id, req.amount,
                        req.description, req.transaction_type, req.slip_image_url, req.user_name
                    )

                    if req.transaction_type == 'income':
                        await conn.execute("UPDATE finance_accounts SET balance = balance + $1 WHERE id = $2", req.amount, req.account_id)
                    elif req.transaction_type == 'expense':
                        await conn.execute("UPDATE finance_accounts SET balance = balance - $1 WHERE id = $2", req.amount, req.account_id)

                    # [DUAL-WRITE] เขียนฝั่ง Double-Entry (หัวบิล + 2 บรรทัด เดบิต/เครดิต) ใน transaction เดียวกัน
                    asset_ledger_id = await cls._resolve_asset_ledger(conn, target_room_id, req.account_id)
                    category_ledger_id = await cls._resolve_category_ledger(conn, target_room_id, req.category_id, req.transaction_type)
                    if req.transaction_type == 'income':
                        lines = [
                            {"ledger_id": asset_ledger_id, "debit": req.amount, "credit": 0, "line_description": f"รับเงินเข้าบัญชี: {req.account_id}"},
                            {"ledger_id": category_ledger_id, "debit": 0, "credit": req.amount, "line_description": f"รายได้: {req.description}"},
                        ]
                    else:  # expense
                        lines = [
                            {"ledger_id": category_ledger_id, "debit": req.amount, "credit": 0, "line_description": f"ค่าใช้จ่าย: {req.description}"},
                            {"ledger_id": asset_ledger_id, "debit": 0, "credit": req.amount, "line_description": f"เงินออกจากบัญชี: {req.account_id}"},
                        ]
                    await cls._insert_journal_entry(
                        conn, target_room_id,
                        reference_type="manual_transaction",
                        reference_id=str(new_tx_id),
                        description=req.description,
                        slip_image_url=req.slip_image_url,
                        recorded_by=req.user_name,
                        metadata={"legacy_transaction_id": new_tx_id},
                        lines=lines,
                    )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_TRANSACTION", status="success",
                        new_values=new_values, endpoint_or_command="FinanceService.add_transaction", execution_time_ms=exec_time
                    )
                    # 📢 แจ้งเตือน Discord: มีรายรับ/รายจ่ายใหม่ (ไม่ @everyone — แค่โชว์ความโปร่งใส)
                    room_server_id = await cls._get_room_server_id(conn, target_room_id)
            if room_server_id:
                # ⚠️ [LOW-PRIORITY] สแปมทุกครั้งที่บันทึกเงิน (รายรับ/รายจ่าย/โอน) → เดี๋ยวจะปิด/ลดการแจ้งเตือนนี้ทีหลัง (คอมเมนต์ไว้เพื่อเตือน)
                await ActionService.notify_new_finance(
                    server_id=room_server_id,
                    txn_type=req.transaction_type,
                    amount=float(req.amount),
                    description=req.description,
                    user_name=req.user_name,
                )
            return {"status": "success", "message": "บันทึกรายการสำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_TRANSACTION", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.add_transaction", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def transfer_money(cls, pool: asyncpg.Pool, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        if req.from_account_id == req.to_account_id: raise ValueError("โอนเงินเข้าบัญชีเดิมไม่ได้!")

        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    current_balance = await conn.fetchval(
                        "SELECT balance FROM finance_accounts WHERE id = $1 AND room_id = $2 AND deleted_at IS NULL FOR UPDATE",
                        req.from_account_id, target_room_id
                    )
                    if current_balance is None: raise RoomNotFoundError("ไม่พบบัญชีต้นทาง")
                    # Decimal vs float — cast float() ทั้งสองฝั่ง (ลบ binary noise) ก่อนเทียบ
                    if float(current_balance) < float(req.amount): raise ValueError("ยอดเงินในบัญชีต้นทางไม่เพียงพอ!")

                    # 🛡️ กันการโอนเงินข้ามห้อง (cross-room leak): ต้องเช็คบัญชีปลายทางด้วย
                    if not await conn.fetchval(
                        "SELECT 1 FROM finance_accounts WHERE id = $1 AND room_id = $2 AND deleted_at IS NULL",
                        req.to_account_id, target_room_id
                    ):
                        raise RoomNotFoundError("ไม่พบบัญชีปลายทาง")

                    group_id = await conn.fetchval("SELECT nextval('transfer_group_id_seq')")
                    
                    await conn.execute("UPDATE finance_accounts SET balance = balance - $1 WHERE id = $2", req.amount, req.from_account_id)
                    # [DUAL-WRITE] ดึง id ของ legacy transaction ขาออก เพื่อเก็บใน journal metadata
                    tx_from_id = await conn.fetchval(
                        """INSERT INTO finance_transactions (room_id, account_id, amount, description, transaction_type, transfer_group_id, recorded_by)
                           VALUES ($1, $2, $3, $4, 'expense', $5, $6) RETURNING id""",
                        target_room_id, req.from_account_id, req.amount, f"โอนออก: {req.description}", group_id, req.user_name
                    )

                    await conn.execute("UPDATE finance_accounts SET balance = balance + $1 WHERE id = $2", req.amount, req.to_account_id)
                    # [DUAL-WRITE] ดึง id ของ legacy transaction ขาเข้า เพื่อเก็บใน journal metadata
                    tx_to_id = await conn.fetchval(
                        """INSERT INTO finance_transactions (room_id, account_id, amount, description, transaction_type, transfer_group_id, recorded_by)
                           VALUES ($1, $2, $3, $4, 'income', $5, $6) RETURNING id""",
                        target_room_id, req.to_account_id, req.amount, f"รับโอน: {req.description}", group_id, req.user_name
                    )

                    # [DUAL-WRITE] เขียนฝั่ง Double-Entry: ย้ายเงินระหว่างบัญชีสินทรัพย์ (Dr ปลายทาง / Cr ต้นทาง)
                    # เก็บ transfer_group_id + legacy_transaction_id ทั้งสองข้างไว้ใน metadata → revert ยกเลิกได้ทั้งกลุ่ม
                    ledger_id_from = await cls._resolve_asset_ledger(conn, target_room_id, req.from_account_id)
                    ledger_id_to = await cls._resolve_asset_ledger(conn, target_room_id, req.to_account_id)
                    await cls._insert_journal_entry(
                        conn, target_room_id,
                        reference_type="transfer",
                        reference_id=str(group_id),
                        description=req.description or "Transfer",
                        recorded_by=req.user_name,
                        metadata={
                            "transfer_group_id": group_id,
                            "legacy_transaction_id": tx_from_id,
                            "legacy_transaction_ids": [tx_from_id, tx_to_id],
                        },
                        lines=[
                            {"ledger_id": ledger_id_to, "debit": req.amount, "credit": 0, "line_description": f"รับโอนเข้าบัญชี: {req.to_account_id}"},
                            {"ledger_id": ledger_id_from, "debit": 0, "credit": req.amount, "line_description": f"โอนออกจากบัญชี: {req.from_account_id}"},
                        ],
                    )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_TRANSFER", entity_id=str(group_id), status="success",
                        new_values=new_values, endpoint_or_command="FinanceService.transfer_money", execution_time_ms=exec_time
                    )
                    return {"status": "success", "message": "โอนเงินสำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_TRANSFER", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.transfer_money", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    # =====================================================================
    # [ROUTER] get_transactions — เลือกอ่านจาก legacy หรือ Double-Entry ตามช่วงเวลา
    # =====================================================================
    @classmethod
    async def get_transactions(
        cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str, limit: int = 50, offset: int = 0,
        start_date: Optional[date] = None, end_date: Optional[date] = None,
        account_id: Optional[int] = None, category_id: Optional[int] = None, transaction_type: Optional[str] = None,
        server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None
    ) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)

                # [ROUTER] ไม่ระบุช่วงเวลา = "ทั้งหมด" → ครอบทั้งสองยุค (มีข้อมูลก่อน cutoff
                # อยู่ในตารางเก่าเท่านั้น) → ต้องอ่าน legacy เสมอ
                if start_date is None and end_date is None:
                    return await cls._get_transactions_legacy(
                        conn=conn, room_id=target_room_id,
                        limit=limit, offset=offset,
                        start_date=start_date, end_date=end_date,
                        account_id=account_id, category_id=category_id,
                        transaction_type=transaction_type,
                        client_source=client_source, actor_identifier=actor_identifier,
                        start_time=start_time,
                    )

                # [ROUTER] มีช่วงเวลา → ตัดสินจากจุดเริ่มของช่วง (ถ้าให้แค่ end_date
                # จะมองว่าจุดเริ่ม = จุดเริ่มต้นประวัติ → ตกรุ่น legacy อย่างปลอดภัย)
                period_start = start_date or date.min
                use_v2 = period_start >= CUTOFF_DATE

                if use_v2:
                    # [ROUTER] ขอข้อมูลหลังวันที่ตัด → อ่านจาก journal_entries/journal_lines
                    # (ถ้าเกินหน้าออกไปก่อนยอดยกมา เช่น start ก่อน 2026-09-01 จะตกรุ่นไปใช้ legacy)
                    return await cls._get_transactions_v2(
                        conn=conn, room_id=target_room_id,
                        limit=limit, offset=offset,
                        start_date=start_date, end_date=end_date,
                        account_id=account_id, category_id=category_id,
                        transaction_type=transaction_type,
                        client_source=client_source, actor_identifier=actor_identifier,
                        start_time=start_time,
                    )
                # [ROUTER] ข้อมูลก่อนวันที่ตัด (หรือข้ามช่วง) → อ่านจากตารางเก่า
                return await cls._get_transactions_legacy(
                    conn=conn, room_id=target_room_id,
                    limit=limit, offset=offset,
                    start_date=start_date, end_date=end_date,
                    account_id=account_id, category_id=category_id,
                    transaction_type=transaction_type,
                    client_source=client_source, actor_identifier=actor_identifier,
                    start_time=start_time,
                )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="FINANCE_TRANSACTION", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_transactions", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def _get_transactions_legacy(
        cls, conn: asyncpg.Connection, *, room_id: int,
        limit: int = 50, offset: int = 0,
        start_date: Optional[date] = None, end_date: Optional[date] = None,
        account_id: Optional[int] = None, category_id: Optional[int] = None, transaction_type: Optional[str] = None,
        client_source: str = "", actor_identifier: str = "", start_time: Optional[float] = None,
    ) -> dict:
        """[ROUTER-LEGACY] Logic เดิมของ get_transactions — อ่านจาก finance_transactions (Single-Entry)."""
        where_clause = "WHERE T.room_id = $1 AND T.deleted_at IS NULL"
        params = [room_id]
        param_idx = 2

        if start_date:
            where_clause += f" AND DATE(T.created_at) >= ${param_idx}"
            params.append(start_date); param_idx += 1
        if end_date:
            where_clause += f" AND DATE(T.created_at) <= ${param_idx}"
            params.append(end_date); param_idx += 1
        if account_id:
            where_clause += f" AND T.account_id = ${param_idx}"
            params.append(account_id); param_idx += 1
        if category_id:
            where_clause += f" AND T.category_id = ${param_idx}"
            params.append(category_id); param_idx += 1
        if transaction_type:
            where_clause += f" AND T.transaction_type = ${param_idx}"
            params.append(transaction_type); param_idx += 1

        total_count = await conn.fetchval(f"SELECT COUNT(*) FROM finance_transactions T {where_clause}", *params)

        data_sql = f"""
            SELECT
                T.id, T.amount, T.description, T.transaction_type, T.created_at,
                T.slip_image_url, T.recorded_by, T.transfer_group_id,
                A.account_name, C.category_name
            FROM finance_transactions T
            LEFT JOIN finance_accounts A ON T.account_id = A.id
            LEFT JOIN finance_categories C ON T.category_id = C.id
            {where_clause}
            ORDER BY T.created_at DESC LIMIT ${param_idx} OFFSET ${param_idx+1}
        """
        data_params = params.copy()
        data_params.extend([limit, offset])

        rows = await conn.fetch(data_sql, *data_params)
        result = {"total_count": total_count, "items": [dict(row) for row in rows]}

        if start_time is not None:
            exec_time = int((time.time() - start_time) * 1000)
            await service_logger.log(
                conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                room_id=room_id, user_id=None, entity_type="FINANCE_TRANSACTION", status="success",
                endpoint_or_command="FinanceService.get_transactions", execution_time_ms=exec_time
            )
        return result

    # =====================================================================
    # [DOUBLE-ENTRY] _get_transactions_v2 — แปลง Dr/Cr จาก journal กลับเป็น
    # schema เดิมที่ frontend ใช้ (TransactionResponse) โดยไม่ให้ frontend แก้ไข
    # =====================================================================
    @classmethod
    async def _get_transactions_v2(
        cls, conn: asyncpg.Connection, *, room_id: int,
        limit: int = 50, offset: int = 0,
        start_date: Optional[date] = None, end_date: Optional[date] = None,
        account_id: Optional[int] = None, category_id: Optional[int] = None, transaction_type: Optional[str] = None,
        client_source: str = "", actor_identifier: str = "", start_time: Optional[float] = None,
    ) -> dict:
        """อ่านประวัติจากระบบบัญชีคู่ (journal_entries + journal_lines) แล้วจัดรูป
        ให้เหมือน legacy (TransactionResponse) เพื่อ frontend ไม่ต้องแก้เลย.

        หลักการจัดประเภท (ตามสเปค Phase 4):
        - income  : สินทรัพย์ เดบิต>0  + บัญชีรายได้ เครดิต>0  (รับเงินเข้า)
        - expense : สินทรัพย์ เครดิต>0 + บัญชีค่าใช้จ่าย เดบิต>0 (เงินออก)
        - transfer: สินทรัพย์ 2 บัญชี (ฝั่งหนึ่ง เดบิต>0 / อีกฝั่ง เครดิต>0)
        - opening_balance: แสดงเป็น income (เงินเข้า Asset) — ยอด = เดบิตฝั่ง Asset
        """
        where_cond, params = ["JE.room_id = $1 AND JE.deleted_at IS NULL"], [room_id]
        idx = 2
        if start_date:
            where_cond.append(f"DATE(JE.transaction_date) >= ${idx}"); params.append(start_date); idx += 1
        if end_date:
            where_cond.append(f"DATE(JE.transaction_date) <= ${idx}"); params.append(end_date); idx += 1
        if account_id:
            # [DOUBLE-ENTRY] กรองด้วยบัญชีสินทรัพย์: journal ใดก็ตามที่ asset ledger นี้มีบทบาท (Dr หรือ Cr)
            where_cond.append(f"""
                EXISTS (
                    SELECT 1 FROM journal_lines JLx
                    JOIN accounting_ledgers ALx ON JLx.ledger_id = ALx.id
                    WHERE JLx.journal_entry_id = JE.id AND ALx.room_id = $1
                      AND ALx.legacy_account_id = ${idx} AND (JLx.debit > 0 OR JLx.credit > 0)
                )""")
            params.append(account_id); idx += 1
        if category_id:
            # [DOUBLE-ENTRY] กรองด้วยหมวดหมู่ (revenue/expense): journal ที่ ledger นั้นมีบทบาท
            where_cond.append(f"""
                EXISTS (
                    SELECT 1 FROM journal_lines JLx2
                    JOIN accounting_ledgers ALx2 ON JLx2.ledger_id = ALx2.id
                    WHERE JLx2.journal_entry_id = JE.id AND ALx2.room_id = $1
                      AND ALx2.legacy_category_id = ${idx} AND (JLx2.debit > 0 OR JLx2.credit > 0)
                )""")
            params.append(category_id); idx += 1

        # [DOUBLE-ENTRY] ยกทุกบรรทัดของ journal ที่ผ่าน filter ขึ้นมา (หลายบรรทัดต่อ 1 บิล)
        # แล้วจัดประเภทที่ฝั่ง Python (อ่านง่ายกว่า SQL หลายชั้น)
        data_sql = f"""
            SELECT
                JE.id AS journal_entry_id,
                JE.reference_type, JE.reference_id,
                JE.description AS entry_description,
                JE.transaction_date, JE.recorded_by,
                JE.slip_image_url, JE.metadata,
                L.id AS line_id, L.debit, L.credit, L.line_description,
                AL.id AS ledger_id, AL.account_code, AL.account_name, AL.account_type,
                AL.legacy_account_id, AL.legacy_category_id
            FROM journal_entries JE
            JOIN journal_lines L ON L.journal_entry_id = JE.id
            JOIN accounting_ledgers AL ON L.ledger_id = AL.id
            WHERE {' AND '.join(where_cond)}
              AND JE.status <> 'voided'
            ORDER BY JE.transaction_date DESC, JE.id DESC
        """
        rows = await conn.fetch(data_sql, *params)

        # [DOUBLE-ENTRY] กลุ่มบรรทัดตามหัวบิล
        entries: Dict[str, dict] = {}
        for r in rows:
            entry_id = str(r["journal_entry_id"])
            if entry_id not in entries:
                entries[entry_id] = {
                    "journal_entry_id": entry_id,
                    "reference_type": r["reference_type"],
                    "reference_id": r["reference_id"],
                    "description": r["entry_description"],
                    "transaction_date": r["transaction_date"],
                    "recorded_by": r["recorded_by"],
                    "slip_image_url": r["slip_image_url"],
                    "metadata": r["metadata"] or {},
                    "lines": [],
                }
            entries[entry_id]["lines"].append(r)

        # [DOUBLE-ENTRY] แปลงแต่ละบิล → TransactionResponse
        # 📌 หมายเหตุเรื่อง id: TransactionResponse.id เป็น int และ frontend ใช้เรียก
        # revert_transaction (ซึ่งค้นจาก finance_transactions.id) → สังเคราะห์จาก legacy
        # transaction id ใน metadata; ถ้าไม่มี (เช่น opening_balance) ใช้ค่าลบจาก UUID
        # เพื่อให้คอลัมน์มีค่าไม่ซ้ำกัน (เป็น id ที่ "ไม่ใช้ได้จริง" กับ revert)
        items: List[dict] = []
        for entry in entries.values():
            txn = cls._classify_journal_entry(entry, transaction_type)
            if txn is None:
                continue  # ถูกกรองด้วย transaction_type ด้านบนแล้ว
            items.append(txn)

        # [DOUBLE-ENTRY] จำลอง pagination ฝั่ง application (ชุดข้อมูลนี้เล็ก —
        # ต่อบิลมีแค่ 2-3 บรรทัด) เพื่อคง API เดิม (limit/offset + total_count)
        total_count = len(items)
        paged = items[offset:offset + limit]

        if start_time is not None:
            exec_time = int((time.time() - start_time) * 1000)
            await service_logger.log(
                conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                room_id=room_id, user_id=None, entity_type="FINANCE_TRANSACTION", status="success",
                endpoint_or_command="FinanceService.get_transactions", execution_time_ms=exec_time
            )
        return {"total_count": total_count, "items": paged}

    @classmethod
    def _classify_journal_entry(cls, entry: dict, transaction_type: Optional[str] = None) -> Optional[dict]:
        """[DOUBLE-ENTRY] จัดประเภทบิลจากชุด journal_lines เป็นแถว TransactionResponse
        ที่ frontend ใช้อยู่ (amount, description, transaction_type, account_name, category_name).

        คืน None ถ้าบิลไม่ตรง transaction_type ที่กรอง (คล้าย WHERE ใน legacy).
        """
        lines = entry["lines"]
        asset_lines = [ln for ln in lines if ln["account_type"] == "asset"]
        revenue_lines = [ln for ln in lines if ln["account_type"] == "revenue"]
        expense_lines = [ln for ln in lines if ln["account_type"] == "expense"]
        # กันพวก liability/equity (เช่น ขา equity ของ opening balance) เข้ามารบกวน
        asset_dr = sum(float(ln["debit"]) for ln in asset_lines)
        asset_cr = sum(float(ln["credit"]) for ln in asset_lines)
        revenue_cr = sum(float(ln["credit"]) for ln in revenue_lines)
        expense_dr = sum(float(ln["debit"]) for ln in expense_lines)

        description = entry["description"] or ""
        recorded_by = entry["recorded_by"]
        txn_type: Optional[str] = None
        amount = 0.0
        account_name: Optional[str] = None
        category_name: Optional[str] = None
        transfer_group_id = None

        # [DOUBLE-ENTRY] 1) ยอดยกมา (Opening Balance) → เงินเข้าสินทรัพย์ = 'income'
        if entry["reference_type"] == "opening_balance":
            txn_type = "income"
            # ยอด = เดบิตฝั่ง Asset (เงินที่มีจริงในกระเป๋า) ไม่รวมขา Equity
            amount = asset_dr
            # ถ้าสินทรัพย์ติดลบ (credit asset) → ให้รวมเครดิตเข้าด้วยเพื่อไม่ให้ amount เป็น 0
            if amount == 0.0:
                amount = asset_cr
            if asset_lines:
                account_name = asset_lines[0]["account_name"]
            category_name = "ยอดยกมา (เปิดระบบบัญชีคู่)"
            description = description or "ยอดยกมา"

        # [DOUBLE-ENTRY] 2) Transfer ระหว่างบัญชีสินทรัพย์ → 'expense' (ขาออก) + transfer_group_id
        elif len(asset_lines) >= 2 and asset_dr > 0 and asset_cr > 0:
            txn_type = "expense"
            amount = asset_cr
            # บัญชีที่เงินออก (credit) — เป็นบัญชีต้นทาง
            out_line = next((ln for ln in asset_lines if float(ln["credit"]) > 0), asset_lines[0])
            account_name = out_line["account_name"]
            category_name = "โอนเงิน"
            transfer_group_id = entry["reference_id"] and int(entry["reference_id"]) or None
            recorded_by = recorded_by

        # [DOUBLE-ENTRY] 3) รายได้: Asset เดบิต + Revenue เครดิต (รับเงินเข้า)
        elif asset_dr > 0 and revenue_cr > 0:
            txn_type = "income"
            amount = asset_dr
            if asset_lines:
                account_name = asset_lines[0]["account_name"]
            if revenue_lines:
                category_name = revenue_lines[0]["account_name"]
            description = description or "รายได้"

        # [DOUBLE-ENTRY] 4) รายจ่าย: Expense เดบิต + Asset เครดิต (เงินออก)
        elif expense_dr > 0 and asset_cr > 0:
            txn_type = "expense"
            amount = asset_cr
            if asset_lines:
                account_name = asset_lines[0]["account_name"]
            if expense_lines:
                category_name = expense_lines[0]["account_name"]
            description = description or "รายจ่าย"

        # [DOUBLE-ENTRY] 5) กรณีโครงสร้างอื่น (ไม่มี asset) → พยายามเดาจากฝั่งที่มี
        else:
            if revenue_cr > 0:
                txn_type = "income"
                amount = revenue_cr
                if revenue_lines:
                    category_name = revenue_lines[0]["account_name"]
            elif expense_dr > 0:
                txn_type = "expense"
                amount = expense_dr
                if expense_lines:
                    category_name = expense_lines[0]["account_name"]
            else:
                return None

        # [DOUBLE-ENTRY] รองรับ filter transaction_type (income/expense) แบบเดียวกับ legacy
        if transaction_type and txn_type != transaction_type:
            return None

        # 💡 เก็บบรรทัดแรกของ asset ไว้เป็น account (หน้า frontend ใช้แสดงกระเป๋าเงิน)
        if not account_name and asset_lines:
            account_name = asset_lines[0]["account_name"]
        if not account_name:
            account_name = "—"

        return {
            "id": _legacy_id_from_journal(entry.get("metadata") or {}, entry.get("journal_entry_id")),
            "amount": float(amount),
            "description": description,
            "transaction_type": txn_type,
            "created_at": entry["transaction_date"],
            "slip_image_url": entry["slip_image_url"],
            "recorded_by": recorded_by,
            "account_name": account_name,
            "category_name": category_name,
            "transfer_group_id": transfer_group_id,
        }

    @classmethod
    async def create_fee_collection(cls, pool: asyncpg.Pool, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    collection_id = await conn.fetchval(
                        "INSERT INTO fee_collections (room_id, title, amount, due_date) VALUES ($1, $2, $3, $4) RETURNING id",
                        target_room_id, req.title, req.amount, req.due_date
                    )
                    
                    target_students = []
                    if req.student_ids is not None:
                        if len(req.student_ids) == 0:
                            raise ValueError("ไม่สามารถสร้างรายการได้ เนื่องจากไม่ได้เลือกนักเรียนเลยแม้แต่คนเดียว")

                        # 💡 กรอง id ซ้ำก่อน query + INSERT กัน UniqueViolationError (student_payments มี UNIQUE(collection_id, student_id))
                        unique_ids = list(dict.fromkeys(req.student_ids))
                        valid_students = await conn.fetch(
                            "SELECT id FROM students WHERE room_id = $1 AND id = ANY($2) AND status = 'active'",
                            target_room_id, unique_ids
                        )
                        target_students = [s['id'] for s in valid_students]
                    else:
                        all_students = await conn.fetch("SELECT id FROM students WHERE room_id = $1 AND status = 'active'", target_room_id)
                        target_students = [s['id'] for s in all_students]

                    if target_students:
                        # 🛡️ กัน id ซ้ำใน target_students (ป้องกัน UniqueViolation ถ้า query คืนค่าซ้ำ)
                        target_students = list(dict.fromkeys(target_students))
                        records = [(collection_id, sid, 'pending') for sid in target_students]
                        await conn.executemany("INSERT INTO student_payments (collection_id, student_id, status) VALUES ($1, $2, $3)", records)
                    
                    new_values = cls._extract_req_data(req)
                    new_values["resolved_target_students"] = target_students
                    msg = f"สร้างแคมเปญสำเร็จ เรียกเก็บเพื่อน {len(target_students)} คน"

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FEE_COLLECTION", entity_id=str(collection_id), status="success",
                        new_values=new_values, endpoint_or_command="FinanceService.create_fee_collection", execution_time_ms=exec_time
                    )
                    # 📢 แจ้งเตือน Discord: สร้างแคมเปญเก็บเงินใหม่ → @everyone (ทุกคนต้องรู้)
                    room_server_id = await cls._get_room_server_id(conn, target_room_id)
            if room_server_id:
                await ActionService.notify_new_collection(
                    server_id=room_server_id,
                    title=req.title,
                    amount=float(req.amount),
                    due_date=req.due_date,
                    user_name=req.user_name,
                )
            return {"status": "success", "message": msg}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FEE_COLLECTION", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.create_fee_collection", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def confirm_payment(cls, pool: asyncpg.Pool, payment_id: int, req, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    # 🛡️ RBAC: มีแค่ผู้ดูแลการเงิน (MANAGE_FINANCE) ถึงจะรับเงินได้
                    if user_id is not None:
                        await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    valid_account = await conn.fetchval("SELECT id FROM finance_accounts WHERE id = $1 AND room_id = $2", req.paid_to_account_id, target_room_id)
                    if not valid_account: raise ValueError("กระเป๋าเงินไม่มีอยู่ หรือไม่ใช่ของห้องนี้!")

                    old_sp_data = await conn.fetchrow("SELECT * FROM student_payments WHERE id = $1 FOR UPDATE", payment_id)
                    old_values = dict(old_sp_data) if old_sp_data else {}

                    payment_info = await conn.fetchrow(
                        """SELECT FC.amount as total_amount, SP.paid_amount as current_paid, FC.title, U.first_name, U.nickname
                           FROM student_payments SP
                           JOIN fee_collections FC ON SP.collection_id = FC.id
                           JOIN students S ON SP.student_id = S.id
                           LEFT JOIN users U ON S.user_id = U.id
                           WHERE SP.id = $1 AND FC.room_id = $2 AND FC.status = 'active'""",
                        payment_id, target_room_id
                    )
                    if not payment_info: raise PaymentNotFoundError("ไม่พบรายการนี้ หรือแคมเปญถูกปิดไปแล้ว")
                    
                    current_paid = float(payment_info['current_paid'])
                    total_amount = float(payment_info['total_amount'])

                    if current_paid >= total_amount: raise ValueError("บิลนี้จ่ายครบไปเรียบร้อยแล้วครับ!")

                    # 🛡️ กัน overpay: ห้ามรับเงินเกินยอดที่เหลือค้าง (current_paid + paid_amount > total)
                    if req.paid_amount > total_amount - current_paid:
                        raise ValueError(
                            f"จำนวนเงินที่รับเกินยอดที่เหลือค้าง! "
                            f"เหลือค้าง {total_amount - current_paid:.2f} บาท แต่ส่งมา {req.paid_amount:.2f} บาท"
                        )

                    new_total_paid = current_paid + req.paid_amount
                    new_status = 'paid' if new_total_paid >= total_amount else 'pending'
                    status_msg = "จ่ายครบแล้ว" if new_status == 'paid' else f"ทยอยจ่าย (ขาดอีก {total_amount - new_total_paid} ฿)"

                    stu_name = payment_info['first_name'] or "Unknown"
                    if payment_info['nickname']: stu_name += f" ({payment_info['nickname']})"
                    dynamic_desc = f"รับเงิน: {payment_info['title']} จาก {stu_name} [{status_msg}]"
                    
                    trans_id = await conn.fetchval(
                        """INSERT INTO finance_transactions 
                           (room_id, account_id, amount, description, transaction_type, slip_image_url, recorded_by, student_payment_id) 
                           VALUES ($1, $2, $3, $4, 'income', $5, $6, $7) RETURNING id""",
                        target_room_id, req.paid_to_account_id, req.paid_amount, dynamic_desc, req.slip_image_url, req.user_name, payment_id
                    )
                    
                    await conn.execute("UPDATE finance_accounts SET balance = balance + $1 WHERE id = $2", req.paid_amount, req.paid_to_account_id)
                    await conn.execute(
                        """UPDATE student_payments
                           SET status = $1, paid_amount = $2, paid_to_account_id = $3, slip_image_url = $4, recorded_by = $5, paid_at = NOW(), transaction_id = $6
                           WHERE id = $7""",
                        new_status, new_total_paid, req.paid_to_account_id, req.slip_image_url, req.user_name, trans_id, payment_id
                    )

                    # [DUAL-WRITE] เขียนฝั่ง Double-Entry: รับชำระเงินจากนักเรียน (Dr สินทรัพย์ / Cr รายได้เก็บเงินห้อง)
                    asset_ledger_id = await cls._resolve_asset_ledger(conn, target_room_id, req.paid_to_account_id)
                    # Credit ขาเป็นรายได้ "เก็บเงินห้องปกติ" — ถ้าไม่มี mapping ให้หา ledger ตามชื่อ
                    # (seed ค่าเริ่มต้น DEFAULT_INCOME_CATEGORIES[0] = '📥 เก็บเงินห้องปกติ')
                    revenue_ledger_id = await cls._find_revenue_ledger_by_name(
                        conn, target_room_id, account_name=DEFAULT_INCOME_CATEGORIES[0]
                    )
                    if revenue_ledger_id is None:
                        # 💡 ห้องที่ยังไม่มี ledger รายได้เลย (เช่น ข้อมูลเก่าที่ยังไม่ผ่าน migration)
                        # → ลอง provision จาก legacy category ที่ชื่อ '📥 เก็บเงินห้องปกติ' (ถ้า seed ไว้)
                        #   เพื่อให้ journal ครบฝั่ง (กันบัญชีไม่สมดุล) — ถ้าไม่มี category นั้นด้วย → ข้าม dual-write
                        legacy_cat_id = await conn.fetchval(
                            """SELECT id FROM finance_categories
                               WHERE room_id = $1 AND category_name = $2 AND category_type = 'income'
                               ORDER BY id LIMIT 1""",
                            target_room_id, DEFAULT_INCOME_CATEGORIES[0],
                        )
                        if legacy_cat_id:
                            revenue_ledger_id = await cls._resolve_category_ledger(conn, target_room_id, legacy_cat_id, 'income')
                    if revenue_ledger_id is None:
                        # ยังไม่มี ledger รายได้ของห้องจริง ๆ → สร้าง journal ฝั่ง Debit อย่างเดียวไม่ได้
                        # (บัญชีไม่สมดุล) → ข้าม Dual-Write ไป (legacy ยังทำงานปกติเหมือนเดิม)
                        pass
                    else:
                        await cls._insert_journal_entry(
                            conn, target_room_id,
                            reference_type="student_payment",
                            reference_id=str(payment_id),
                            description=dynamic_desc,
                            slip_image_url=req.slip_image_url,
                            recorded_by=req.user_name,
                            metadata={"student_payment_id": payment_id, "legacy_transaction_id": trans_id},
                            lines=[
                                {"ledger_id": asset_ledger_id, "debit": req.paid_amount, "credit": 0, "line_description": f"รับเงินจากนักเรียน (student_payment #{payment_id})"},
                                {"ledger_id": revenue_ledger_id, "debit": 0, "credit": req.paid_amount, "line_description": f"รายได้: {payment_info['title']}"},
                            ],
                        )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="STUDENT_PAYMENT", entity_id=str(payment_id), status="success",
                        old_values=old_values, new_values=new_values, endpoint_or_command="FinanceService.confirm_payment", execution_time_ms=exec_time
                    )
                    # 📢 แจ้งเตือน Discord: มีคนจ่ายเงินแล้ว (ไม่ @everyone — โชว์ความโปร่งใส)
                    room_server_id = await cls._get_room_server_id(conn, target_room_id)
            if room_server_id:
                # ⚠️ [LOW-PRIORITY] สแปมทุกครั้งที่มีคนจ่ายเงิน ("คนนี้จ่ายตังค์แล้ว") → เดี๋ยวจะปิด/ลดการแจ้งเตือนนี้ทีหลัง (คอมเมนต์ไว้เพื่อเตือน)
                await ActionService.notify_payment_confirmed(
                    server_id=room_server_id,
                    payer_name=stu_name,
                    title=payment_info['title'],
                    amount=float(req.paid_amount),
                    user_name=req.user_name,
                )
            return {"status": "success", "message": f"รับเงินสำเร็จ! สถานะ: {status_msg}"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="STUDENT_PAYMENT", entity_id=str(payment_id), status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.confirm_payment", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def get_collection_status(cls, pool: asyncpg.Pool, collection_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                sql = """
                    SELECT 
                        SP.id as payment_id, SP.student_id, SP.status, SP.paid_amount, SP.paid_at, SP.slip_image_url,
                        S.student_no, U.first_name, U.last_name, U.nickname, FC.amount as total_amount
                    FROM student_payments SP
                    JOIN students S ON SP.student_id = S.id
                    LEFT JOIN users U ON S.user_id = U.id
                    JOIN fee_collections FC ON SP.collection_id = FC.id
                    WHERE SP.collection_id = $1 AND S.room_id = $2
                    ORDER BY S.student_no ASC
                """
                rows = await conn.fetch(sql, collection_id, target_room_id)
                total = len(rows)
                paid_count = sum(1 for r in rows if r['status'] == 'paid')
                result = {"collection_id": collection_id, "summary": {"total": total, "paid": paid_count, "pending": total - paid_count}, "students": [dict(r) for r in rows]}

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="STUDENT_PAYMENT", entity_id=str(collection_id), status="success",
                    endpoint_or_command="FinanceService.get_collection_status", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="STUDENT_PAYMENT", entity_id=str(collection_id), status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_collection_status", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def remove_student_from_collection(cls, pool: asyncpg.Pool, collection_id: int, student_id: int, user_id: int, client_source: str, actor_identifier: str, user_name: str = "—", server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")

                    payment = await conn.fetchrow("""
                        SELECT SP.*, FC.title 
                        FROM student_payments SP
                        JOIN fee_collections FC ON SP.collection_id = FC.id
                        WHERE SP.collection_id = $1 AND SP.student_id = $2 AND FC.room_id = $3
                        FOR UPDATE OF SP
                    """, collection_id, student_id, target_room_id)

                    if not payment:
                        raise PaymentNotFoundError("ไม่พบข้อมูลการเรียกเก็บเงินของนักเรียนคนนี้")
                    
                    old_values = dict(payment)
                    if payment['paid_amount'] > 0:
                        raise ValueError("ไม่สามารถลบได้ เนื่องจากนักเรียนมีการจ่ายเงิน (หรือทยอยจ่าย) เข้ามาแล้ว ให้ใช้วิธียกเลิกธุรกรรมการเงินแทน")

                    await conn.execute("DELETE FROM student_payments WHERE id = $1", payment['id'])

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="STUDENT_PAYMENT", entity_id=str(payment['id']), status="success",
                        old_values=old_values, endpoint_or_command="FinanceService.remove_student_from_collection", execution_time_ms=exec_time
                    )
                    return {"status": "success", "message": "ลบรายชื่อนักเรียนออกจากรายการนี้สำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="DELETE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="STUDENT_PAYMENT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.remove_student_from_collection", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def create_category(cls, pool: asyncpg.Pool, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    # [DUAL-WRITE] ดึง id ของแถว legacy เพื่อ map ลง accounting_ledgers
                    new_category_id = await conn.fetchval(
                        "INSERT INTO finance_categories (room_id, category_name, category_type) VALUES ($1, $2, $3) RETURNING id",
                        target_room_id, req.category_name, req.category_type
                    )

                    # [DUAL-WRITE] สร้าง ledger ฝั่ง Double-Entry (income → 'revenue' 4xxxx, expense → 5xxxx)
                    # รหัสบัญชีตรงกับ migrate_phase2_ledgers.py
                    if req.category_type == 'income':
                        ledger_type, ledger_code = 'revenue', f"4{new_category_id:04d}"
                    else:
                        ledger_type, ledger_code = 'expense', f"5{new_category_id:04d}"
                    await conn.execute(
                        """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, legacy_category_id, description)
                           VALUES ($1, $2, $3, $4, $5, 'Created via dual-write (create_category)')""",
                        target_room_id, ledger_code, req.category_name, ledger_type, new_category_id
                    )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_CATEGORY", status="success",
                        new_values=new_values, endpoint_or_command="FinanceService.create_category", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": f"เพิ่มหมวดหมู่ {req.category_name} แล้ว"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_CATEGORY", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.create_category", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def get_categories(cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str, cat_type: Optional[str] = None, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> List[dict]:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                if cat_type:
                    rows = await conn.fetch("SELECT id, category_name, category_type FROM finance_categories WHERE room_id = $1 AND category_type = $2 ORDER BY id", target_room_id, cat_type)
                else:
                    rows = await conn.fetch("SELECT id, category_name, category_type FROM finance_categories WHERE room_id = $1 ORDER BY id", target_room_id)
                result = [dict(row) for row in rows]

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="FINANCE_CATEGORY", status="success",
                    endpoint_or_command="FinanceService.get_categories", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="FINANCE_CATEGORY", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_categories", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def revert_transaction(cls, pool: asyncpg.Pool, transaction_id: int, user_id: int, client_source: str, actor_identifier: str, user_name: str = "—", server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    t = await conn.fetchrow(
                        "SELECT * FROM finance_transactions WHERE id = $1 AND room_id = $2 AND deleted_at IS NULL FOR UPDATE",
                        transaction_id, target_room_id
                    )
                    if not t: raise TransactionNotFoundError("ไม่พบรายการธุรกรรมนี้")
                    old_values = dict(t)

                    if t['transfer_group_id']:
                        group_trans = await conn.fetch("SELECT * FROM finance_transactions WHERE transfer_group_id = $1 AND room_id = $2 AND deleted_at IS NULL FOR UPDATE", t['transfer_group_id'], target_room_id)
                        for gt in group_trans:
                            if gt['transaction_type'] == 'expense': 
                                await conn.execute("UPDATE finance_accounts SET balance = balance + $1 WHERE id = $2", gt['amount'], gt['account_id'])
                            elif gt['transaction_type'] == 'income':
                                curr_bal = await conn.fetchval("SELECT balance FROM finance_accounts WHERE id = $1 FOR UPDATE", gt['account_id'])
                                if float(curr_bal) < float(gt['amount']): raise ValueError("เงินในบัญชีรับโอนไม่พอหักคืน")
                                await conn.execute("UPDATE finance_accounts SET balance = balance - $1 WHERE id = $2", gt['amount'], gt['account_id'])
                        await conn.execute("UPDATE finance_transactions SET deleted_at = NOW() WHERE transfer_group_id = $1 AND room_id = $2", t['transfer_group_id'], target_room_id)
                        # [DUAL-WRITE] ยกเลิก journal ของรายการโอนเงิน (หาด้วย metadata.transfer_group_id)
                        # 💡 ถ้ายังไม่มี journal (ข้อมูลเก่าที่ไม่ได้ผ่าน dual-write) → no-op ไม่พัง
                        await conn.execute(
                            """UPDATE journal_entries
                               SET status = 'voided', deleted_at = NOW(), updated_at = CURRENT_TIMESTAMP
                               WHERE room_id = $1 AND status <> 'voided' AND deleted_at IS NULL
                                 AND metadata->>'transfer_group_id' = $2""",
                            target_room_id, str(t['transfer_group_id']),
                        )
                        action_detail = "ยกเลิกรายการโอนเงิน"
                    else:
                        if t['transaction_type'] == 'income':
                            curr_bal = await conn.fetchval("SELECT balance FROM finance_accounts WHERE id = $1 FOR UPDATE", t['account_id'])
                            if float(curr_bal) < float(t['amount']): raise ValueError("เงินในบัญชีไม่พอหักคืน")
                            await conn.execute("UPDATE finance_accounts SET balance = balance - $1 WHERE id = $2", t['amount'], t['account_id'])
                            
                            if t['student_payment_id']:
                                sp_id = t['student_payment_id']
                                sp_info = await conn.fetchrow("SELECT paid_amount, FC.amount as total_amount FROM student_payments SP JOIN fee_collections FC ON SP.collection_id = FC.id WHERE SP.id = $1 FOR UPDATE", sp_id)
                                new_paid = float(sp_info['paid_amount']) - float(t['amount'])
                                new_status = 'paid' if new_paid >= float(sp_info['total_amount']) else 'pending'
                                # 💡 ถ้ายกเลิกจน paid_amount กลับเป็น 0 ให้ล้าง field การชำระทั้งหมด
                                # (กันสถานะ "จ่ายครบ" ค้างทั้งที่ยอดโดนหักคืนแล้ว)
                                if new_paid <= 0:
                                    await conn.execute(
                                        """UPDATE student_payments
                                           SET paid_amount = 0, status = 'pending', paid_to_account_id = NULL,
                                               slip_image_url = NULL, recorded_by = NULL, paid_at = NULL, transaction_id = NULL
                                           WHERE id = $1""", sp_id
                                    )
                                else:
                                    await conn.execute("UPDATE student_payments SET paid_amount = $1, status = $2 WHERE id = $3", new_paid, new_status, sp_id)

                        elif t['transaction_type'] == 'expense': 
                            await conn.execute("UPDATE finance_accounts SET balance = balance + $1 WHERE id = $2", t['amount'], t['account_id'])
                        
                        await conn.execute("UPDATE finance_transactions SET deleted_at = NOW() WHERE id = $1", transaction_id)
                        # [DUAL-WRITE] ยกเลิก journal ของรายการปกติ (หาด้วย metadata.legacy_transaction_id)
                        # 💡 ถ้ายังไม่มี journal (ข้อมูลเก่า/จาก test ที่ insert ตรง) → no-op ไม่พัง
                        await conn.execute(
                            """UPDATE journal_entries
                               SET status = 'voided', deleted_at = NOW(), updated_at = CURRENT_TIMESTAMP
                               WHERE room_id = $1 AND status <> 'voided' AND deleted_at IS NULL
                                 AND metadata->>'legacy_transaction_id' = $2""",
                            target_room_id, str(transaction_id),
                        )
                        action_detail = f"ยกเลิกรายการ {t['transaction_type']}"

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_TRANSACTION", entity_id=str(transaction_id), status="success",
                        old_values=old_values, new_values={"action": action_detail}, endpoint_or_command="FinanceService.revert_transaction", execution_time_ms=exec_time
                    )
                    return {"status": "success", "message": action_detail}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_TRANSACTION", entity_id=str(transaction_id), status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.revert_transaction", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e
            
    @classmethod
    # =====================================================================
    # [ROUTER] get_summary — เลือกอ่านจาก legacy หรือ Double-Entry ตามช่วงเวลา
    # =====================================================================
    @classmethod
    async def get_summary(
        cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str,
        month: Optional[int] = None, year: Optional[int] = None,
        server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None
    ) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)

                # 💡 month/year ต้องระบุพร้อมกันเสมอ (ไม่งั้น params เลื่อนทำให้ SQL error)
                if (month is None) != (year is None):
                    raise ValueError("ต้องระบุทั้ง month และ year พร้อมกัน หรือไม่ระบุทั้งคู่")

                # [ROUTER] จุดแบ่งเวลา: งวดที่เริ่มหลัง CUTOFF_DATE → ระบบบัญชีคู่
                # (ไม่ระบุงวด = เดือนปัจจุบัน → ขึ้นอยู่กับ wall clock ว่าเลยวันที่ตัดหรือยัง)
                period_start = cls._period_start(month, year, None, None)
                use_v2 = period_start >= CUTOFF_DATE

                if use_v2:
                    return await cls._get_summary_v2(
                        conn=conn, room_id=target_room_id,
                        month=month, year=year,
                        client_source=client_source, actor_identifier=actor_identifier,
                        start_time=start_time,
                    )
                return await cls._get_summary_legacy(
                    conn=conn, room_id=target_room_id,
                    month=month, year=year,
                    client_source=client_source, actor_identifier=actor_identifier,
                    start_time=start_time,
                )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="FINANCE_SUMMARY", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_summary", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def _get_summary_legacy(
        cls, conn: asyncpg.Connection, *, room_id: int,
        month: Optional[int] = None, year: Optional[int] = None,
        client_source: str = "", actor_identifier: str = "", start_time: Optional[float] = None,
    ) -> dict:
        """[ROUTER-LEGACY] Logic เดิมของ get_summary — อ่านจาก finance_accounts + finance_transactions."""
        net_worth = await conn.fetchval("SELECT SUM(balance) FROM finance_accounts WHERE room_id = $1", room_id) or 0.0

        params = [room_id]
        if month and year:
            date_cond = "AND EXTRACT(MONTH FROM created_at) = $2 AND EXTRACT(YEAR FROM created_at) = $3"
            date_cond_t = "AND EXTRACT(MONTH FROM T.created_at) = $2 AND EXTRACT(YEAR FROM T.created_at) = $3"
            params.extend([month, year])
            period_str = f"{year}-{month:02d}"
        else:
            date_cond = "AND date_trunc('month', created_at) = date_trunc('month', CURRENT_DATE)"
            date_cond_t = "AND date_trunc('month', T.created_at) = date_trunc('month', CURRENT_DATE)"
            period_str = "current_month"

        stats = await conn.fetchrow(f"""
            SELECT
                SUM(CASE WHEN transaction_type = 'income' AND transfer_group_id IS NULL THEN amount ELSE 0 END) as total_inc,
                SUM(CASE WHEN transaction_type = 'expense' AND transfer_group_id IS NULL THEN amount ELSE 0 END) as total_exp
            FROM finance_transactions WHERE room_id = $1 AND deleted_at IS NULL {date_cond}
        """, *params)

        breakdown = await conn.fetch(f"""
            SELECT C.category_name, SUM(T.amount) as total_amount
            FROM finance_transactions T
            JOIN finance_categories C ON T.category_id = C.id
            WHERE T.room_id = $1 AND T.transaction_type = 'expense' AND T.transfer_group_id IS NULL AND T.deleted_at IS NULL {date_cond_t}
            GROUP BY C.category_name ORDER BY total_amount DESC
        """, *params)

        pending_collection = await conn.fetchval("""
            SELECT SUM(FC.amount - SP.paid_amount) FROM student_payments SP
            JOIN fee_collections FC ON SP.collection_id = FC.id
            WHERE SP.status = 'pending' AND FC.room_id = $1 AND FC.status = 'active'
        """, room_id) or 0.0

        result = {
            "net_worth": float(net_worth), "total_income": float(stats['total_inc'] or 0),
            "total_expense": float(stats['total_exp'] or 0), "pending_collection_amount": float(pending_collection),
            "period": period_str, "expense_breakdown": [dict(b) for b in breakdown]
        }

        if start_time is not None:
            exec_time = int((time.time() - start_time) * 1000)
            await service_logger.log(
                conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                room_id=room_id, user_id=None, entity_type="FINANCE_SUMMARY", status="success",
                endpoint_or_command="FinanceService.get_summary", execution_time_ms=exec_time
            )
        return result

    # =====================================================================
    # [DOUBLE-ENTRY] _get_summary_v2 — สรุปยอดจาก ledger ตามหลักบัญชีคู่
    # =====================================================================
    @classmethod
    async def _get_summary_v2(
        cls, conn: asyncpg.Connection, *, room_id: int,
        month: Optional[int] = None, year: Optional[int] = None,
        client_source: str = "", actor_identifier: str = "", start_time: Optional[float] = None,
    ) -> dict:
        """สรุปยอดจากระบบบัญชีคู่ (journal_lines) แทนการรวมจาก finance_transactions.

        - Net Worth : SUM(Dr) − SUM(Cr) ของทุก ledger ประเภท 'asset' (ยอดสะสมทั้งหมด)
        - รายได้     : SUM(Cr) − SUM(Dr) ของ 'revenue' ในงวดที่ขอ
        - รายจ่าย   : SUM(Dr) − SUM(Cr) ของ 'expense' ในงวดที่ขอ
        - Expense Breakdown: group ตาม account_name ของ ledger ประเภท 'expense'
        - ยอดยกมา (opening_balance) ไม่ถูกนับเป็นรายได้ของงวด (มันคือทุน ไม่ใช่รายได้)
        """
        if month is not None and year is not None:
            start_dt = date(year, month, 1)
            if month == 12:
                end_dt = date(year + 1, 1, 1)
            else:
                end_dt = date(year, month + 1, 1)
            period_str = f"{year}-{month:02d}"
        else:
            today = datetime.now(THAI_TZ).date()
            start_dt = date(today.year, today.month, 1)
            if today.month == 12:
                end_dt = date(today.year + 1, 1, 1)
            else:
                end_dt = date(today.year, today.month + 1, 1)
            period_str = "current_month"

        # [DOUBLE-ENTRY] ยอดสินทรัพย์สะสมทั้งห้อง (ไม่จำกัดงวด) — เทียบเท่า SUM(balance)
        # 💡 นับเฉพาะ journal ที่ไม่ได้ void และไม่ได้ลบ (ลบ legacy ที่ delete ไป)
        net_worth = await conn.fetchval(
            """SELECT COALESCE(SUM(L.debit - L.credit), 0)
               FROM journal_lines L
               JOIN journal_entries JE ON L.journal_entry_id = JE.id
               JOIN accounting_ledgers AL ON L.ledger_id = AL.id
               WHERE JE.room_id = $1 AND AL.account_type = 'asset'
                 AND JE.deleted_at IS NULL AND JE.status <> 'voided'""",
            room_id,
        ) or 0.0

        # [DOUBLE-ENTRY] ยอดรายได้/รายจ่ายภายในงวด (เฉพาะ journal ที่ไม่ใช่ opening_balance)
        period_stats = await conn.fetchrow(
            """SELECT
                 COALESCE(SUM(CASE WHEN AL.account_type = 'revenue' THEN L.credit - L.debit ELSE 0 END), 0) AS total_inc,
                 COALESCE(SUM(CASE WHEN AL.account_type = 'expense' THEN L.debit - L.credit ELSE 0 END), 0) AS total_exp
               FROM journal_lines L
               JOIN journal_entries JE ON L.journal_entry_id = JE.id
               JOIN accounting_ledgers AL ON L.ledger_id = AL.id
               WHERE JE.room_id = $1 AND JE.deleted_at IS NULL AND JE.status <> 'voided'
                 AND JE.reference_type <> 'opening_balance'
                 AND JE.transaction_date >= $2 AND JE.transaction_date < $3""",
            room_id, start_dt, end_dt,
        )

        # [DOUBLE-ENTRY] รายจ่ายรายหมวด (จากชื่อ ledger ฝั่ง expense) ภายในงวด
        breakdown_rows = await conn.fetch(
            """SELECT AL.account_name AS category_name, SUM(L.debit - L.credit) AS total_amount
               FROM journal_lines L
               JOIN journal_entries JE ON L.journal_entry_id = JE.id
               JOIN accounting_ledgers AL ON L.ledger_id = AL.id
               WHERE JE.room_id = $1 AND JE.deleted_at IS NULL AND JE.status <> 'voided'
                 AND JE.reference_type <> 'opening_balance'
                 AND AL.account_type = 'expense'
                 AND JE.transaction_date >= $2 AND JE.transaction_date < $3
               GROUP BY AL.account_name
               ORDER BY total_amount DESC""",
            room_id, start_dt, end_dt,
        )

        pending_collection = await conn.fetchval("""
            SELECT SUM(FC.amount - SP.paid_amount) FROM student_payments SP
            JOIN fee_collections FC ON SP.collection_id = FC.id
            WHERE SP.status = 'pending' AND FC.room_id = $1 AND FC.status = 'active'
        """, room_id) or 0.0

        result = {
            "net_worth": float(net_worth),
            "total_income": float(period_stats["total_inc"] or 0),
            "total_expense": float(period_stats["total_exp"] or 0),
            "pending_collection_amount": float(pending_collection),
            "period": period_str,
            "expense_breakdown": [
                {"category_name": b["category_name"], "total_amount": float(b["total_amount"])}
                for b in breakdown_rows
            ],
        }

        if start_time is not None:
            exec_time = int((time.time() - start_time) * 1000)
            await service_logger.log(
                conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                room_id=room_id, user_id=None, entity_type="FINANCE_SUMMARY", status="success",
                endpoint_or_command="FinanceService.get_summary", execution_time_ms=exec_time
            )
        return result

    # =====================================================================
    # [DOUBLE-ENTRY] งบทดลอง (Trial Balance) — จุดแข็งของระบบบัญชีคู่
    # =====================================================================
    @classmethod
    async def get_trial_balance(
        cls, pool: asyncpg.Pool, room_id: int,
        client_source: str = "", actor_identifier: str = "",
        user_id: Optional[int] = None, as_of_date: Optional[date] = None, server_id: Optional[int] = None,
    ) -> dict:
        """งบทดลอง: ยอด YTD (Year-to-Date) ของทุก ledger ที่ยัง active ในห้อง.

        กติกาการหักยอดตามประเภทบัญชี (สเปค Phase 4):
        - Assets & Expenses   : Dr − Cr
        - Liabilities, Equity, Revenue : Cr − Dr

        คืน {"ledgers": [...], "total_debit": X, "total_credit": Y, "is_balanced": bool}
        โดย total_debit/total_credit คือผลรวมของยอด Dr/Cr รวม (ไม่ใช่สุทธิ) ของทุก ledger
        → ถ้า journal ทุกรายการสมดุล (Dr = Cr เสมอ) ค่าเท่ากัน และ is_balanced = True
        """
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ ข้อมูลการเงิน → ต้องเป็นสมาชิกห้องเท่านั้น
                await require_member(conn, target_room_id, user_id)

                # [DOUBLE-ENTRY] ขอบเขตเวลา: ถึง as_of_date (ถ้าไม่ระบุ = ทั้งหมดจนถึงตอนนี้)
                if as_of_date is not None:
                    date_filter = "AND JE.transaction_date < $2"
                    params = [target_room_id, datetime.combine(as_of_date, dtime(23, 59, 59))]
                else:
                    date_filter = ""
                    params = [target_room_id]

                # [DOUBLE-ENTRY] รวม Dr/Cr ของทุก ledger ที่ active ยังไม่ void
                rows = await conn.fetch(
                    f"""SELECT
                            AL.id AS ledger_id,
                            AL.account_code,
                            AL.account_name,
                            AL.account_type,
                            COALESCE(SUM(L.debit), 0)  AS total_debit,
                            COALESCE(SUM(L.credit), 0) AS total_credit
                        FROM accounting_ledgers AL
                        LEFT JOIN journal_lines L ON L.ledger_id = AL.id
                        LEFT JOIN journal_entries JE ON L.journal_entry_id = JE.id
                            AND JE.deleted_at IS NULL AND JE.status <> 'voided'
                        WHERE AL.room_id = $1 AND AL.is_active = TRUE {date_filter}
                        GROUP BY AL.id, AL.account_code, AL.account_name, AL.account_type
                        ORDER BY AL.account_code NULLS LAST, AL.id""",
                    *params,
                )

                ledgers: List[dict] = []
                grand_debit = 0.0
                grand_credit = 0.0
                for r in rows:
                    dr = float(r["total_debit"])
                    cr = float(r["total_credit"])
                    # [DOUBLE-ENTRY] ตามสมการปกติของงบทดลอง
                    if r["account_type"] in ("asset", "expense"):
                        balance = dr - cr
                    else:  # liability / equity / revenue
                        balance = cr - dr
                    grand_debit += dr
                    grand_credit += cr
                    ledgers.append({
                        "ledger_id": r["ledger_id"],
                        "account_code": r["account_code"],
                        "account_name": r["account_name"],
                        "account_type": r["account_type"],
                        "total_debit": dr,
                        "total_credit": cr,
                        "balance": balance,
                    })

                result = {
                    "ledgers": ledgers,
                    "total_debit": grand_debit,
                    "total_credit": grand_credit,
                    "is_balanced": abs(grand_debit - grand_credit) < 0.01,  # เผื่อ floating noise
                }

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="TRIAL_BALANCE", status="success",
                    endpoint_or_command="FinanceService.get_trial_balance", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="TRIAL_BALANCE", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_trial_balance", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    # =====================================================================
    # [DOUBLE-ENTRY] งบกำไรขาดทุน (Income Statement) — ตามงวดเวลา
    # =====================================================================
    @classmethod
    async def get_income_statement(
        cls, pool: asyncpg.Pool, room_id: int, start_date: date, end_date: date,
        client_source: str = "", actor_identifier: str = "",
        user_id: Optional[int] = None, server_id: Optional[int] = None,
    ) -> dict:
        """งบกำไรขาดทุน: รวมรายได้/ค่าใช้จ่ายภายในช่วงเวลาที่กำหนด.

        - revenues: group ตามชื่อ ledger ประเภท 'revenue' (SUM(credit − debit))
        - expenses: group ตามชื่อ ledger ประเภท 'expense' (SUM(debit − credit))
        - net_income = Total Revenue − Total Expense
        - ไม่นับยอดยกมา (opening_balance) เพราะมันคือทุน ไม่ใช่รายได้ของงวด
        """
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                await require_member(conn, target_room_id, user_id)

                if start_date > end_date:
                    raise ValueError("วันที่เริ่มต้นต้องไม่เกินวันที่สิ้นสุด")

                # [DOUBLE-ENTRY] ขอบเขตปลาย → คร่อมทั้งวันของ end_date
                end_bound = datetime.combine(end_date, dtime(23, 59, 59))

                rev_rows = await conn.fetch(
                    """SELECT AL.account_name,
                              COALESCE(SUM(L.credit - L.debit), 0) AS total
                       FROM journal_lines L
                       JOIN journal_entries JE ON L.journal_entry_id = JE.id
                       JOIN accounting_ledgers AL ON L.ledger_id = AL.id
                       WHERE JE.room_id = $1 AND JE.deleted_at IS NULL AND JE.status <> 'voided'
                         AND JE.reference_type <> 'opening_balance'
                         AND AL.account_type = 'revenue'
                         AND JE.transaction_date >= $2 AND JE.transaction_date <= $3
                       GROUP BY AL.account_name
                       ORDER BY total DESC""",
                    target_room_id, start_date, end_bound,
                )

                exp_rows = await conn.fetch(
                    """SELECT AL.account_name,
                              COALESCE(SUM(L.debit - L.credit), 0) AS total
                       FROM journal_lines L
                       JOIN journal_entries JE ON L.journal_entry_id = JE.id
                       JOIN accounting_ledgers AL ON L.ledger_id = AL.id
                       WHERE JE.room_id = $1 AND JE.deleted_at IS NULL AND JE.status <> 'voided'
                         AND JE.reference_type <> 'opening_balance'
                         AND AL.account_type = 'expense'
                         AND JE.transaction_date >= $2 AND JE.transaction_date <= $3
                       GROUP BY AL.account_name
                       ORDER BY total DESC""",
                    target_room_id, start_date, end_bound,
                )

                revenues = [
                    {"account_name": r["account_name"], "amount": float(r["total"])} for r in rev_rows
                ]
                expenses = [
                    {"account_name": r["account_name"], "amount": float(r["total"])} for r in exp_rows
                ]
                total_revenue = sum(r["amount"] for r in revenues)
                total_expense = sum(r["amount"] for r in expenses)

                result = {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "revenues": revenues,
                    "expenses": expenses,
                    "total_revenue": total_revenue,
                    "total_expense": total_expense,
                    "net_income": total_revenue - total_expense,
                }

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="INCOME_STATEMENT", status="success",
                    endpoint_or_command="FinanceService.get_income_statement", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="INCOME_STATEMENT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_income_statement", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def get_student_debts(cls, pool: asyncpg.Pool, student_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                student = await conn.fetchrow("SELECT S.id, U.first_name, U.nickname FROM students S LEFT JOIN users U ON S.user_id = U.id WHERE S.id = $1 AND S.room_id = $2", student_id, target_room_id)
                if not student: raise RoomNotFoundError("ไม่พบข้อมูลนักเรียนคนนี้ในห้อง")

                rows = await conn.fetch("""
                    SELECT SP.id as payment_id, FC.id as collection_id, FC.title, (FC.amount - COALESCE(SP.paid_amount, 0)) AS amount, FC.due_date, FC.status AS collection_status
                    FROM student_payments SP JOIN fee_collections FC ON SP.collection_id = FC.id
                    WHERE SP.student_id = $1 AND SP.status = 'pending' AND FC.room_id = $2
                    ORDER BY FC.status ASC, FC.due_date ASC
                """, student_id, target_room_id)

                formatted_debts, total_pending = [], 0.0
                for r in rows:
                    row_dict = dict(r)
                    row_dict['amount'] = float(row_dict['amount'])
                    formatted_debts.append(row_dict)
                    total_pending += row_dict['amount']

                formatted_name = student['first_name'] or "Unknown"
                if student['nickname']: formatted_name += f" ({student['nickname']})"

                result = {"student_id": student_id, "student_name": formatted_name, "total_pending_amount": total_pending, "debts": formatted_debts}

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="STUDENT_DEBT", entity_id=str(student_id), status="success",
                    endpoint_or_command="FinanceService.get_student_debts", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="STUDENT_DEBT", entity_id=str(student_id), status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_student_debts", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def add_student_to_collection(cls, pool: asyncpg.Pool, collection_id: int, student_id: int, user_id: int, client_source: str, actor_identifier: str, user_name: str = "—", server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")

                    # 🛡️ ต้องเป็นสมาชิก active เท่านั้น (กัน pending/left student เข้ารายการเก็บเงิน)
                    if not await conn.fetchval("SELECT id FROM students WHERE id = $1 AND room_id = $2 AND status = 'active'", student_id, target_room_id):
                        raise RoomNotFoundError("ไม่พบเด็กคนนี้ในห้อง")
                    if not await conn.fetchval("SELECT id FROM fee_collections WHERE id = $1 AND room_id = $2 AND status = 'active'", collection_id, target_room_id):
                        raise ValueError("ไม่พบรายการเรียกเก็บเงินนี้ หรือแคมเปญถูกปิดไปแล้ว!")

                    try:
                        await conn.execute("INSERT INTO student_payments (collection_id, student_id, status) VALUES ($1, $2, 'pending')", collection_id, student_id)
                    except asyncpg.exceptions.UniqueViolationError:
                        raise ValueError("เพื่อนคนนี้มีชื่อในรายการนี้อยู่แล้ว")

                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="STUDENT_PAYMENT", status="success",
                        new_values={"collection_id": collection_id, "student_id": student_id}, endpoint_or_command="FinanceService.add_student_to_collection", execution_time_ms=exec_time
                    )
                    return {"status": "success", "message": "เพิ่มเพื่อนเข้าสู่การเก็บเงินแล้ว"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="CREATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="STUDENT_PAYMENT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.add_student_to_collection", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e
    
    @classmethod
    async def get_all_collections(cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> List[dict]:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                rows = await conn.fetch("SELECT id, title, amount, due_date, status FROM fee_collections WHERE room_id = $1 ORDER BY id DESC", target_room_id)
                result = [dict(row) for row in rows]

                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="FEE_COLLECTION", status="success",
                    endpoint_or_command="FinanceService.get_all_collections", execution_time_ms=exec_time
                )
                return result
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="FEE_COLLECTION", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_all_collections", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def update_collection(cls, pool: asyncpg.Pool, collection_id: int, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    current_data = await conn.fetchrow("SELECT * FROM fee_collections WHERE id = $1 AND room_id = $2", collection_id, target_room_id)
                    if not current_data: raise RoomNotFoundError("ไม่พบแคมเปญนี้")
                    old_values = dict(current_data)

                    updates, values, idx, changed_labels = [], [], 1, []
                    
                    if req.title is not None:
                        updates.append(f"title = ${idx}"); values.append(req.title); idx += 1; changed_labels.append("title")
                    if req.amount is not None and float(req.amount) != float(current_data['amount']):
                        if await conn.fetchval("SELECT 1 FROM student_payments WHERE collection_id = $1 AND paid_amount > 0 LIMIT 1", collection_id):
                            raise ValueError("ไม่สามารถแก้จำนวนเงินได้ เนื่องจากมีเงินโอนเข้ามาแล้ว!")
                        updates.append(f"amount = ${idx}"); values.append(req.amount); idx += 1; changed_labels.append("amount")
                    if req.due_date is not None:
                        updates.append(f"due_date = ${idx}"); values.append(req.due_date); idx += 1; changed_labels.append("due_date")
                    if req.status is not None:
                        updates.append(f"status = ${idx}"); values.append(req.status); idx += 1; changed_labels.append("status")

                    if not updates: return {"status": "success", "message": "ไม่มีข้อมูลให้เปลี่ยนแปลง"}

                    values.extend([collection_id, target_room_id])
                    res = await conn.execute(f"UPDATE fee_collections SET {', '.join(updates)} WHERE id = ${idx} AND room_id = ${idx + 1}", *values)
                    if res == "UPDATE 0": raise RoomNotFoundError("ไม่พบแคมเปญนี้")

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FEE_COLLECTION", entity_id=str(collection_id), status="success",
                        old_values=old_values, new_values=new_values, endpoint_or_command="FinanceService.update_collection", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": "อัปเดตข้อมูลสำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FEE_COLLECTION", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.update_collection", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def update_account(cls, pool: asyncpg.Pool, account_id: int, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    old_data = await conn.fetchrow("SELECT * FROM finance_accounts WHERE id = $1 AND room_id = $2", account_id, target_room_id)
                    if not old_data: raise RoomNotFoundError("ไม่พบบัญชีนี้")
                    old_values = dict(old_data)

                    res = await conn.execute("UPDATE finance_accounts SET account_name = $1 WHERE id = $2 AND room_id = $3", req.account_name, account_id, target_room_id)
                    if res == "UPDATE 0": raise RoomNotFoundError("ไม่พบบัญชีนี้")

                    # [DUAL-WRITE] ซิงก์ชื่อไปยัง accounting_ledgers (ถ้ามี) — กัน ledger ค้างชื่อเก่า
                    await conn.execute(
                        "UPDATE accounting_ledgers SET account_name = $1, updated_at = CURRENT_TIMESTAMP WHERE legacy_account_id = $2 AND room_id = $3",
                        req.account_name, account_id, target_room_id
                    )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_ACCOUNT", entity_id=str(account_id), status="success",
                        old_values=old_values, new_values=new_values, endpoint_or_command="FinanceService.update_account", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": "อัปเดตชื่อบัญชีสำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_ACCOUNT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.update_account", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def delete_account(cls, pool: asyncpg.Pool, account_id: int, user_id: int, client_source: str, actor_identifier: str, user_name: str = "—", server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    old_data = await conn.fetchrow("SELECT * FROM finance_accounts WHERE id = $1 AND room_id = $2", account_id, target_room_id)
                    if not old_data: raise RoomNotFoundError("ไม่พบบัญชีนี้")
                    old_values = dict(old_data)
                    
                    bal = old_data['balance']
                    if bal > 0: raise ValueError("ไม่สามารถลบบัญชีได้ เนื่องจากยังมีเงินคงเหลืออยู่!")
                    if await conn.fetchval("SELECT 1 FROM student_payments WHERE paid_to_account_id = $1 LIMIT 1", account_id):
                        raise ValueError("ไม่สามารถลบบัญชีได้ เนื่องจากมีประวัติการรับเงินผูกกับบัญชีนี้อยู่!")
                    # 🛡️ กันประวัติธุรกรรมหาย: ถ้ามี finance_transactions อ้างถึงบัญชีนี้ ห้าม hard-delete
                    # (FK account_id ON DELETE SET NULL → ประวัติรายรับ/รายจ่ายจะกลายเป็น NULL)
                    if await conn.fetchval("SELECT 1 FROM finance_transactions WHERE account_id = $1 LIMIT 1", account_id):
                        raise ValueError("ไม่สามารถลบบัญชีได้ เนื่องจากมีประวัติธุรกรรมผูกกับบัญชีนี้!")

                    await conn.execute("DELETE FROM finance_accounts WHERE id = $1", account_id)
                    
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_ACCOUNT", entity_id=str(account_id), status="success",
                        old_values=old_values, endpoint_or_command="FinanceService.delete_account", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": "ลบบัญชีสำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="DELETE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_ACCOUNT", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.delete_account", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def get_all_debtors(cls, pool: asyncpg.Pool, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None) -> List[dict]:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)
                rows = await conn.fetch("""
                    SELECT S.id as student_id, S.student_no, U.first_name, U.nickname, COUNT(SP.id) as overdue_count, SUM(FC.amount - SP.paid_amount) as total_pending_amount
                    FROM students S LEFT JOIN users U ON S.user_id = U.id
                    JOIN student_payments SP ON S.id = SP.student_id JOIN fee_collections FC ON SP.collection_id = FC.id
                    WHERE S.room_id = $1 AND SP.status = 'pending'
                    GROUP BY S.id, S.student_no, U.first_name, U.nickname ORDER BY S.student_no ASC
                """, target_room_id)
                debtors = []
                for r in rows:
                    name = r['first_name'] or "Unknown"
                    if r['nickname']: name += f" ({r['nickname']})"
                    debtors.append({"student_id": r['student_id'], "student_no": r['student_no'], "student_name": name, "overdue_count": r['overdue_count'], "total_pending_amount": float(r['total_pending_amount'])})
                
                exec_time = int((time.time() - start_time) * 1000)
                await service_logger.log(
                    conn=conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                    room_id=target_room_id, user_id=None, entity_type="DEBTOR_LIST", status="success",
                    endpoint_or_command="FinanceService.get_all_debtors", execution_time_ms=exec_time
                )
                return debtors
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="VIEW", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="DEBTOR_LIST", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.get_all_debtors", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e
    
    @classmethod
    async def update_category(cls, pool: asyncpg.Pool, category_id: int, req, user_id: int, client_source: str, actor_identifier: str, server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    old_data = await conn.fetchrow("SELECT * FROM finance_categories WHERE id = $1 AND room_id = $2", category_id, target_room_id)
                    if not old_data: raise RoomNotFoundError("ไม่พบหมวดหมู่นี้")
                    old_values = dict(old_data)

                    res = await conn.execute("UPDATE finance_categories SET category_name = $1 WHERE id = $2 AND room_id = $3", req.category_name, category_id, target_room_id)
                    if res == "UPDATE 0": raise RoomNotFoundError("ไม่พบหมวดหมู่นี้")

                    # [DUAL-WRITE] ซิงก์ชื่อไปยัง accounting_ledgers (ถ้ามี) — กัน ledger ค้างชื่อเก่า
                    await conn.execute(
                        "UPDATE accounting_ledgers SET account_name = $1, updated_at = CURRENT_TIMESTAMP WHERE legacy_category_id = $2 AND room_id = $3",
                        req.category_name, category_id, target_room_id
                    )

                    new_values = cls._extract_req_data(req)
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_CATEGORY", entity_id=str(category_id), status="success",
                        old_values=old_values, new_values=new_values, endpoint_or_command="FinanceService.update_category", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": "อัปเดตชื่อหมวดหมู่สำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="UPDATE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_CATEGORY", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.update_category", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def delete_category(cls, pool: asyncpg.Pool, category_id: int, user_id: int, client_source: str, actor_identifier: str, user_name: str = "—", server_id: Optional[int] = None, room_id: Optional[int] = None) -> dict:
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                    await require_permission(conn, target_room_id, user_id, "MANAGE_FINANCE")
                    
                    old_data = await conn.fetchrow("SELECT * FROM finance_categories WHERE id = $1 AND room_id = $2", category_id, target_room_id)
                    if not old_data: raise RoomNotFoundError("ไม่พบหมวดหมู่นี้")
                    old_values = dict(old_data)

                    if await conn.fetchval("SELECT 1 FROM finance_transactions WHERE category_id = $1 LIMIT 1", category_id):
                        raise ValueError("ไม่สามารถลบได้ เนื่องจากมีการใช้หมวดหมู่นี้อยู่!")
                    res = await conn.execute("DELETE FROM finance_categories WHERE id = $1 AND room_id = $2", category_id, target_room_id)
                    if res == "DELETE 0": raise RoomNotFoundError("ไม่พบหมวดหมู่นี้")
                    
                    exec_time = int((time.time() - start_time) * 1000)
                    await service_logger.log(
                        conn=conn, action="DELETE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_CATEGORY", entity_id=str(category_id), status="success",
                        old_values=old_values, endpoint_or_command="FinanceService.delete_category", execution_time_ms=exec_time
                    )
                return {"status": "success", "message": "ลบหมวดหมู่สำเร็จ"}
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="DELETE", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=user_id, entity_type="FINANCE_CATEGORY", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.delete_category", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    # =====================================================================
    # 📤 Export ประวัติการเงินเป็นไฟล์ Excel (.xlsx)
    # =====================================================================

    # =====================================================================
    # [ROUTER] export_transactions_excel — เลือกอ่านจาก legacy หรือ Double-Entry ตามช่วงเวลา
    # =====================================================================
    @classmethod
    async def export_transactions_excel(
        cls, pool: asyncpg.Pool, req, client_source: str, actor_identifier: str,
        server_id: Optional[int] = None, room_id: Optional[int] = None, user_id: Optional[int] = None
    ) -> io.BytesIO:
        """ส่งออกประวัติการทำรายการการเงินของห้องเป็น .xlsx ที่จัดรูปแบบสวยงาม.

        ช่วงเวลาที่รองรับ (เลือกอย่างใดอย่างหนึ่ง):
        - start_date + end_date  → ช่วงวันที่ที่กำหนด
        - month + year           → ทั้งเดือน
        - ไม่ระบุเลย             → ทุกอย่าง

        คัดเฉพาะคอลัมน์ที่ "คนอ่านเอาไปใช้ต่อได้" (ไม่มี system values เช่น
        deleted_at/transfer_group_id/slip URL) และรวมขาโอนเงินเข้าด้วยกัน
        เพื่อให้ตัวเลขรายรับ/รายจ่ายสะท้อนเงินจริงที่เข้า/ออกห้อง
        """
        start_time = time.time()
        target_room_id = room_id
        try:
            async with pool.acquire() as conn:
                target_room_id = await cls.resolve_room_id(conn, server_id, room_id)
                # 🛡️ สมาชิกห้องดูได้ (transparency) แต่ต้องเป็นสมาชิกห้องนี้เท่านั้น (กันข้ามห้อง)
                await require_member(conn, target_room_id, user_id)

                # [ROUTER] ตัดสินใจด้วยจุดเริ่มต้นของช่วง (ก่อน/หลัง CUTOFF_DATE)
                month, year = getattr(req, "month", None), getattr(req, "year", None)
                start_date = getattr(req, "start_date", None)
                end_date = getattr(req, "end_date", None)
                # โจทย์ export ไม่ระบุ → "ทั้งหมด" ซึ่งรวมทั้งก่อนและหลัง cutoff → ต้องอ่าน legacy
                # (เพราะข้อมูลก่อน cutoff มีแค่ในตารางเก่า) → ใช้ month/start_date ตัดสินใจหลัก
                if (month is not None and year is not None) or start_date is not None or end_date is not None:
                    period_start = cls._period_start(month, year, start_date, end_date)
                    use_v2 = period_start >= CUTOFF_DATE
                else:
                    # [ROUTER] ไม่ระบุช่วง = ขอทุกอย่าง → ครอบคลุมทั้งสองยุค → ใช้ legacy ทั้งหมด
                    # (หลีกเลี่ยงการอ่านข้อมูลก่อน cutoff ผ่าน Double-Entry ที่ไม่มี ledger)
                    use_v2 = False

                if use_v2:
                    return await cls._export_transactions_excel_v2(
                        conn=conn, room_id=target_room_id,
                        month=month, year=year, start_date=start_date, end_date=end_date,
                        client_source=client_source, actor_identifier=actor_identifier,
                        start_time=start_time,
                    )
                return await cls._export_transactions_excel_legacy(
                    conn=conn, room_id=target_room_id,
                    month=month, year=year, start_date=start_date, end_date=end_date,
                    client_source=client_source, actor_identifier=actor_identifier,
                    start_time=start_time,
                )
        except Exception as e:
            exec_time = int((time.time() - start_time) * 1000)
            try:
                async with pool.acquire() as log_conn:
                    await service_logger.log(
                        conn=log_conn, action="EXPORT", actor_identifier=actor_identifier, client_source=client_source,
                        room_id=target_room_id, user_id=None, entity_type="FINANCE_TRANSACTION", status="failed", error_detail=str(e),
                        endpoint_or_command="FinanceService.export_transactions_excel", execution_time_ms=exec_time
                    )
            except Exception:
                pass
            raise e

    @classmethod
    async def _export_transactions_excel_legacy(
        cls, conn: asyncpg.Connection, *, room_id: int,
        month: Optional[int] = None, year: Optional[int] = None,
        start_date: Optional[date] = None, end_date: Optional[date] = None,
        client_source: str = "", actor_identifier: str = "", start_time: Optional[float] = None,
    ) -> io.BytesIO:
        """[ROUTER-LEGACY] Logic เดิมของ export — อ่านจาก finance_transactions (Single-Entry)."""
        where_clause, period_params, period_label = cls._resolve_export_period(
            _ExportPeriodView(month=month, year=year, start_date=start_date, end_date=end_date)
        )

        room = await conn.fetchrow("SELECT room_name FROM rooms WHERE id = $1", room_id)
        room_name = room["room_name"] if room else f"ห้อง #{room_id}"

        rows = await conn.fetch(
            f"""
            SELECT
                T.id, T.transaction_type, T.amount, T.description, T.recorded_by,
                T.created_at, T.transfer_group_id, T.student_payment_id,
                A.account_name, C.category_name
            FROM finance_transactions T
            LEFT JOIN finance_accounts A ON T.account_id = A.id
            LEFT JOIN finance_categories C ON T.category_id = C.id
            WHERE T.room_id = $1 AND T.deleted_at IS NULL {where_clause}
            ORDER BY T.created_at ASC, T.id ASC
            """,
            room_id, *period_params,
        )

        # ยอดคงเหลือปัจจุบันของแต่ละบัญชี (ดึงจาก finance_accounts ตรง ๆ
        # เพื่อสะท้อนยอดจริงรวม seed/เปิดบัญชี — ไม่ใช่แค่เงินที่เคลื่อนในงวดนี้)
        account_balances = await conn.fetch(
            "SELECT account_name, balance FROM finance_accounts WHERE room_id = $1 AND deleted_at IS NULL ORDER BY id",
            room_id,
        )

        final_rows = cls._consolidate_transfers([dict(r) for r in rows])
        excel_file = cls._build_finance_workbook(
            room_name=room_name, period_label=period_label,
            rows=final_rows, account_balances=[(r["account_name"], r["balance"]) for r in account_balances],
            generated_at=datetime.now(THAI_TZ),
        )

        if start_time is not None:
            exec_time = int((time.time() - start_time) * 1000)
            await service_logger.log(
                conn=conn, action="EXPORT", actor_identifier=actor_identifier, client_source=client_source,
                room_id=room_id, user_id=None, entity_type="FINANCE_TRANSACTION", status="success",
                new_values={"period": period_label, "rows": len(final_rows)},
                endpoint_or_command="FinanceService.export_transactions_excel", execution_time_ms=exec_time
            )
        return excel_file

    # =====================================================================
    # [DOUBLE-ENTRY] _export_transactions_excel_v2 — สร้าง Excel จาก journal
    # (แปล Dr/Cr → แถว TransactionResponse แล้วใช้ _build_finance_workbook
    #  เหมือนเดิม แต่ยอดคงเหลือรายบัญชีคำนวณจาก Net Balance ของ ledger)
    # =====================================================================
    @classmethod
    async def _export_transactions_excel_v2(
        cls, conn: asyncpg.Connection, *, room_id: int,
        month: Optional[int] = None, year: Optional[int] = None,
        start_date: Optional[date] = None, end_date: Optional[date] = None,
        client_source: str = "", actor_identifier: str = "", start_time: Optional[float] = None,
    ) -> io.BytesIO:
        # [DOUBLE-ENTRY] แปลงช่วงเวลา → ฉลากเหมือน legacy (ปี-เดือน / ช่วงวันที่ / ทั้งหมด)
        # ⚠️ end_dt ต้องเป็นแบบ "ครอบถึง" (inclusive) เพราะ _get_transactions_v2 กรองด้วย `<= $3`
        # (ต่างจาก legacy month/year ที่ใช้ `<` แบบ exclusive) — กันรายการวันที่ 1 ของเดือนถัดไปหลุดเข้า
        if month is not None and year is not None:
            period_label = f"{year}-{month:02d}"
            start_dt = date(year, month, 1)
            end_dt = date(year, 12, 31) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
        elif start_date is not None or end_date is not None:
            if start_date and end_date and start_date > end_date:
                raise ValueError("วันที่เริ่มต้นต้องไม่เกินวันที่สิ้นสุด")
            if start_date and end_date:
                period_label = f"{start_date.isoformat()} ถึง {end_date.isoformat()}"
            elif start_date:
                period_label = f"ตั้งแต่วันที่ {start_date.isoformat()}"
            else:
                period_label = f"จนถึงวันที่ {end_date.isoformat()}"
            # วันที่เดียว → คร่อมทั้งวัน (inclusive อยู่แล้ว)
            start_dt = start_date or date.min
            end_dt = end_date or date.max
        else:
            period_label = "ทั้งหมด"
            start_dt, end_dt = None, None

        room = await conn.fetchrow("SELECT room_name FROM rooms WHERE id = $1", room_id)
        room_name = room["room_name"] if room else f"ห้อง #{room_id}"

        # [DOUBLE-ENTRY] รายการทั้งหมด (แปลจาก journal) — เรียงตามเวลาเหมือน legacy export
        txn_result = await cls._get_transactions_v2(
            conn=conn, room_id=room_id,
            limit=100000, offset=0,
            start_date=start_dt, end_date=end_dt,
            client_source=client_source, actor_identifier=actor_identifier,
            start_time=None,  # ไม่ log อีกครั้ง (export จะ log เอง)
        )
        raw_items = txn_result["items"]

        # [DOUBLE-ENTRY] จัดรูปให้ _build_finance_workbook ใช้ได้ (รายรับ/รายจ่าย/หมวด/บัญชี)
        final_rows = cls._format_v2_rows(raw_items)
        final_rows.sort(key=lambda x: (x["created_at"] or datetime.min, x["id"] or ""))

        # [DOUBLE-ENTRY] ยอดคงเหลือรายบัญชีจาก Net Balance ของ ledger สินทรัพย์
        # (SUM(debit) − SUM(credit)) — สะท้อนยอดจริงจากระบบบัญชีคู่ ไม่ใช่ finance_accounts
        balances = await conn.fetch(
            """SELECT AL.account_name,
                      COALESCE(SUM(L.debit - L.credit), 0) AS net_balance
               FROM accounting_ledgers AL
               LEFT JOIN journal_lines L ON L.ledger_id = AL.id
               LEFT JOIN journal_entries JE ON L.journal_entry_id = JE.id
                   AND JE.deleted_at IS NULL AND JE.status <> 'voided'
               WHERE AL.room_id = $1 AND AL.account_type = 'asset' AND AL.is_active = TRUE
               GROUP BY AL.id, AL.account_name
               ORDER BY AL.id""",
            room_id,
        )

        excel_file = cls._build_finance_workbook(
            room_name=room_name, period_label=period_label,
            rows=final_rows,
            account_balances=[(r["account_name"], r["net_balance"]) for r in balances],
            generated_at=datetime.now(THAI_TZ),
        )

        if start_time is not None:
            exec_time = int((time.time() - start_time) * 1000)
            await service_logger.log(
                conn=conn, action="EXPORT", actor_identifier=actor_identifier, client_source=client_source,
                room_id=room_id, user_id=None, entity_type="FINANCE_TRANSACTION", status="success",
                new_values={"period": period_label, "rows": len(final_rows)},
                endpoint_or_command="FinanceService.export_transactions_excel", execution_time_ms=exec_time
            )
        return excel_file

    @classmethod
    def _format_v2_rows(cls, items: List[dict]) -> List[dict]:
        """[DOUBLE-ENTRY] แปลงแถว TransactionResponse (จาก _get_transactions_v2)
        → แถวที่ _build_finance_workbook ใช้ (income/expense/category/account/type)."""
        formatted: List[dict] = []
        for t in items:
            amount = float(t["amount"] or 0.0)
            if t["transaction_type"] == "income":
                formatted.append({
                    "id": t.get("id"),
                    "created_at": t.get("created_at"),
                    "type": "รายรับ",
                    "income": amount,
                    "expense": 0.0,
                    "description": t.get("description") or "",
                    "category": t.get("category_name") or "—",
                    "account": t.get("account_name") or "—",
                    "recorded_by": t.get("recorded_by") or "—",
                })
            else:  # expense (รวม transfer ที่แสดงเป็นรายจ่ายขาออก)
                formatted.append({
                    "id": t.get("id"),
                    "created_at": t.get("created_at"),
                    "type": "รายจ่าย",
                    "income": 0.0,
                    "expense": amount,
                    "description": t.get("description") or "",
                    "category": t.get("category_name") or "—",
                    "account": t.get("account_name") or "—",
                    "recorded_by": t.get("recorded_by") or "—",
                })
        return formatted

    @staticmethod
    def _resolve_export_period(req) -> tuple:
        """แปล req (FinanceExportRequest) → (where_sql, params, period_label).

        ถ้าใช้ month/year → ครอบทั้งเดือน (created_at >= วันที่ 1, < วันที่ 1 เดือนถัดไป)
        ถ้าใช้ start_date/end_date → คร่อมวันที่ (ให้ตัวเดียว → ตัวเดียวถูกบังคับ)
        ไม่ระบุเลย → ครอบทุกอย่าง (คอลัมน์ created_at ทั้งหมด)

        หมายเหตุ: ตำแหน่ง placeholder เริ่มที่ $2 เสมอ (เพราะ $1 คือ room_id ใน query หลัก)
        """
        month, year = getattr(req, "month", None), getattr(req, "year", None)
        start_date = getattr(req, "start_date", None)
        end_date = getattr(req, "end_date", None)

        if month is not None and year is not None:
            start = date(year, month, 1)
            if month == 12:
                end = date(year + 1, 1, 1)
            else:
                end = date(year, month + 1, 1)
            return " AND T.created_at >= $2 AND T.created_at < $3", [start, end], f"{year}-{month:02d}"

        if start_date is not None and end_date is not None:
            if start_date > end_date:
                raise ValueError("วันที่เริ่มต้นต้องไม่เกินวันที่สิ้นสุด")
            return (
                " AND DATE(T.created_at) >= $2 AND DATE(T.created_at) <= $3",
                [start_date, end_date],
                f"{start_date.isoformat()} ถึง {end_date.isoformat()}",
            )
        if start_date is not None:
            return " AND DATE(T.created_at) >= $2", [start_date], f"ตั้งแต่วันที่ {start_date.isoformat()}"
        if end_date is not None:
            return " AND DATE(T.created_at) <= $2", [end_date], f"จนถึงวันที่ {end_date.isoformat()}"
        return "", [], "ทั้งหมด"

    @staticmethod
    def _clean_transfer_desc(description: Optional[str], transfer_group_id: Optional[int]) -> str:
        """คำอธิบายรายการโอนเงิน: ตัดคำว่า 'โอนออก:'/'รับโอน:' ซ้ำออก เหลือแค่เรื่องที่โอน."""
        if not description:
            return ""
        # ขาโอนทั้งสองข้างมี transfer_group_id → ใช้คำอธิบายดิบ (มี โอนออก:/รับโอน: ข้างหน้า)
        if transfer_group_id is not None:
            return re.sub(r"^(โอนออก:|รับโอน:)\s*", "", description.strip())
        return description

    @classmethod
    def _consolidate_transfers(cls, rows: List[dict]) -> List[dict]:
        """รวมขาโอนเงิน (transfer_group_id เดียวกัน) เข้าเป็นรายการเดียว.

        ปัญหาของข้อมูลดิบ: การโอนเงินระหว่างบัญชีจะสร้าง 2 รายการ
        (ขาออก 'โอนออก: ...' จากบัญชีต้นทาง + ขาเข้า 'รับโอน: ...' เข้าบัญชีปลายทาง)
        ซึ่งถ้าใส่ลงตารางตรง ๆ จะทำให้รายรับ/รายจ่าย "เกินจริง" (เงินแค่ย้ายบัญชีในห้อง ไม่ได้ออกนอกห้อง)

        → จัดการโดยจับคู่ขาที่มี transfer_group_id เดียวกันเป็น 1 แถว
          โดยแสดงเป็น รายจ่ายต้นทาง (amount ลบ) และปล่อยให้ยอดรวมรายรับ/รายจ่ายสะท้อนเงินจริง
        """
        transfer_groups: Dict[int, dict] = {}
        regular_rows: List[dict] = []

        for r in rows:
            group_id = r.get("transfer_group_id")
            if group_id is None:
                regular_rows.append(r)
                continue
            # เลือก "ขาต้นทาง" (transaction_type = expense) เป็นตัวแทนกลุ่ม
            # เพื่อให้ account_name ในแถวชี้ไปที่บัญชีที่เงินออกจริง
            if group_id not in transfer_groups or r["transaction_type"] == "expense":
                transfer_groups[group_id] = r

        final_rows = []
        for r in regular_rows:
            final_rows.append(cls._format_row(r, is_transfer=False))
        for group_id in sorted(transfer_groups.keys()):
            leg = transfer_groups[group_id]
            final_rows.append(cls._format_row(leg, is_transfer=True))
        # ยังคงเรียงตามเวลาจริง (created_at + id)
        final_rows.sort(key=lambda x: (x["created_at"], x["id"]))
        return final_rows

    @classmethod
    def _format_row(cls, r: dict, is_transfer: bool) -> dict:
        """แปลงแถว asyncpg → dict ที่พร้อมใส่ Excel (คัดเฉพาะคอลัมน์ที่คนอ่านเอาไปใช้ต่อได้)."""
        txn_type = r.get("transaction_type")
        amount = float(r.get("amount") or 0.0)
        account_name = r.get("account_name") or "—"
        category_name = r.get("category_name") or "—"

        if is_transfer:
            return {
                "id": r.get("id"),
                "created_at": r.get("created_at"),
                "type": "โอนเงินระหว่างบัญชี",
                "income": 0.0,
                "expense": amount,
                "description": cls._clean_transfer_desc(r.get("description"), r.get("transfer_group_id")),
                "category": "โอนเงิน",
                "account": account_name,
                "recorded_by": r.get("recorded_by") or "—",
            }

        if txn_type == "income":
            return {
                "id": r.get("id"),
                "created_at": r.get("created_at"),
                "type": "รายรับ",
                "income": amount,
                "expense": 0.0,
                "description": r.get("description") or "",
                "category": category_name,
                "account": account_name,
                "recorded_by": r.get("recorded_by") or "—",
            }
        return {
            "id": r.get("id"),
            "created_at": r.get("created_at"),
            "type": "รายจ่าย",
            "income": 0.0,
            "expense": amount,
            "description": r.get("description") or "",
            "category": category_name,
            "account": account_name,
            "recorded_by": r.get("recorded_by") or "—",
        }

    @classmethod
    def _build_finance_workbook(
        cls, room_name: str, period_label: str, rows: List[dict],
        account_balances: Optional[List[tuple]] = None, generated_at: datetime = None
    ) -> io.BytesIO:
        """สร้าง Workbook 3 แผ่นที่จัดรูปแบบสวยงาม:
          Sheet 1 'สรุปยอด' — ภาพรวมรายรับ/รายจ่าย/ยอดคงเหลือของบัญชี
          Sheet 2 'ประวัติรายการ' — ทุกรายการที่ดึงออกมา (หัวข้อหลัก)
          Sheet 3 'สรุปรายหมวดหมู่' — รวมยอดรายรับ/รายจ่ายรายหมวด
        """
        if generated_at is None:
            generated_at = datetime.now(THAI_TZ)
        income_total = round(sum(r["income"] for r in rows), 2)
        expense_total = round(sum(r["expense"] for r in rows), 2)

        wb = Workbook()
        # ---- Sheet 1: สรุปยอด ----
        ws_summary = wb.active
        ws_summary.title = "สรุปยอด"
        ws_summary.sheet_view.showGridLines = False
        ws_summary.column_dimensions["A"].width = 34
        ws_summary.column_dimensions["B"].width = 26

        HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")   # น้ำเงินเข้ม
        TOTAL_FILL = PatternFill("solid", fgColor="D1D5DB")    # เทาอ่อน
        INCOME_FILL = PatternFill("solid", fgColor="D1FAE5")   # เขียวอ่อน
        EXPENSE_FILL = PatternFill("solid", fgColor="FEE2E2")  # แดงอ่อน
        white_bold = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="0F172A")

        ws_summary["A1"] = f"สรุปการเงิน — {room_name}"
        ws_summary["A1"].font = title_font
        ws_summary["A2"] = f"รอบระยะเวลา: {period_label} · สร้างเมื่อ {generated_at.strftime('%d/%m/%Y %H:%M')} น. (เวลาไทย)"
        ws_summary["A2"].font = Font(color="64748B", size=10)

        # กลุ่มรายรับ/รายจ่าย
        ws_summary["A4"] = "รายรับรวม"
        ws_summary["B4"] = income_total
        ws_summary["A5"] = "รายจ่ายรวม"
        ws_summary["B5"] = expense_total
        ws_summary["A6"] = "คงเหลือ (รายรับ − รายจ่าย)"
        ws_summary["B6"] = round(income_total - expense_total, 2)
        for cell in ("A4", "B4"):
            ws_summary[cell].fill = INCOME_FILL
            ws_summary[cell].font = Font(bold=True)
        for cell in ("A5", "B5"):
            ws_summary[cell].fill = EXPENSE_FILL
            ws_summary[cell].font = Font(bold=True)
        for cell in ("A6", "B6"):
            ws_summary[cell].fill = TOTAL_FILL
            ws_summary[cell].font = Font(bold=True, size=12)
        ws_summary["B4"].number_format = "#,##0.00 \"฿\""
        ws_summary["B5"].number_format = "#,##0.00 \"฿\""
        ws_summary["B6"].number_format = "#,##0.00 \"฿\""

        # ยอดคงเหลือรายบัญชี
        ws_summary["A8"] = "ยอดคงเหลือรายบัญชี"
        ws_summary["A8"].font = Font(bold=True, size=12)
        ws_summary["A9"] = "บัญชี"
        ws_summary["B9"] = "ยอดคงเหลือ (บาท)"
        for cell in ("A9", "B9"):
            ws_summary[cell].fill = HEADER_FILL
            ws_summary[cell].font = white_bold
        row_idx = 10
        for account, bal in (account_balances or []):
            ws_summary.cell(row=row_idx, column=1, value=account)
            ws_summary.cell(row=row_idx, column=2, value=float(bal)).number_format = "#,##0.00 \"฿\""
            row_idx += 1
        # ไม่มีบัญชีในห้องนี้เลย (เช่น ยังไม่เคยเปิดบัญชี) → ใส่ placeholder
        if not account_balances:
            ws_summary.cell(row=row_idx, column=1, value="(ไม่มีรายการในช่วงนี้)")
            ws_summary.cell(row=row_idx, column=2, value=0.0).number_format = "#,##0.00 \"฿\""

        # ---- Sheet 2: ประวัติรายการ (หัวข้อหลัก) ----
        ws_data = wb.create_sheet("ประวัติรายการ")
        headers = [
            ("ลำดับ", 6), ("วันที่", 14), ("เวลา", 10), ("ประเภท", 14), ("รายรับ (บาท)", 14),
            ("รายจ่าย (บาท)", 14), ("รายการ", 42), ("หมวดหมู่", 20), ("บัญชี", 18), ("ผู้บันทึก", 16),
        ]
        ws_data.append([h[0] for h in headers])
        for idx, (_, width) in enumerate(headers, start=1):
            ws_data.column_dimensions[get_column_letter(idx)].width = width
            cell = ws_data.cell(row=1, column=idx)
            cell.fill = HEADER_FILL
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, r in enumerate(rows, start=1):
            ts = r["created_at"]
            if ts is None:
                date_str, time_str = "", ""
            else:
                if isinstance(ts, datetime):
                    dt_local = ts.astimezone(THAI_TZ)
                else:
                    dt_local = datetime.combine(ts, dtime(0))
                date_str = dt_local.strftime("%d/%m/%Y")
                time_str = dt_local.strftime("%H:%M")
            ws_data.append([
                i, date_str, time_str, r["type"], r["income"], r["expense"],
                r["description"], r["category"], r["account"], r["recorded_by"],
            ])
            ws_data.cell(row=i + 1, column=5).number_format = "#,##0.00"
            ws_data.cell(row=i + 1, column=6).number_format = "#,##0.00"
            if i % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws_data.cell(row=i + 1, column=col_idx).fill = PatternFill("solid", fgColor="F8FAFC")
        ws_data.freeze_panes = "A2"

        # ---- Sheet 3: สรุปรายหมวดหมู่ ----
        ws_cat = wb.create_sheet("สรุปรายหมวดหมู่")
        ws_cat.append(["หมวดหมู่", "ประเภทรายการ", "ยอดรวม (บาท)"])
        ws_cat.column_dimensions["A"].width = 30
        ws_cat.column_dimensions["B"].width = 14
        ws_cat.column_dimensions["C"].width = 18
        for idx in range(1, 4):
            cell = ws_cat.cell(row=1, column=idx)
            cell.fill = HEADER_FILL
            cell.font = white_bold
        cat_totals: Dict[str, dict] = {}
        for r in rows:
            key = r["category"]
            entry = cat_totals.setdefault(key, {"type": r["type"], "total": 0.0})
            if r["type"] == "รายรับ":
                entry["total"] += r["income"]
            elif r["type"] == "รายจ่าย":
                entry["total"] += r["expense"]
        for idx, (name, info) in enumerate(sorted(cat_totals.items()), start=2):
            ws_cat.cell(row=idx, column=1, value=name)
            ws_cat.cell(row=idx, column=2, value=info["type"])
            cell = ws_cat.cell(row=idx, column=3, value=round(info["total"], 2))
            cell.number_format = "#,##0.00 \"฿\""

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output