"""
Integration tests for FinanceService.export_transactions_excel (services/finance_service.py)

ครอบคลุม:
  - ช่วงเวลาที่รองรับ: start_date+end_date / month+year / ไม่ระบุเลย (ทั้งหมด)
  - การรวมขาโอนเงิน (transfer_group_id) → 1 แถว เพื่อไม่ให้รายรับ/รายจ่ายเกินจริง
  - การคัดคอลัมน์: ไม่มี system values (slip_image_url / deleted_at / transfer_group_id)
  - ความถูกต้องของตัวเลขใน Excel (รายรับ/รายจ่าย/ยอดคงเหลือรายบัญชี)
  - RBAC: สมาชิกห้องอ่านได้, สมาชิกต่างห้อง/คนนอกโดน ForbiddenError

Pattern ตาม docs/rules/testing.md: service-level เรียกตรง, deep DB verification,
ไม่แตะ Redis (FinanceService ไม่ publish)
"""
import io
import random
import string
import uuid
from datetime import date

import openpyxl
import pytest

from core.exceptions import ForbiddenError
from models.finance_schemas import FinanceExportRequest, TransactionCreate, TransferCreate
from services.finance_service import FinanceService

pytestmark = pytest.mark.asyncio


# === Fixtures & Setup (ลอก pattern จาก test_finance.py) ===


async def _insert_user(pool, *, first_name="Test", last_name="User") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO users (first_name, last_name, username)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            first_name, last_name, f"u{uuid.uuid4().hex[:12]}",
        )


async def _insert_room(pool, owner_id: int, room_name="Test Room") -> int:
    async with pool.acquire() as conn:
        while True:
            code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE room_code = $1", code):
                break
        room_id = await conn.fetchval(
            """
            INSERT INTO rooms (room_name, room_code, owner_id)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            room_name, code, owner_id,
        )
        await conn.execute(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, 0, 'president', 'active', TRUE, $3::jsonb)
            """,
            room_id, owner_id, '["all"]',
        )
        return room_id


async def _insert_student(pool, room_id: int, user_id: int, student_no: int, *, status="active") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, $3, 'student', $4, FALSE, '[]'::jsonb)
            RETURNING id
            """,
            room_id, user_id, student_no, status,
        )


async def _insert_finance_account(pool, room_id: int, account_name="กระเป๋ากลาง", balance=0.0) -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO finance_accounts (room_id, account_name, balance)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            room_id, account_name, balance,
        )


async def _insert_category(pool, room_id: int, category_name="ค่าอาหาร", category_type="expense") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO finance_categories (room_id, category_name, category_type)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            room_id, category_name, category_type,
        )


async def _set_created_at(pool, tx_id: int, dt) -> None:
    """คอลัมน์ TIMESTAMP ต้องส่ง datetime object (asyncpg ปฏิเสธ string) — ตาม docs/skills.md"""
    async with pool.acquire() as conn:
        await conn.execute("UPDATE finance_transactions SET created_at = $2 WHERE id = $1", tx_id, dt)


def _load_data_sheet(excel_file) -> list:
    """อ่าน Sheet 'ประวัติรายการ' กลับมาเป็น list[dict] (row แรกคือ header)."""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["ประวัติรายการ"]
    rows = list(ws.values)
    header = list(rows[0])
    return [dict(zip(header, row)) for row in rows[1:]]


def _load_summary_balances(excel_file) -> dict:
    """อ่าน Sheet 'สรุปยอด' → dict ของ บัญชี → ยอดคงเหลือ."""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["สรุปยอด"]
    values = list(ws.values)
    balances = {}
    in_table = False
    for row in values:
        if row[0] == "ยอดคงเหลือรายบัญชี":
            in_table = True
            continue
        if in_table and row[0] and row[0] not in ("บัญชี", "ยอดคงเหลือ (บาท)"):
            balances[row[0]] = row[1]
    return balances


# === Tests ===


async def test_export_empty_room_returns_valid_workbook(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner, room_name="ห้องเทส")

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
        room_id=room_id, user_id=owner,
    )
    assert isinstance(excel_file, io.BytesIO)

    wb = openpyxl.load_workbook(excel_file)
    # 3 แผ่นตามโครงสร้าง
    assert wb.sheetnames == ["สรุปยอด", "ประวัติรายการ", "สรุปรายหมวดหมู่"]

    ws_data = wb["ประวัติรายการ"]
    rows = list(ws_data.values)
    assert len(rows) == 1  # มีแค่ header
    header = rows[0]
    assert "วันที่" in header and "รายรับ (บาท)" in header and "รายจ่าย (บาท)" in header

    # สรุปยอดเป็นศูนย์ และไม่มีบัญชีใด (มีแค่ placeholder "(ไม่มีรายการในช่วงนี้)")
    assert _load_summary_balances(excel_file) == {"(ไม่มีรายการในช่วงนี้)": 0}


async def test_export_includes_income_and_expense_rows(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner, room_name="ห้องเทส")
    account_id = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=account_id, category_id=inc_cat, amount=100.0,
                              description="รับบริจาค", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=account_id, category_id=exp_cat, amount=30.0,
                              description="ซื้อของ", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
        room_id=room_id, user_id=owner,
    )
    rows = _load_data_sheet(excel_file)
    assert len(rows) == 2
    income_row = next(r for r in rows if r["รายรับ (บาท)"] == 100.0)
    expense_row = next(r for r in rows if r["รายจ่าย (บาท)"] == 30.0)
    assert income_row["หมวดหมู่"] == "เงินบริจาค"
    assert income_row["บัญชี"] == "กองกลาง"
    assert expense_row["หมวดหมู่"] == "ค่าอาหาร"
    assert expense_row["ผู้บันทึก"] == "Owner"

    # ต้องไม่มี system values หลุดออกมา
    assert "slip_image_url" not in income_row
    assert "transfer_group_id" not in income_row
    assert "deleted_at" not in income_row


async def test_export_filters_by_date_range(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    from datetime import datetime
    account_id = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=account_id, category_id=cat, amount=100.0,
                              description="ม.ค.", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=account_id, category_id=cat, amount=200.0,
                              description="ก.พ.", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        ids = await conn.fetch("SELECT id FROM finance_transactions WHERE room_id = $1 ORDER BY id", room_id)
    await _set_created_at(db_pool, ids[0]["id"], datetime(2026, 1, 15, 10))
    await _set_created_at(db_pool, ids[1]["id"], datetime(2026, 2, 15, 10))

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(start_date=date(2026, 1, 1), end_date=date(2026, 1, 31)),
        client_source="test", actor_identifier="test", room_id=room_id, user_id=owner,
    )
    rows = _load_data_sheet(excel_file)
    assert len(rows) == 1
    assert "ม.ค." in rows[0]["รายการ"]


async def test_export_month_year_filter(db_pool):
    from datetime import datetime
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    account_id = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    async def _add(amount, desc):
        return await FinanceService.add_transaction(
            pool=db_pool,
            req=TransactionCreate(account_id=account_id, category_id=cat, amount=amount,
                                  description=desc, transaction_type="income", user_name="Owner"),
            user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
        )

    await _add(100.0, "ม.ค.")
    await _add(200.0, "ก.พ.")
    async with db_pool.acquire() as conn:
        ids = await conn.fetch("SELECT id FROM finance_transactions WHERE room_id = $1 ORDER BY id", room_id)
    await _set_created_at(db_pool, ids[0]["id"], datetime(2026, 1, 10, 9))
    await _set_created_at(db_pool, ids[1]["id"], datetime(2026, 2, 10, 9))

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(month=2, year=2026),
        client_source="test", actor_identifier="test", room_id=room_id, user_id=owner,
    )
    rows = _load_data_sheet(excel_file)
    assert len(rows) == 1
    assert "ก.พ." in rows[0]["รายการ"]


async def test_export_consolidates_transfer_legs_into_one_row(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    from_acc = await _insert_finance_account(db_pool, room_id, "บัญชีหลัก", 1000.0)
    to_acc = await _insert_finance_account(db_pool, room_id, "บัญชีย่อย", 0.0)

    # โอน 400 จากหลัก → ย่อย (สร้าง 2 ขาใน finance_transactions)
    await FinanceService.transfer_money(
        pool=db_pool,
        req=TransferCreate(from_account_id=from_acc, to_account_id=to_acc, amount=400.0,
                           description="ฝากสำรอง", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    # ยืนยันว่ามี 2 ขาจริงใน DB
    async with db_pool.acquire() as conn:
        leg_count = await conn.fetchval(
            "SELECT COUNT(*) FROM finance_transactions WHERE room_id = $1 AND transfer_group_id IS NOT NULL",
            room_id,
        )
    assert leg_count == 2

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
        room_id=room_id, user_id=owner,
    )
    rows = _load_data_sheet(excel_file)
    # ต้องเหลือ 1 แถว และแสดงเป็นรายจ่าย (เงินออกจากบัญชีต้นทาง)
    assert len(rows) == 1
    assert rows[0]["ประเภท"] == "โอนเงินระหว่างบัญชี"
    assert rows[0]["รายจ่าย (บาท)"] == 400.0
    assert rows[0]["รายรับ (บาท)"] == 0.0
    # แถวชี้ไปที่บัญชีต้นทาง (บัญชีหลัก) ที่เงินออกจริง
    assert rows[0]["บัญชี"] == "บัญชีหลัก"
    # คำอธิบายถูกตัด 'โอนออก:' ออก
    assert "ฝากสำรอง" in rows[0]["รายการ"]


async def test_export_summary_sheet_has_correct_totals(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    # acc1 เปิดด้วยยอด seed 200 (ยอดสะสมก่อนหน้างวด) — ต้องโชว์ในสรุปยอด
    acc1 = await _insert_finance_account(db_pool, room_id, "กองกลาง", 200.0)
    acc2 = await _insert_finance_account(db_pool, room_id, "เงินสด", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    async def _add(acc, cat, amount, ttype, desc):
        return await FinanceService.add_transaction(
            pool=db_pool,
            req=TransactionCreate(account_id=acc, category_id=cat, amount=amount,
                                  description=desc, transaction_type=ttype, user_name="Owner"),
            user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
        )

    await _add(acc1, inc_cat, 500.0, "income", "บริจาค 500")
    await _add(acc2, inc_cat, 300.0, "income", "บริจาค 300")
    await _add(acc1, exp_cat, 200.0, "expense", "ซื้อของ 200")

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
        room_id=room_id, user_id=owner,
    )
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["สรุปยอด"]
    values = list(ws.values)
    by_label = {row[0]: row[1] for row in values if row[0] and isinstance(row[1], (int, float))}
    # ยอดรวมคำนวณจากรายการในช่วง (เฉพาะที่ผ่าน filter)
    assert by_label["รายรับรวม"] == 800.0
    assert by_label["รายจ่ายรวม"] == 200.0
    assert by_label["คงเหลือ (รายรับ − รายจ่าย)"] == 600.0
    # ยอดคงเหลือรายบัญชี = ยอดจริงจาก finance_accounts (รวม seed 200)
    balances = _load_summary_balances(excel_file)
    assert balances["กองกลาง"] == 500.0   # 200 seed + 500 − 200
    assert balances["เงินสด"] == 300.0


async def test_export_soft_deleted_transactions_excluded(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    account_id = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=account_id, category_id=cat, amount=100.0,
                              description="รายการจริง", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("SELECT id FROM finance_transactions WHERE room_id = $1", room_id)
        live_id = rows[0]["id"]
    # สร้างรายการที่ soft-delete แล้ว
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=account_id, category_id=cat, amount=999.0,
                              description="ถูกลบ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        del_id = await conn.fetchval(
            "SELECT id FROM finance_transactions WHERE room_id = $1 AND id != $2", room_id, live_id
        )
        await conn.execute("UPDATE finance_transactions SET deleted_at = NOW() WHERE id = $1", del_id)

    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
        room_id=room_id, user_id=owner,
    )
    rows = _load_data_sheet(excel_file)
    assert len(rows) == 1
    assert "ถูกลบ" not in rows[0]["รายการ"]


async def test_export_non_member_forbidden(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    outsider = await _insert_user(db_pool, first_name="Outsider", last_name="User")

    with pytest.raises(ForbiddenError):
        await FinanceService.export_transactions_excel(
            pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
            room_id=room_id, user_id=outsider,
        )


async def test_export_cross_room_member_forbidden(db_pool):
    owner_a = await _insert_user(db_pool, first_name="Admin", last_name="A")
    room_a = await _insert_room(db_pool, owner_a, room_name="ห้อง A")
    owner_b = await _insert_user(db_pool, first_name="Admin", last_name="B")
    room_b = await _insert_room(db_pool, owner_b, room_name="ห้อง B")
    member_a = await _insert_user(db_pool, first_name="Member", last_name="A")
    await _insert_student(db_pool, room_a, member_a, 1, status="active")

    # สมาชิกห้อง A พยายาม export ห้อง B → ForbiddenError
    with pytest.raises(ForbiddenError):
        await FinanceService.export_transactions_excel(
            pool=db_pool, req=FinanceExportRequest(), client_source="test", actor_identifier="test",
            room_id=room_b, user_id=member_a,
        )


async def test_export_start_after_end_raises(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)

    with pytest.raises(ValueError):
        await FinanceService.export_transactions_excel(
            pool=db_pool, req=FinanceExportRequest(start_date=date(2026, 2, 1), end_date=date(2026, 1, 1)),
            client_source="test", actor_identifier="test", room_id=room_id, user_id=owner,
        )
