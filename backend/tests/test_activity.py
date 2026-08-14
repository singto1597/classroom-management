"""
Activity & Role Management — integration tests (service + HTTP).

ครอบคลุมตาม docs/rules/testing.md:
  - Real Postgres (test_db) + clean_database fixture (state isolation)
  - Randomized IDs — ห้าม hardcode
  - Deep DB verification: หลัง HTTP assert สถานะ → query DB ตรง ๆ พิสูจน์ mutation
  - Mock ActionService.notify_new_activity / aioredis (กันแตะ Redis จริง)
  - JSONB parse quirk: asyncpg คืน str หรือ dict ตามเวอร์ชัน → normalize เสมอ
"""
import io
import json
import random
import string
import uuid
from datetime import date

import openpyxl
import pytest
from unittest.mock import AsyncMock, patch

from core.config import settings
from services.action_service import ActionService
from services.activity_service import ActivityService

pytestmark = pytest.mark.asyncio


# === Fixtures & Setup (copy pattern จาก test_student.py) ===


async def _insert_user(pool, *, first_name="Test", last_name="User", email=None, discord_id=None) -> int:
    if email is None:
        email = f"u{uuid.uuid4().hex[:12]}@test.local"
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO users (email, first_name, last_name, username, discord_id)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING id
            """,
            email, first_name, last_name, f"user_{uuid.uuid4().hex[:8]}", discord_id,
        )


async def _insert_room(pool, owner_id: int, room_name="Test Room", server_id=None) -> int:
    async with pool.acquire() as conn:
        while True:
            code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE room_code = $1", code):
                break
        room_id = await conn.fetchval(
            """
            INSERT INTO rooms (room_name, room_code, owner_id, server_id)
            VALUES ($1, $2, $3, $4)
            RETURNING id
            """,
            room_name, code, owner_id, server_id,
        )
        # ผู้สร้างห้องเป็น admin ทันที
        await conn.execute(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, 0, 'president', 'active', TRUE, $3::jsonb)
            """,
            room_id, owner_id, '["all"]',
        )
        return room_id


async def _insert_student(pool, room_id: int, user_id: int, student_no: int, *, status="active", permissions="[]") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, $3, 'student', $4, FALSE, $5::jsonb)
            RETURNING id
            """,
            room_id, user_id, student_no, status, permissions,
        )


def _parse_metadata(raw) -> dict:
    """asyncpg คืน JSONB เป็น str/dict ตามเวอร์ชัน → normalize"""
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {}
    return {}


def _make_web_headers(user_id: int) -> dict:
    from jose import jwt
    token = jwt.encode(
        {"user_id": user_id, "exp": 9999999999},
        settings.JWT_SECRET,
        algorithm=settings.JWT_ALGORITHM,
    )
    return {"Authorization": f"Bearer {token}"}


def _activity_api(room_id: int, path: str = "", target_type: str = "room") -> str:
    return f"/api/classroom/{room_id}/activities{path}?target_type={target_type}"


async def _create_activity_http(client, room_id, owner_id, *, title="กีฬาสี", activity_date="2026-10-15",
                                base_hours=8.0, status="upcoming", metadata=None, participants=None, **kw):
    payload = {
        "title": title,
        "activity_date": activity_date,
        "base_hours": base_hours,
        "status": status,
        "metadata": metadata or {},
        "participants": participants or [],
        "user_name": "ผู้ดูแล",
        **kw,
    }
    return client.post(_activity_api(room_id), json=payload, headers=_make_web_headers(owner_id))


# ================================================================
# 🧩 Service-level: JSONB parsing + create atomicity
# ================================================================

async def test_create_activity_persists_jsonb_metadata_and_participants(db_pool):
    """สร้างกิจกรรม + ผู้เข้าร่วม (executemany) → metadata เก็บครบ + ผู้เข้าร่วม insert ครบ
    ⚠️ ห้องต้องมี server_id (ผูก Discord) ถึงจะ publish NEW_ACTIVITY — เหมือน notify_new_task"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    server_id = random.randint(1_000_000, 9_999_999)
    room_id = await _insert_room(db_pool, owner, server_id=server_id)
    s1 = await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="สมชาย", last_name="ใจดี"), 1)
    s2 = await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="สมหญิง", last_name="สวยงาม"), 2)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock) as mock_notify:
        result = await ActivityService.create_activity(
            pool=db_pool,
            title="กีฬาสีประจำปี 2026",
            description="งานกีฬาสี",
            activity_date=date(2026, 10, 15),
            base_hours=8.0,
            status="upcoming",
            metadata={"tags": ["กีฬา"], "location_url": "https://maps.example.com/x", "agenda": ["08:00 เปิดงาน", "10:00 แข่ง"]},
            participants=[
                {"student_no": 1, "role_type": "leader", "role_detail": "ถือป้าย", "metadata": {"bus_number": "A1"}},
                {"student_no": 2, "role_type": "staff", "role_detail": "สวัสดิการ", "metadata": {"shirt_size": "M"}},
            ],
            user_name="ผู้ดูแล",
            client_source="WEB_APP",
            actor_identifier="user_id:1",
            room_id=room_id,
            actor_user_id=owner,
        )

    activity_id = result["activity_id"]
    async with db_pool.acquire() as conn:
        act = await conn.fetchrow("SELECT * FROM activities WHERE id = $1", activity_id)
        assert act["title"] == "กีฬาสีประจำปี 2026"
        assert act["status"] == "upcoming"
        meta = _parse_metadata(act["metadata"])
        assert meta["tags"] == ["กีฬา"]
        assert meta["location_url"].startswith("https://")

        parts = await conn.fetch(
            "SELECT * FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL ORDER BY student_id",
            activity_id,
        )
        assert len(parts) == 2
        assert parts[0]["role_type"] == "leader"
        assert parts[0]["role_detail"] == "ถือป้าย"
        assert _parse_metadata(parts[0]["metadata"])["bus_number"] == "A1"
        assert parts[1]["role_type"] == "staff"
        assert _parse_metadata(parts[1]["metadata"])["shirt_size"] == "M"

    # 📢 publish NEW_ACTIVITY หลัง commit (ต้องเป็นห้องที่ผูก Discord แล้ว)
    mock_notify.assert_awaited_once()
    kwargs = mock_notify.await_args.kwargs
    assert kwargs["title"] == "กีฬาสีประจำปี 2026"
    assert kwargs["participant_count"] == 2
    assert kwargs["metadata"]["tags"] == ["กีฬา"]


async def test_create_activity_audit_log_written(db_pool):
    """ทุก mutation ต้องเขียน audit_logs ใน transaction เดียวกัน"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ค่ายลูกเสือ", activity_date=date(2026, 11, 1),
            base_hours=12.0, status="upcoming", metadata={"tags": ["ค่าย"]},
            participants=[{"student_no": 1}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    async with db_pool.acquire() as conn:
        log = await conn.fetchrow(
            "SELECT * FROM audit_logs WHERE entity_type = 'ACTIVITY' AND entity_id = $1 ORDER BY created_at DESC LIMIT 1",
            str(activity_id),
        )
        assert log is not None
        assert log["action"] == "CREATE"
        new_vals = json.loads(log["new_values"])
        assert new_vals["title"] == "ค่ายลูกเสือ"


async def test_create_activity_rbac_forbidden_for_plain_member(db_pool):
    """สมาชิกธรรมดา (ไม่ใช่ admin / ไม่มี MANAGE_ACTIVITIES) → ForbiddenError"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Plain", last_name="Member")
    await _insert_student(db_pool, room_id, member, 5)

    from core.exceptions import ForbiddenError
    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock), pytest.raises(ForbiddenError):
        await ActivityService.create_activity(
            pool=db_pool, title="กิจกรรม", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={}, participants=[],
            user_name="Plain", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=member,
        )


async def test_create_activity_duplicate_participant_blocked(db_pool):
    """เลขที่ซ้ำในรายชื่อผู้เข้าร่วม → ValidationError + rollback (ไม่มี activity หลุดออกมา)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    from core.exceptions import ValidationError
    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock), pytest.raises(ValidationError):
        await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={},
            participants=[{"student_no": 1}, {"student_no": 1}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )

    # 🚨 Atomic: transaction ต้อง rollback ทั้ง activity + participants
    async with db_pool.acquire() as conn:
        count = await conn.fetchval("SELECT COUNT(*) FROM activities WHERE room_id = $1", room_id)
        assert count == 0


async def test_create_activity_pending_student_rejected(db_pool):
    """เพิ่มผู้เข้าร่วมที่เป็น pending (ยังไม่ active) → StudentNotFoundError"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    pending_user = await _insert_user(db_pool, first_name="รอ", last_name="อนุมัติ")
    await _insert_student(db_pool, room_id, pending_user, 3, status="pending")

    from core.exceptions import StudentNotFoundError
    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock), pytest.raises(StudentNotFoundError):
        await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={},
            participants=[{"student_no": 3}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )


async def test_update_activity_merges_metadata(db_pool):
    """PATCH metadata ต้อง merge กับของเดิม ไม่ทับคีย์ที่ไม่ได้ส่ง"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming",
            metadata={"tags": ["กีฬา"], "location_url": "https://old.example.com"},
            participants=[], user_name="ผู้ดูแล", client_source="WEB_APP",
            actor_identifier="user_id:1", room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    updated = await ActivityService.update_activity(
        pool=db_pool, activity_id=activity_id,
        update_data={"metadata": {"location_url": "https://new.example.com"}},
        user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
        room_id=room_id, actor_user_id=owner,
    )
    assert updated["metadata"]["tags"] == ["กีฬา"]          # merge เก็บคีย์เก่า
    assert updated["metadata"]["location_url"] == "https://new.example.com"


async def test_delete_activity_soft_deletes_participants(db_pool):
    """Soft delete กิจกรรม → deleted_at ถูกตั้ง + participants ถูก soft delete ด้วย"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={},
            participants=[{"student_no": 1}], user_name="ผู้ดูแล",
            client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    await ActivityService.delete_activity(
        pool=db_pool, activity_id=activity_id, user_name="ผู้ดูแล", user_id=owner,
        client_source="WEB_APP", actor_identifier="user_id:1", room_id=room_id,
    )

    async with db_pool.acquire() as conn:
        act = await conn.fetchrow("SELECT deleted_at FROM activities WHERE id = $1", activity_id)
        assert act["deleted_at"] is not None
        parts = await conn.fetch("SELECT deleted_at FROM activity_participants WHERE activity_id = $1", activity_id)
        assert all(p["deleted_at"] is not None for p in parts)


async def test_add_participant_revives_soft_deleted(db_pool):
    """เพิ่มผู้เข้าร่วมที่เคยถูกลบ (soft delete) → กู้กลับมา ไม่ชน UNIQUE partial index"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={},
            participants=[{"student_no": 1}], user_name="ผู้ดูแล",
            client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    # ลบ participant
    async with db_pool.acquire() as conn:
        part_id = await conn.fetchval("SELECT id FROM activity_participants WHERE activity_id = $1", activity_id)
    await ActivityService.remove_participant(
        pool=db_pool, activity_id=activity_id, participant_id=part_id,
        user_name="ผู้ดูแล", user_id=owner, client_source="WEB_APP",
        actor_identifier="user_id:1", room_id=room_id,
    )

    # เพิ่มคนเดิมกลับ → ไม่ error (revive)
    res = await ActivityService.add_participant(
        pool=db_pool, activity_id=activity_id, student_no=1, user_name="ผู้ดูแล",
        client_source="WEB_APP", actor_identifier="user_id:1",
        room_id=room_id, actor_user_id=owner, role_detail="ใหม่",
    )
    assert res["status"] == "success"

    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM activity_participants WHERE id = $1", part_id)
        assert row["deleted_at"] is None
        assert row["role_detail"] == "ใหม่"


async def test_get_student_activity_roles_returns_bus_number(db_pool):
    """get_student_activity_roles → คืน role_detail + bus_number จาก metadata (ใช้กับบอท /my_roles)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    student_user = await _insert_user(db_pool, first_name="นัก", last_name="เรียน", discord_id=777001)
    student_id = await _insert_student(db_pool, room_id, student_user, 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ทัศนศึกษา", activity_date=date(2026, 9, 20),
            base_hours=6.0, status="upcoming", metadata={"location_url": "https://goo.gl/maps/abc"},
            participants=[{"student_no": 1, "role_type": "staff", "role_detail": "นับคน", "metadata": {"bus_number": "B2"}}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    roles = await ActivityService.get_student_activity_roles(
        pool=db_pool, user_id=student_user, client_source="WEB_APP",
        actor_identifier="user_id:1", room_id=room_id,
    )
    assert len(roles) == 1
    assert roles[0]["activity_id"] == activity_id
    assert roles[0]["role_detail"] == "นับคน"
    assert roles[0]["participant_metadata"]["bus_number"] == "B2"


async def test_export_activity_excel_extracts_metadata_columns(db_pool):
    """Excel export → เอาคีย์ (เบอร์รถบัส) ออกมาเป็นคอลัมน์จริง — bus_number จาก metadata"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    stu_user = await _insert_user(db_pool, first_name="สมชาย", last_name="ใจดี")
    # 🌟 shirt_size เป็น Type A → ตั้งค่าในโปรไฟล์ users (ไม่ใช่ metadata)
    await _insert_student_with_profile(db_pool, room_id, stu_user, 1, shirt_size="L")

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ไปทัศนศึกษา", activity_date=date(2026, 9, 20),
            base_hours=6.0, status="upcoming", metadata={"tags": ["ทริป"]},
            participants=[{"student_no": 1, "metadata": {"bus_number": "B2"}}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    excel = await ActivityService.export_activity_excel(
        pool=db_pool, activity_id=activity_id, metadata_keys=["bus_number", "shirt_size"],
        user_name="ผู้ดูแล", user_id=owner, client_source="WEB_APP",
        actor_identifier="user_id:1", room_id=room_id,
    )
    wb = openpyxl.load_workbook(io.BytesIO(excel.getvalue()))
    assert wb.sheetnames == ["สรุป", "รายชื่อผู้เข้าร่วม"]
    ws = wb["รายชื่อผู้เข้าร่วม"]
    header = list(ws.values)[0]
    assert "หมายเลขรถบัส" in header
    assert "ไซส์เสื้อ" in header
    # data row: ชื่อ + bus_number (metadata) + shirt_size (Type A จากโปรไฟล์)
    rows = list(ws.values)
    assert len(rows) == 2  # header + 1 participant
    data_row = rows[1]
    idx = header.index("หมายเลขรถบัส")
    assert data_row[idx] == "B2"
    idx_shirt = header.index("ไซส์เสื้อ")
    assert data_row[idx_shirt] == "L"


# ================================================================
# 🌐 HTTP-level: auth + status-code mapping
# ================================================================

async def test_http_create_activity_200_and_db_verification(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        resp = await _create_activity_http(
            client, room_id, owner,
            title="กีฬาสี",
            metadata={"tags": ["กีฬา"], "location_url": "https://maps.example.com"},
            participants=[{"student_no": 1, "role_detail": "ถือป้าย", "metadata": {"bus_number": "A1"}}],
        )
    assert resp.status_code == 200, resp.text

    # 🚨 Deep DB verification
    async with db_pool.acquire() as conn:
        act = await conn.fetchrow("SELECT * FROM activities WHERE room_id = $1", room_id)
        assert act["title"] == "กีฬาสี"
        assert _parse_metadata(act["metadata"])["tags"] == ["กีฬา"]
        part = await conn.fetchrow("SELECT * FROM activity_participants WHERE activity_id = $1", act["id"])
        assert part["role_detail"] == "ถือป้าย"
        assert _parse_metadata(part["metadata"])["bus_number"] == "A1"


async def test_http_create_activity_requires_auth(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    resp = client.post(
        _activity_api(room_id),
        json={"title": "x", "activity_date": "2026-10-15", "user_name": "x"},
    )
    assert resp.status_code == 401


async def test_http_create_activity_member_forbidden(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Plain", last_name="Member")
    await _insert_student(db_pool, room_id, member, 5)

    resp = await _create_activity_http(client, room_id, member)
    assert resp.status_code == 403  # ForbiddenError → 403


async def test_http_list_activities_member_can_view(client, db_pool):
    """อ่าน (GET) — สมาชิกทุกคนดูได้ (require_member ไม่ใช่ MANAGE)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Plain", last_name="Member")
    await _insert_student(db_pool, room_id, member, 5)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        await _create_activity_http(client, room_id, owner, title="งาน 1")

    resp = client.get(_activity_api(room_id), headers=_make_web_headers(member))
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1
    assert data[0]["title"] == "งาน 1"
    assert "participant_count" in data[0]


async def test_http_get_activity_me_roles_route_not_shadowed(client, db_pool):
    """GET /activities/me/roles ต้องไม่โดน /activities/{activity_id} กลืน (route ordering)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    student_user = await _insert_user(db_pool, first_name="นัก", last_name="เรียน")
    await _insert_student(db_pool, room_id, student_user, 2)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        await _create_activity_http(
            client, room_id, owner,
            participants=[{"student_no": 2, "role_detail": "สวัสดิการ"}],
        )

    resp = client.get(
        _activity_api(room_id, "/me/roles"),
        headers=_make_web_headers(student_user),
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert len(data) == 1
    assert data[0]["role_detail"] == "สวัสดิการ"


async def test_http_export_activity_excel(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        create_resp = await _create_activity_http(
            client, room_id, owner,
            participants=[{"student_no": 1, "metadata": {"bus_number": "C3"}}],
        )
        activity_id = create_resp.json()["message"].split("ID: ")[1].rstrip(")")

    resp = client.post(
        _activity_api(room_id, "/export"),
        json={"activity_id": int(activity_id), "metadata_keys": ["bus_number"], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["รายชื่อผู้เข้าร่วม"]
    header = list(ws.values)[0]
    assert "หมายเลขรถบัส" in header
    idx = header.index("หมายเลขรถบัส")
    assert list(ws.values)[1][idx] == "C3"


async def test_http_export_activity_requires_manage_permission(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Plain", last_name="Member")
    await _insert_student(db_pool, room_id, member, 5)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        create_resp = await _create_activity_http(
            client, room_id, owner, participants=[{"student_no": 5}],
        )
        activity_id = create_resp.json()["message"].split("ID: ")[1].rstrip(")")

    resp = client.post(
        _activity_api(room_id, "/export"),
        json={"activity_id": int(activity_id), "metadata_keys": [], "user_name": "Plain"},
        headers=_make_web_headers(member),
    )
    assert resp.status_code == 403


async def test_http_bot_path_create_activity(client, db_pool):
    """บอท path (X-API-Key + X-Discord-Id) สร้างกิจกรรมได้ — resolve target เป็น server_id"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner", discord_id=778901)
    server_id = random.randint(1_000_000, 9_999_999)
    room_id = await _insert_room(db_pool, owner, server_id=server_id)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    bot_headers = {"X-API-Key": settings.API_KEY, "X-Discord-Id": str(778901)}
    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock) as mock_notify:
        resp = client.post(
            _activity_api(server_id, target_type="server"),
            json={
                "title": "กิจกรรมจากบอท",
                "activity_date": "2026-12-05",
                "base_hours": 4.0,
                "metadata": {},
                "participants": [{"student_no": 1}],
                "user_name": "บอท",
            },
            headers=bot_headers,
        )
    assert resp.status_code == 200, resp.text
    # publish ไป server_id ที่ resolve ได้
    mock_notify.assert_awaited_once()
    assert mock_notify.await_args.kwargs["title"] == "กิจกรรมจากบอท"

    async with db_pool.acquire() as conn:
        count = await conn.fetchval("SELECT COUNT(*) FROM activities WHERE room_id = $1", room_id)
        assert count == 1


# ================================================================
# 🌟 Dynamic Smart Forms — Field Selector (required_fields) + Type A profile JOIN
# ================================================================

async def _insert_student_with_profile(pool, room_id: int, user_id: int, student_no: int, *, blood_group="B", shirt_size="M",
                                       food_allergy="กุ้ง", congenital_disease=None, phone="0800000001",
                                       parent_phone="0800000002") -> int:
    """สร้าง student + อัปเดตโปรไฟล์ users ให้ครบ Type A (จำลอง Field Selector ติ๊กแล้ว)"""
    async with pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE users SET blood_group = $1, shirt_size = $2, food_allergy = $3,
                congenital_disease = $4, phone_number = $5, phone_number_parent = $6
            WHERE id = $7
            """,
            blood_group, shirt_size, food_allergy, congenital_disease, phone, parent_phone, user_id,
        )
    return await _insert_student(pool, room_id, user_id, student_no)


async def test_create_activity_stores_required_fields_in_metadata(db_pool):
    """Field Selector → activities.metadata.required_fields เก็บ Array คีย์ที่ติ๊ก"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="กีฬาสี", activity_date=date(2026, 10, 15),
            base_hours=8.0, status="upcoming",
            metadata={"required_fields": ["bus_number", "shirt_size", "food_allergy"]},
            participants=[{"student_no": 1}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT metadata FROM activities WHERE id = $1", activity_id)
        meta = _parse_metadata(row["metadata"])
        assert meta["required_fields"] == ["bus_number", "shirt_size", "food_allergy"]


async def test_get_activity_returns_profile_fields_joined(db_pool):
    """GET participants → JOIN users → คืน Type A profile fields (blood_group, shirt_size, ...)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    stu_user = await _insert_user(db_pool, first_name="สมชาย", last_name="ใจดี")
    await _insert_student_with_profile(db_pool, room_id, stu_user, 1, blood_group="B", shirt_size="L",
                                       food_allergy="ถั่ว", phone="0812345678")

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ค่าย", activity_date=date(2026, 11, 1),
            base_hours=6.0, status="upcoming", metadata={},
            participants=[{"student_no": 1}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    activity = await ActivityService.get_activity(
        pool=db_pool, activity_id=activity_id,
        client_source="WEB_APP", actor_identifier="user_id:1", room_id=room_id,
    )
    assert len(activity["participants"]) == 1
    p = activity["participants"][0]
    # 🌟 Type A — มาแบบ JOIN จาก users ไม่ต้องอยู่ใน metadata
    assert p["blood_group"] == "B"
    assert p["shirt_size"] == "L"
    assert p["food_allergy"] == "ถั่ว"
    assert p["phone_number"] == "0812345678"
    # metadata ยังว่าง (ไม่ได้บันทึกซ้ำ)
    assert "blood_group" not in p["metadata"]


async def test_export_dynamic_uses_required_fields_for_headers(db_pool):
    """Export Dynamic — header มาจาก required_fields ของกิจกรรม (ไม่ต้องส่ง metadata_keys)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    stu_user = await _insert_user(db_pool, first_name="สมชาย", last_name="ใจดี")
    await _insert_student_with_profile(db_pool, room_id, stu_user, 1, blood_group="B", shirt_size="L",
                                       food_allergy="กุ้ง", phone="0800000001", parent_phone="0800000002")

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ไปทัศนศึกษา", activity_date=date(2026, 9, 20),
            base_hours=6.0, status="upcoming",
            metadata={"required_fields": ["bus_number", "shirt_size", "food_allergy", "phone_number"]},
            participants=[{"student_no": 1, "metadata": {"bus_number": "B2"}}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    # 🔑 ไม่ส่ง metadata_keys (backend ต้องอ่าน required_fields เอง) — backward compat + dynamic
    excel = await ActivityService.export_activity_excel(
        pool=db_pool, activity_id=activity_id, metadata_keys=[],
        user_name="ผู้ดูแล", user_id=owner, client_source="WEB_APP",
        actor_identifier="user_id:1", room_id=room_id,
    )
    wb = openpyxl.load_workbook(io.BytesIO(excel.getvalue()))
    ws = wb["รายชื่อผู้เข้าร่วม"]
    header = list(ws.values)[0]
    assert "หมายเลขรถบัส" in header          # Type B จาก metadata
    assert "ไซส์เสื้อ" in header              # Type A จากโปรไฟล์
    assert "อาหารที่แพ้" in header
    assert "เบอร์โทรศัพท์นักเรียน" in header   # Type A (phone_number) — label ใหม่

    # 🚨 ต้องไม่มีคอลัมน์ขยะ (ฟิลด์ที่ไม่ได้เลือก)
    assert "ห้องพัก" not in header
    assert "กรุ๊ปเลือด" not in header

    rows = list(ws.values)
    data_row = rows[1]
    # Type B — อ่านจาก metadata
    assert data_row[header.index("หมายเลขรถบัส")] == "B2"
    # Type A — อ่านจากโปรไฟล์ (JOIN users) ไม่ใช่ metadata
    assert data_row[header.index("ไซส์เสื้อ")] == "L"
    assert data_row[header.index("อาหารที่แพ้")] == "กุ้ง"
    assert data_row[header.index("เบอร์โทรศัพท์นักเรียน")] == "0800000001"


async def test_export_dynamic_prefers_required_fields_over_metadata_keys(db_pool):
    """ถ้ามีทั้ง required_fields และ metadata_keys → ให้ required_fields ชนะ (Field Selector เป็น source of truth)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    stu_user = await _insert_user(db_pool, first_name="A", last_name="B")
    await _insert_student_with_profile(db_pool, room_id, stu_user, 1, shirt_size="S")

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming",
            metadata={"required_fields": ["shirt_size"]},
            participants=[{"student_no": 1}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    # ส่ง metadata_keys=["room_number"] แต่กิจกรรมเลือก shirt_size → header ต้องเป็น shirt_size
    excel = await ActivityService.export_activity_excel(
        pool=db_pool, activity_id=activity_id, metadata_keys=["room_number"],
        user_name="ผู้ดูแล", user_id=owner, client_source="WEB_APP",
        actor_identifier="user_id:1", room_id=room_id,
    )
    wb = openpyxl.load_workbook(io.BytesIO(excel.getvalue()))
    header = list(wb["รายชื่อผู้เข้าร่วม"].values)[0]
    assert "ไซส์เสื้อ" in header
    assert "ห้องพัก" not in header


async def test_export_is_paid_boolean_translated(db_pool):
    """is_paid (Boolean) → แปลงเป็น '✅ จ่ายแล้ว'/'⏳ ยังไม่จ่าย' ใน Excel"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ค่ายลูกเสือ", activity_date=date(2026, 11, 1),
            base_hours=12.0, status="upcoming",
            metadata={"required_fields": ["is_paid"]},
            participants=[{"student_no": 1, "metadata": {"is_paid": True}}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    excel = await ActivityService.export_activity_excel(
        pool=db_pool, activity_id=activity_id, metadata_keys=[],
        user_name="ผู้ดูแล", user_id=owner, client_source="WEB_APP",
        actor_identifier="user_id:1", room_id=room_id,
    )
    wb = openpyxl.load_workbook(io.BytesIO(excel.getvalue()))
    ws = wb["รายชื่อผู้เข้าร่วม"]
    header = list(ws.values)[0]
    assert "สถานะจ่ายเงินค่าค่าย" in header
    assert list(ws.values)[1][header.index("สถานะจ่ายเงินค่าค่าย")] == "✅ จ่ายแล้ว"


async def test_http_get_activity_returns_profile_fields(client, db_pool):
    """HTTP GET activity → participants มี Type A profile fields (web ใช้ render ตาราง Smart)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    stu_user = await _insert_user(db_pool, first_name="สมหญิง", last_name="สวยงาม")
    await _insert_student_with_profile(db_pool, room_id, stu_user, 3, blood_group="O", shirt_size="M",
                                       food_allergy="ถั่วลิสง")

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        create_resp = await _create_activity_http(
            client, room_id, owner,
            participants=[{"student_no": 3}],
        )
        activity_id = create_resp.json()["message"].split("ID: ")[1].rstrip(")")

    resp = client.get(_activity_api(room_id, f"/{activity_id}"), headers=_make_web_headers(owner))
    assert resp.status_code == 200
    data = resp.json()
    assert len(data["participants"]) == 1
    p = data["participants"][0]
    assert p["blood_group"] == "O"
    assert p["food_allergy"] == "ถั่วลิสง"
    assert p["shirt_size"] == "M"
    # Pydantic response_model ต้อง allow ฟิลด์ใหม่ (ไม่โดน strip)
    assert "phone_number" in p


# ================================================================
# 🎯 Batch Apply (คลุมดำตั้งค่า) — atomic metadata update
# ================================================================

async def test_batch_update_participants_merges_metadata(db_pool):
    """Batch Apply → metadata ถูก merge กับของเดิมของแต่ละคน (ไม่ทับคีย์เก่า)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    u1 = await _insert_user(db_pool, first_name="หนึ่ง", last_name="แรก")
    u2 = await _insert_user(db_pool, first_name="สอง", last_name="สอง")
    await _insert_student(db_pool, room_id, u1, 1)
    await _insert_student(db_pool, room_id, u2, 2)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="ค่าย", activity_date=date(2026, 11, 1),
            base_hours=6.0, status="upcoming", metadata={},
            participants=[
                {"student_no": 1, "metadata": {"room_number": "501"}},
                {"student_no": 2, "metadata": {"room_number": "502"}},
            ],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    async with db_pool.acquire() as conn:
        p1, p2 = await conn.fetch(
            "SELECT id, student_id FROM activity_participants WHERE activity_id = $1 ORDER BY student_id", activity_id
        )

    # 🎯 ตั้งค่ารถบัสให้ทั้ง 2 คน (merge → room_number เดิมยังอยู่)
    res = await ActivityService.batch_update_participants(
        pool=db_pool, activity_id=activity_id,
        items=[
            {"participant_id": p1["id"], "metadata": {"bus_number": "1"}},
            {"participant_id": p2["id"], "metadata": {"bus_number": "1"}},
        ],
        user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
        room_id=room_id, actor_user_id=owner,
    )
    assert res["status"] == "success"
    assert res["updated_count"] == 2

    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT metadata FROM activity_participants WHERE activity_id = $1 ORDER BY student_id", activity_id
        )
        for row in rows:
            meta = _parse_metadata(row["metadata"])
            assert meta["bus_number"] == "1"
            assert "room_number" in meta  # 🌟 merge ไม่ทับ


async def test_batch_update_participants_atomic_rollback_on_error(db_pool):
    """Batch Atomic — participant_id หลุด (ไม่ใช่ของกิจกรรม) → rollback ทั้งก้อน (ไม่มีใครโดนตั้งค่า)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="C", last_name="D"), 2)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={},
            participants=[{"student_no": 1}, {"student_no": 2}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    async with db_pool.acquire() as conn:
        p1 = await conn.fetchval(
            "SELECT ap.id FROM activity_participants ap JOIN students s ON ap.student_id = s.id "
            "WHERE ap.activity_id = $1 AND s.student_no = 1", activity_id)

    from core.exceptions import ParticipantNotFoundError
    with pytest.raises(ParticipantNotFoundError):
        await ActivityService.batch_update_participants(
            pool=db_pool, activity_id=activity_id,
            items=[
                {"participant_id": p1, "metadata": {"bus_number": "9"}},
                {"participant_id": 999999, "metadata": {"bus_number": "9"}},  # ไม่มี → rollback
            ],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )

    # 🚨 atomic: participant ที่ถูกต้องก็ต้องไม่โดนตั้งค่า
    async with db_pool.acquire() as conn:
        meta = await conn.fetchval(
            "SELECT metadata FROM activity_participants ap JOIN students s ON ap.student_id = s.id "
            "WHERE ap.activity_id = $1 AND s.student_no = 1", activity_id)
        assert "bus_number" not in _parse_metadata(meta)


async def test_batch_update_participants_rbac_forbidden(db_pool):
    """Batch Apply — สมาชิกธรรมดา (ไม่มี MANAGE_ACTIVITIES) → ForbiddenError"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Plain", last_name="Member")
    await _insert_student(db_pool, room_id, member, 5)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        result = await ActivityService.create_activity(
            pool=db_pool, title="งาน", activity_date=date(2026, 12, 1),
            base_hours=1.0, status="upcoming", metadata={},
            participants=[{"student_no": 5}],
            user_name="ผู้ดูแล", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=owner,
        )
        activity_id = result["activity_id"]

    from core.exceptions import ForbiddenError
    with pytest.raises(ForbiddenError):
        await ActivityService.batch_update_participants(
            pool=db_pool, activity_id=activity_id,
            items=[{"participant_id": 1, "metadata": {"bus_number": "1"}}],
            user_name="Plain", client_source="WEB_APP", actor_identifier="user_id:1",
            room_id=room_id, actor_user_id=member,
        )


async def test_http_batch_update_participants_route_not_shadowed(client, db_pool):
    """HTTP: PATCH /participants/batch ต้องไม่โดน /participants/{participant_id} กลืน (route ordering)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="A", last_name="B"), 1)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="C", last_name="D"), 2)

    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        create_resp = await _create_activity_http(
            client, room_id, owner,
            participants=[{"student_no": 1}, {"student_no": 2}],
        )
        activity_id = create_resp.json()["message"].split("ID: ")[1].rstrip(")")

    async with db_pool.acquire() as conn:
        p1 = await conn.fetchval(
            "SELECT ap.id FROM activity_participants ap JOIN students s ON ap.student_id = s.id "
            "WHERE ap.activity_id = $1 AND s.student_no = 1", int(activity_id))
        p2 = await conn.fetchval(
            "SELECT ap.id FROM activity_participants ap JOIN students s ON ap.student_id = s.id "
            "WHERE ap.activity_id = $1 AND s.student_no = 2", int(activity_id))

    resp = client.patch(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={
            "items": [
                {"participant_id": p1, "metadata": {"bus_number": "2"}},
                {"participant_id": p2, "metadata": {"bus_number": "2"}},
            ],
            "user_name": "ผู้ดูแล",
        },
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200, resp.text

    # 🚨 Deep DB verification — ทั้ง 2 คนมี bus_number="2"
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT ap.metadata FROM activity_participants ap JOIN students s ON ap.student_id = s.id "
            "WHERE ap.activity_id = $1 ORDER BY s.student_no", int(activity_id))
        for row in rows:
            assert _parse_metadata(row["metadata"])["bus_number"] == "2"
