"""
HTTP-layer integration tests for student_router.py POST /{target_id}/export.

ครอบคลุมการทำงานผ่าน HTTP (TestClient) ที่ชุดเทส service-level (test_student.py) ยังไม่จับ:
  - Web (JWT Bearer) auth path
  - Status-code mapping ของ router (401/403)
  - target_type=room resolution ผ่าน URL path (web ต้องส่ง ?target_type=room เสมอ)
  - Response headers: content-type .xlsx + Content-Disposition attachment
  - ไฟล์จริงเปิดอ่านได้: sheetnames ['สรุป', 'รายชื่อ'] + หัวคอลัมน์ภาษาไทย

Pattern ตาม docs/rules/testing.md: ใช้ client/db_pool fixtures จาก conftest.py,
randomized server_id, mock ActionService/aioredis เพื่อไม่แตะ Redis จริง
"""
import io
import random
import string
import uuid

import openpyxl
import pytest

from core.config import settings

pytestmark = pytest.mark.asyncio


# === Fixtures: HTTP client with real auth ===


async def _insert_user(pool, *, first_name="Test", last_name="User", username=None) -> int:
    if username is None:
        username = f"u{uuid.uuid4().hex[:12]}"
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO users (first_name, last_name, username)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            first_name, last_name, username,
        )


async def _insert_room(pool, owner_id: int, room_name="Test Room") -> int:
    async with pool.acquire() as conn:
        while True:
            code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE room_code = $1", code):
                break
        room_id = await conn.fetchval(
            """
            INSERT INTO rooms (room_name, room_code, owner_id)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            room_name, code, owner_id,
        )
        # ผู้สร้างห้องเป็น admin ทันที (เหมือน test_student.py)
        await conn.execute(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, 0, 'president', 'active', TRUE, $3::jsonb)
            """,
            room_id, owner_id, '["all"]',
        )
        return room_id


async def _insert_student(pool, room_id: int, user_id: int, student_no: int, *, status="active") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, $3, 'student', $4, FALSE, '[]'::jsonb)
            RETURNING id
            """,
            room_id, user_id, student_no, status,
        )


def _make_web_headers(user_id: int) -> dict:
    from jose import jwt
    token = jwt.encode(
        {"user_id": user_id, "exp": 9999999999},
        settings.JWT_SECRET,
        algorithm=settings.JWT_ALGORITHM,
    )
    return {"Authorization": f"Bearer {token}"}


def _room_api(target_id: int, path: str, target_type: str = "room") -> str:
    """สร้าง URL student API โดยผูก target_type ตามจริง (web=room)"""
    return f"/api/classroom/{target_id}{path}?target_type={target_type}"


# === Tests ===


async def test_web_export_students_excel_200(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Excel", last_name="Export")
    await _insert_student(db_pool, room_id, member, 2)

    resp = client.post(
        _room_api(room_id, "/export"),
        json={"fields": ["student_no", "first_name", "last_name"], "user_name": "Owner"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["สรุป", "รายชื่อ"]
    ws = wb["รายชื่อ"]
    rows = list(ws.values)
    assert rows[0] == ("เลขที่", "ชื่อจริง", "นามสกุล")
    assert len(rows) == 3  # header + owner + member


async def test_web_export_students_excel_requires_auth(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)

    resp = client.post(
        _room_api(room_id, "/export"),
        json={"fields": ["student_no"], "user_name": "Owner"},
    )
    assert resp.status_code == 401


async def test_web_export_students_excel_member_forbidden(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    member = await _insert_user(db_pool, first_name="Plain", last_name="Member")
    await _insert_student(db_pool, room_id, member, 5)

    resp = client.post(
        _room_api(room_id, "/export"),
        json={"fields": ["student_no"], "user_name": "Plain"},
        headers=_make_web_headers(member),
    )
    assert resp.status_code == 403  # router maps ForbiddenError → 403
