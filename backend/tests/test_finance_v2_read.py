"""
Integration tests for Phase 4 — Time-Based Routing & Double-Entry Read Models.

ครอบคลุม:
- [ROUTER] get_transactions / get_summary / export_transactions_excel ตัดสินใจ
  ถูกต้องตามช่วงเวลา (ก่อน CUTOFF_DATE → legacy, หลัง → v2)
- [DOUBLE-ENTRY] _get_transactions_v2 แปลง Dr/Cr กลับเป็น TransactionResponse
  (income / expense / transfer / opening_balance)
- [DOUBLE-ENTRY] _get_summary_v2 รวมยอดจาก ledger (Net Worth / Income / Expense / Breakdown)
- [DOUBLE-ENTRY] get_trial_balance (Dr = Cr → is_balanced)
- [DOUBLE-ENTRY] get_income_statement (Revenue − Expense = net_income)

ใช้ service โดยตรง (ไม่ผ่าน HTTP) ตาม convention ของ test_finance.py
และพึ่ง dual-write ของ add_transaction/transfer_money/confirm_payment ที่สร้าง journal ให้เอง
"""
import random
import string
import uuid
from datetime import date, datetime, timedelta

import pytest

from core.exceptions import ForbiddenError
from models.finance_schemas import (
    AccountCreate,
    CategoryCreate,
    FinanceExportRequest,
    TransactionCreate,
    TransferCreate,
)
from services.finance_service import FinanceService, CUTOFF_DATE

pytestmark = pytest.mark.asyncio

# วันที่เทียบเคียงสำหรับงวด "หลัง" cutoff (ต้องหลุดจากวันจริงของ CI เสมอ)
POST_CUTOFF = date(2026, 10, 1)
PRE_CUTOFF = date(2026, 1, 1)


# === Fixtures & Setup (ลอกจาก test_finance.py เพื่อ isolation) ===


async def _insert_user(pool, *, email=None, first_name="Test", last_name="User", username=None) -> int:
    if username is None:
        username = f"u{uuid.uuid4().hex[:12]}"
    async with pool.acquire() as conn:
        return await conn.fetchval(
            "INSERT INTO users (email, first_name, last_name, username) VALUES ($1, $2, $3, $4) RETURNING id",
            email, first_name, last_name, username,
        )


async def _insert_room(pool, owner_id: int, room_name="Test Room", server_id=None) -> int:
    async with pool.acquire() as conn:
        while True:
            code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE room_code = $1", code):
                break
        room_id = await conn.fetchval(
            "INSERT INTO rooms (room_name, room_code, owner_id, server_id) VALUES ($1, $2, $3, $4) RETURNING id",
            room_name, code, owner_id, server_id,
        )
        await conn.execute(
            "INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions) "
            "VALUES ($1, $2, 0, 'president', 'active', TRUE, $3::jsonb)",
            room_id, owner_id, '["all"]',
        )
        return room_id


async def _insert_finance_account(pool, room_id: int, account_name="กระเป๋ากลาง", balance=0.0) -> int:
    async with pool.acquire() as conn:
        # [DUAL-WRITE] สร้างทั้ง legacy account + accounting_ledgers (asset 1xxxx)
        # เพื่อให้ dual-write (add_transaction/transfer_money) และ opening balance หา ledger เจอ
        account_id = await conn.fetchval(
            "INSERT INTO finance_accounts (room_id, account_name, balance) VALUES ($1, $2, $3) RETURNING id",
            room_id, account_name, balance,
        )
        await conn.execute(
            """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, legacy_account_id, description)
               VALUES ($1, $2, $3, 'asset', $4, 'test')""",
            room_id, f"1{account_id:04d}", account_name, account_id,
        )
        return account_id


async def _insert_category(pool, room_id: int, category_name="เงินบริจาค", category_type="income") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            "INSERT INTO finance_categories (room_id, category_name, category_type) VALUES ($1, $2, $3) RETURNING id",
            room_id, category_name, category_type,
        )


async def _set_tx_dates(pool, room_id: int, dates: dict):
    """ตั้ง created_at (legacy) + transaction_date (journal) ของห้องให้ตรงกับ dates.
    dates: { "mapping_key": datetime } — ใช้เรียงตาม legacy id ลำดับที่ insert"""
    async with pool.acquire() as conn:
        # ปรับ legacy finance_transactions (เรียงตาม id)
        rows = await conn.fetch(
            "SELECT id FROM finance_transactions WHERE room_id = $1 ORDER BY id", room_id
        )
        for i, r in enumerate(rows):
            dt = dates.get(f"legacy_{i}", None)
            if dt is not None:
                await conn.execute("UPDATE finance_transactions SET created_at = $2 WHERE id = $1", r["id"], dt)
        # ปรับ journal_entries (เรียงตาม created_at ของหัวบิล — ใช้ id uuid)
        jrows = await conn.fetch(
            """SELECT id, created_at FROM journal_entries
               WHERE room_id = $1 ORDER BY created_at""", room_id
        )
        for i, jr in enumerate(jrows):
            dt = dates.get(f"journal_{i}", None)
            if dt is not None:
                await conn.execute("UPDATE journal_entries SET transaction_date = $2 WHERE id = $1", jr["id"], dt)


async def _insert_opening_balance(pool, room_id: int, asset_ledger_id: int, amount: float, equity_ledger_id: int):
    """สร้าง journal 'opening_balance' ตรง ๆ (เลียนแบบ migrate_phase2_5_opening_balance.py)."""
    import json
    async with pool.acquire() as conn:
        entry_id = await conn.fetchval(
            """INSERT INTO journal_entries (room_id, reference_type, description, recorded_by, metadata)
               VALUES ($1, 'opening_balance', 'ตั้งยอดยกมา (ระบบบัญชีคู่)', 'SYSTEM', $2::jsonb)
               RETURNING id""",
            room_id, json.dumps({"note": "test"}),
        )
        await conn.execute(
            "INSERT INTO journal_lines (journal_entry_id, ledger_id, debit, credit, line_description) VALUES ($1, $2, $3, 0, 'ยอดยกมา')",
            entry_id, asset_ledger_id, amount,
        )
        await conn.execute(
            "INSERT INTO journal_lines (journal_entry_id, ledger_id, debit, credit, line_description) VALUES ($1, $2, 0, $3, 'ทุน')",
            entry_id, equity_ledger_id, amount,
        )
        await conn.execute("UPDATE journal_entries SET transaction_date = $2 WHERE id = $1", entry_id, datetime(2026, 9, 1, 0, 0, 0))
        return entry_id


# === [ROUTER] get_transactions ===


async def test_router_pre_cutoff_uses_legacy(db_pool):
    """ช่วงก่อน cutoff (เช่น ม.ค. 2026) → อ่านจาก finance_transactions ยังมีข้อมูลอยู่"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=100.0,
                              description="บริจาคก่อน cutoff", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await _set_tx_dates(db_pool, room_id, {"legacy_0": datetime(2026, 1, 10, 9), "journal_0": datetime(2026, 1, 10, 9)})

    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test",
        start_date=date(2026, 1, 1), end_date=date(2026, 1, 31), room_id=room_id, user_id=owner,
    )
    assert data["total_count"] == 1
    assert data["items"][0]["transaction_type"] == "income"
    assert data["items"][0]["amount"] == pytest.approx(100.0)


async def test_router_post_cutoff_uses_v2(db_pool):
    """ช่วงหลัง cutoff (เช่น ต.ค. 2026) → อ่านจาก journal (v2) แต่ legacy ยังมีข้อมูล"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 1000.0)
    cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=50.0,
                              description="ซื้อของ", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await _set_tx_dates(db_pool, room_id, {"legacy_0": datetime(2026, 10, 5, 9), "journal_0": datetime(2026, 10, 5, 9)})

    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test",
        start_date=POST_CUTOFF, end_date=date(2026, 10, 31), room_id=room_id, user_id=owner,
    )
    assert data["total_count"] == 1
    item = data["items"][0]
    # [DOUBLE-ENTRY] จัดเป็น expense (Asset credit + Expense debit)
    assert item["transaction_type"] == "expense"
    assert item["amount"] == pytest.approx(50.0)
    assert item["account_name"] == "กองกลาง"
    assert item["category_name"] == "ค่าอาหาร"
    assert isinstance(item["id"], int)  # สังเคราะห์เป็น int (TransactionResponse.id)


async def test_router_no_filter_uses_legacy(db_pool):
    """ไม่ระบุช่วงเวลา = "ทั้งหมด" → ครอบทั้งสองยุค → อ่าน legacy เสมอ"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=100.0,
                              description="รายการ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test", room_id=room_id, user_id=owner,
    )
    assert data["total_count"] == 1
    assert data["items"][0]["description"] == "รายการ"


async def test_router_cross_period_start_before_cutoff_uses_legacy(db_pool):
    """ช่วงคร่อมวันที่ตัด (start ก่อน cutoff) → ต้องอ่าน legacy เพื่อไม่ให้ข้อมูลหาย"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=100.0,
                              description="รายการ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await _set_tx_dates(db_pool, room_id, {"legacy_0": datetime(2026, 8, 15, 9), "journal_0": datetime(2026, 8, 15, 9)})

    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test",
        start_date=date(2026, 8, 1), end_date=date(2026, 10, 31), room_id=room_id, user_id=owner,
    )
    # ข้อมูลอยู่ช่วง ต.ค. (หลัง cutoff) แต่ช่วงเริ่มก่อน → legacy → ยังอ่านเจอจาก legacy table
    assert data["total_count"] == 1


# === [DOUBLE-ENTRY] _get_transactions_v2: ประเภทต่าง ๆ ===


async def test_v2_reads_income_expense_transfer(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc1 = await _insert_finance_account(db_pool, room_id, "กระเป๋าหลัก", 1000.0)
    acc2 = await _insert_finance_account(db_pool, room_id, "กระเป๋ารอง", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    # income
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc1, category_id=inc_cat, amount=500.0,
                              description="รับบริจาค", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    # expense
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc1, category_id=exp_cat, amount=200.0,
                              description="ซื้อของ", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    # transfer
    await FinanceService.transfer_money(
        pool=db_pool,
        req=TransferCreate(from_account_id=acc1, to_account_id=acc2, amount=100.0,
                           description="ฝาก", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    # ปรับทุกวันที่ให้อยู่หลัง cutoff
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test",
        start_date=POST_CUTOFF, end_date=date(2026, 10, 31), room_id=room_id, user_id=owner,
    )
    assert data["total_count"] == 3

    by_type = {item["transaction_type"]: item for item in data["items"]}
    # income 500
    assert by_type["income"]["amount"] == pytest.approx(500.0)
    assert by_type["income"]["category_name"] == "เงินบริจาค"
    # expense จริง 200 (รายจ่ายจากหมวดค่าอาหาร) — transfer ก็นับเป็น expense ด้วย
    expense_item = next(i for i in data["items"] if i["category_name"] == "ค่าอาหาร")
    assert expense_item["amount"] == pytest.approx(200.0)
    # transfer แสดงเป็น expense + transfer_group_id
    transfer_items = [i for i in data["items"] if i.get("transfer_group_id") is not None]
    assert len(transfer_items) == 1
    assert transfer_items[0]["category_name"] == "โอนเงิน"


async def test_v2_reads_opening_balance_as_income(db_pool):
    """ยอดยกมา → income โดยยอด = เดบิตฝั่ง Asset"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กระเป๋ากลาง", 5000.0)
    # หา ledger ของ asset + equity
    async with db_pool.acquire() as conn:
        asset_ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND legacy_account_id = $2", room_id, acc
        )
        equity_ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND account_code = '3000'", room_id
        )
        if not equity_ledger_id:
            equity_ledger_id = await conn.fetchval(
                """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, description)
                   VALUES ($1, '3000', 'ทุน-ยอดยกมา', 'equity', 'test') RETURNING id""", room_id
            )
    await _insert_opening_balance(db_pool, room_id, asset_ledger_id, 5000.0, equity_ledger_id)

    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test",
        start_date=CUTOFF_DATE, end_date=date(2026, 10, 31), room_id=room_id, user_id=owner,
    )
    assert data["total_count"] == 1
    item = data["items"][0]
    assert item["transaction_type"] == "income"
    assert item["amount"] == pytest.approx(5000.0)
    assert "ยอดยกมา" in item["category_name"]


async def test_v2_returns_empty_when_room_has_no_journal(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)

    data = await FinanceService.get_transactions(
        pool=db_pool, client_source="test", actor_identifier="test",
        start_date=POST_CUTOFF, end_date=date(2026, 10, 31), room_id=room_id, user_id=owner,
    )
    assert data["total_count"] == 0
    assert data["items"] == []


# === [DOUBLE-ENTRY] _get_summary_v2 ===


async def test_summary_v2_aggregates_from_ledgers(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc1 = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    acc2 = await _insert_finance_account(db_pool, room_id, "เงินสด", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    # income 500 ไป กองกลาง, expense 200 จาก กองกลาง, income 300 ไป เงินสด
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc1, category_id=inc_cat, amount=500.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc1, category_id=exp_cat, amount=200.0,
                              description="จ่าย", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc2, category_id=inc_cat, amount=300.0,
                              description="รับ2", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    summary = await FinanceService.get_summary(
        pool=db_pool, client_source="test", actor_identifier="test",
        month=10, year=2026, room_id=room_id, user_id=owner,
    )
    # Net Worth = สินทรัพย์สะสม = 500 − 200 + 300 = 600
    assert float(summary["net_worth"]) == pytest.approx(600.0)
    # รายได้รวมในงวด = 500 + 300
    assert float(summary["total_income"]) == pytest.approx(800.0)
    # รายจ่ายรวมในงวด = 200
    assert float(summary["total_expense"]) == pytest.approx(200.0)
    assert summary["period"] == "2026-10"
    # expense breakdown ตาม ledger
    assert summary["expense_breakdown"][0]["category_name"] == "ค่าอาหาร"
    assert float(summary["expense_breakdown"][0]["total_amount"]) == pytest.approx(200.0)


async def test_summary_v2_net_worth_includes_opening_balance(db_pool):
    """Net Worth ต้องรวมยอดยกมา (Asset Dr) แต่ยอดยกมาไม่นับเป็นรายได้งวด"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 5000.0)
    async with db_pool.acquire() as conn:
        asset_ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND legacy_account_id = $2", room_id, acc
        )
        equity_ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND account_code = '3000'", room_id
        )
        if not equity_ledger_id:
            equity_ledger_id = await conn.fetchval(
                """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, description)
                   VALUES ($1, '3000', 'ทุน-ยอดยกมา', 'equity', 'test') RETURNING id""", room_id
            )
    await _insert_opening_balance(db_pool, room_id, asset_ledger_id, 5000.0, equity_ledger_id)
    # เพิ่มรายได้งวดถัดมา (ต.ค.)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=700.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1 AND reference_type <> 'opening_balance'", room_id)

    summary = await FinanceService.get_summary(
        pool=db_pool, client_source="test", actor_identifier="test",
        month=10, year=2026, room_id=room_id, user_id=owner,
    )
    # Net Worth = 5000 (ยกมา) + 700 (รายได้)
    assert float(summary["net_worth"]) == pytest.approx(5700.0)
    # รายได้งวด = 700 เท่านั้น (ไม่รวมยอดยกมา 5000)
    assert float(summary["total_income"]) == pytest.approx(700.0)
    assert float(summary["total_expense"]) == pytest.approx(0.0)


# === [ROUTER] export_transactions_excel ===


async def test_export_router_post_cutoff_uses_v2(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=inc_cat, amount=100.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=exp_cat, amount=30.0,
                              description="จ่าย", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    import io
    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(month=10, year=2026),
        client_source="test", actor_identifier="test", room_id=room_id, user_id=owner,
    )
    assert isinstance(excel_file, io.BytesIO)

    import openpyxl
    wb = openpyxl.load_workbook(excel_file)
    assert wb.sheetnames == ["สรุปยอด", "ประวัติรายการ", "สรุปรายหมวดหมู่"]

    ws_summary = wb["สรุปยอด"]
    by_label = {row[0]: row[1] for row in ws_summary.values if row[0] and isinstance(row[1], (int, float))}
    assert by_label["รายรับรวม"] == 100.0
    assert by_label["รายจ่ายรวม"] == 30.0

    ws_data = wb["ประวัติรายการ"]
    rows = list(ws_data.values)
    assert len(rows) == 3  # header + 2 rows
    data_rows = rows[1:]
    income_rows = [r for r in data_rows if r[3] == "รายรับ"]
    expense_rows = [r for r in data_rows if r[3] == "รายจ่าย"]
    assert len(income_rows) == 1 and income_rows[0][4] == 100.0
    assert len(expense_rows) == 1 and expense_rows[0][5] == 30.0


async def test_export_router_no_filter_uses_legacy(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=100.0,
                              description="รายการ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    excel_file = await FinanceService.export_transactions_excel(
        pool=db_pool, req=FinanceExportRequest(),
        client_source="test", actor_identifier="test", room_id=room_id, user_id=owner,
    )
    import openpyxl
    wb = openpyxl.load_workbook(excel_file)
    rows = list(wb["ประวัติรายการ"].values)
    assert len(rows) == 2  # header + 1 row


# === [DOUBLE-ENTRY] get_trial_balance ===


async def test_trial_balance_is_balanced_and_lists_ledgers(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=250.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )

    tb = await FinanceService.get_trial_balance(
        pool=db_pool, room_id=room_id, user_id=owner, client_source="test", actor_identifier="test",
    )
    assert tb["is_balanced"] is True
    # Dr รวม = Cr รวม (250 ทั้งคู่)
    assert float(tb["total_debit"]) == pytest.approx(250.0)
    assert float(tb["total_credit"]) == pytest.approx(250.0)

    ledgers = tb["ledgers"]
    types = {l["account_type"] for l in ledgers}
    assert "asset" in types and "revenue" in types
    asset = next(l for l in ledgers if l["account_type"] == "asset")
    revenue = next(l for l in ledgers if l["account_type"] == "revenue")
    assert asset["balance"] == pytest.approx(250.0)   # Dr − Cr
    assert revenue["balance"] == pytest.approx(250.0)  # Cr − Dr


async def test_trial_balance_as_of_date(db_pool):
    """as_of_date ก่อนรายการ → ยอดเป็น 0"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=250.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    tb = await FinanceService.get_trial_balance(
        pool=db_pool, room_id=room_id, user_id=owner, client_source="test", actor_identifier="test",
        as_of_date=date(2026, 9, 30),
    )
    assert float(tb["total_debit"]) == pytest.approx(0.0)
    assert float(tb["total_credit"]) == pytest.approx(0.0)
    assert tb["is_balanced"] is True


# === [DOUBLE-ENTRY] get_income_statement ===


async def test_income_statement_calculates_net_income(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=inc_cat, amount=500.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=exp_cat, amount=120.0,
                              description="จ่าย", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    stmt = await FinanceService.get_income_statement(
        pool=db_pool, room_id=room_id, start_date=POST_CUTOFF, end_date=date(2026, 10, 31),
        client_source="test", actor_identifier="test", user_id=owner,
    )
    assert float(stmt["total_revenue"]) == pytest.approx(500.0)
    assert float(stmt["total_expense"]) == pytest.approx(120.0)
    assert float(stmt["net_income"]) == pytest.approx(380.0)
    assert stmt["revenues"][0]["account_name"] == "เงินบริจาค"
    assert stmt["expenses"][0]["account_name"] == "ค่าอาหาร"


async def test_income_statement_excludes_opening_balance(db_pool):
    """ยอดยกมาไม่นับเป็นรายได้"""
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 5000.0)
    async with db_pool.acquire() as conn:
        asset_ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND legacy_account_id = $2", room_id, acc
        )
        equity_ledger_id = await conn.fetchval(
            "SELECT id FROM accounting_ledgers WHERE room_id = $1 AND account_code = '3000'", room_id
        )
        if not equity_ledger_id:
            equity_ledger_id = await conn.fetchval(
                """INSERT INTO accounting_ledgers (room_id, account_code, account_name, account_type, description)
                   VALUES ($1, '3000', 'ทุน-ยอดยกมา', 'equity', 'test') RETURNING id""", room_id
            )
    await _insert_opening_balance(db_pool, room_id, asset_ledger_id, 5000.0, equity_ledger_id)

    stmt = await FinanceService.get_income_statement(
        pool=db_pool, room_id=room_id, start_date=POST_CUTOFF, end_date=date(2026, 10, 31),
        client_source="test", actor_identifier="test", user_id=owner,
    )
    assert float(stmt["total_revenue"]) == pytest.approx(0.0)
    assert float(stmt["net_income"]) == pytest.approx(0.0)


# === RBAC: v2 methods ต้องเช็คสมาชิกเหมือนเดิม ===


async def test_v2_reads_require_membership(db_pool):
    owner = await _insert_user(db_pool)
    room_id = await _insert_room(db_pool, owner)
    outsider = await _insert_user(db_pool, first_name="Outsider", last_name="User")

    with pytest.raises(ForbiddenError):
        await FinanceService.get_transactions(
            pool=db_pool, client_source="test", actor_identifier="test",
            start_date=POST_CUTOFF, end_date=date(2026, 10, 31), room_id=room_id, user_id=outsider,
        )
    with pytest.raises(ForbiddenError):
        await FinanceService.get_trial_balance(
            pool=db_pool, room_id=room_id, user_id=outsider, client_source="test", actor_identifier="test",
        )
    with pytest.raises(ForbiddenError):
        await FinanceService.get_income_statement(
            pool=db_pool, room_id=room_id, start_date=POST_CUTOFF, end_date=date(2026, 10, 31),
            client_source="test", actor_identifier="test", user_id=outsider,
        )


# =====================================================================
# HTTP layer — trial-balance & income-statement endpoints
# =====================================================================
from fastapi.testclient import TestClient
from core.config import settings


def _room_api(target_id: int, path: str) -> str:
    return f"/api/classroom/{target_id}{path}?target_type=room"


def _web_headers(user_id: int) -> dict:
    from jose import jwt
    token = jwt.encode(
        {"user_id": user_id, "exp": 9999999999},
        settings.JWT_SECRET,
        algorithm=settings.JWT_ALGORITHM,
    )
    return {"Authorization": f"Bearer {token}"}


async def test_http_trial_balance_endpoint_200(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=250.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    resp = client.get(_room_api(room_id, "/finance/trial-balance"), headers=_web_headers(owner))
    assert resp.status_code == 200
    body = resp.json()
    assert body["is_balanced"] is True
    assert float(body["total_debit"]) == pytest.approx(250.0)
    assert float(body["total_credit"]) == pytest.approx(250.0)
    assert any(l["account_type"] == "asset" for l in body["ledgers"])
    assert any(l["account_type"] == "revenue" for l in body["ledgers"])


async def test_http_trial_balance_as_of_query_200(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=cat, amount=250.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    resp = client.get(
        _room_api(room_id, "/finance/trial-balance") + "&as_of_date=2026-09-30",
        headers=_web_headers(owner),
    )
    assert resp.status_code == 200
    assert float(resp.json()["total_debit"]) == pytest.approx(0.0)


async def test_http_income_statement_endpoint_200(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    acc = await _insert_finance_account(db_pool, room_id, "กองกลาง", 0.0)
    inc_cat = await _insert_category(db_pool, room_id, "เงินบริจาค", "income")
    exp_cat = await _insert_category(db_pool, room_id, "ค่าอาหาร", "expense")

    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=inc_cat, amount=500.0,
                              description="รับ", transaction_type="income", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    await FinanceService.add_transaction(
        pool=db_pool,
        req=TransactionCreate(account_id=acc, category_id=exp_cat, amount=120.0,
                              description="จ่าย", transaction_type="expense", user_name="Owner"),
        user_id=owner, client_source="test", actor_identifier="test", room_id=room_id,
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE journal_entries SET transaction_date = '2026-10-10' WHERE room_id = $1", room_id)

    resp = client.get(
        _room_api(room_id, "/finance/income-statement") + "&start_date=2026-10-01&end_date=2026-10-31",
        headers=_web_headers(owner),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert float(body["total_revenue"]) == pytest.approx(500.0)
    assert float(body["total_expense"]) == pytest.approx(120.0)
    assert float(body["net_income"]) == pytest.approx(380.0)


async def test_http_income_statement_requires_membership_403(client, db_pool):
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    room_id = await _insert_room(db_pool, owner)
    outsider = await _insert_user(db_pool, first_name="Outsider", last_name="User")

    resp = client.get(
        _room_api(room_id, "/finance/income-statement") + "&start_date=2026-10-01&end_date=2026-10-31",
        headers=_web_headers(outsider),
    )
    assert resp.status_code == 403
