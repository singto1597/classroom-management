"""
Part 1 — ระบบเช็คชื่อแยกแผ่น + Bulk Update (ขยาย) + เพิ่มนักเรียน + Dynamic Fields — integration tests.

ครอบคลุมตาม docs/rules/testing.md:
  - Real Postgres + clean_database fixture (state isolation)
  - Randomized IDs — ห้าม hardcode
  - Deep DB verification: หลัง HTTP assert → query DB ตรง ๆ พิสูจน์ mutation
  - Mock ActionService.notify_new_activity (กันแตะ Redis จริง)
  - JSONB parse quirk → normalize ด้วย _parse_metadata เสมอ
"""
import io
import json
import random
import string
import uuid
from datetime import date

import openpyxl
import pytest
from unittest.mock import AsyncMock, patch

from core.config import settings
from services.action_service import ActionService
from services.activity_service import ActivityService

pytestmark = pytest.mark.asyncio


# === Fixtures & Setup (copy pattern จาก test_activity.py) ===


async def _insert_user(pool, *, first_name="Test", last_name="User", email=None, discord_id=None) -> int:
    if email is None:
        email = f"u{uuid.uuid4().hex[:12]}@test.local"
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO users (email, first_name, last_name, username, discord_id)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING id
            """,
            email, first_name, last_name, f"user_{uuid.uuid4().hex[:8]}", discord_id,
        )


async def _insert_room(pool, owner_id: int, room_name="Test Room", server_id=None) -> int:
    async with pool.acquire() as conn:
        while True:
            code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not await conn.fetchval("SELECT 1 FROM rooms WHERE room_code = $1", code):
                break
        room_id = await conn.fetchval(
            """
            INSERT INTO rooms (room_name, room_code, owner_id, server_id)
            VALUES ($1, $2, $3, $4)
            RETURNING id
            """,
            room_name, code, owner_id, server_id,
        )
        await conn.execute(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, 0, 'president', 'active', TRUE, $3::jsonb)
            """,
            room_id, owner_id, '["all"]',
        )
        return room_id


async def _insert_student(pool, room_id: int, user_id: int, student_no: int, *, status="active", permissions="[]") -> int:
    async with pool.acquire() as conn:
        return await conn.fetchval(
            """
            INSERT INTO students (room_id, user_id, student_no, class_role, status, is_admin, permissions)
            VALUES ($1, $2, $3, 'student', $4, FALSE, $5::jsonb)
            RETURNING id
            """,
            room_id, user_id, student_no, status, permissions,
        )


def _parse_metadata(raw) -> dict:
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {}
    return {}


def _make_web_headers(user_id: int) -> dict:
    from jose import jwt
    token = jwt.encode(
        {"user_id": user_id, "exp": 9999999999},
        settings.JWT_SECRET,
        algorithm=settings.JWT_ALGORITHM,
    )
    return {"Authorization": f"Bearer {token}"}


def _activity_api(room_id: int, path: str = "", target_type: str = "room") -> str:
    return f"/api/classroom/{room_id}/activities{path}?target_type={target_type}"


async def _create_activity_http(client, room_id, owner_id, *, title="กีฬาสี", activity_date="2026-10-15",
                                base_hours=8.0, status="upcoming", metadata=None, participants=None, **kw):
    payload = {
        "title": title,
        "activity_date": activity_date,
        "base_hours": base_hours,
        "status": status,
        "metadata": metadata or {},
        "participants": participants or [],
        "user_name": "ผู้ดูแล",
        **kw,
    }
    return client.post(_activity_api(room_id), json=payload, headers=_make_web_headers(owner_id))


async def _setup_activity(client, db_pool, *, student_nos=(1, 2, 3), with_server_id=False, activity_metadata=None):
    """สร้างห้อง + นักเรียน + กิจกรรม (ผ่าน HTTP) → คืน (room_id, activity_id, owner_id)"""
    owner = await _insert_user(db_pool, first_name="Admin", last_name="Owner")
    server_id = random.randint(1_000_000, 9_999_999) if with_server_id else None
    room_id = await _insert_room(db_pool, owner, server_id=server_id)
    for no in student_nos:
        await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name=f"ชื่อ{no}", last_name=f"นามสกุล{no}"), no)
    participants = [
        {"student_no": no, "role_type": "participant", "role_detail": None, "metadata": {}}
        for no in student_nos
    ]
    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        resp = await _create_activity_http(client, room_id, owner, participants=participants, metadata=activity_metadata or {})
    activity_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))
    return room_id, activity_id, owner


async def _participant_ids(db_pool, activity_id: int) -> list:
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT ap.id FROM activity_participants ap "
            "JOIN students s ON ap.student_id = s.id "
            "WHERE ap.activity_id = $1 AND ap.deleted_at IS NULL ORDER BY s.student_no ASC",
            activity_id,
        )
        return [r["id"] for r in rows]


async def _add_plain_member(db_pool, room_id: int, *, student_no: int = 99) -> int:
    member = await _insert_user(db_pool, first_name="สมาชิก", last_name="ธรรมดา")
    await _insert_student(db_pool, room_id, member, student_no, permissions="[]")
    return member


# ================================================================
# ✅ Multiple Attendance Sheets — service + HTTP
# ================================================================

async def test_checkin_sheet_create_and_list_counts(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    with patch.object(ActionService, "notify_new_activity", new_callable=AsyncMock):
        resp = client.post(
            _activity_api(room_id, f"/{activity_id}/checkins"),
            json={"title": "เช็คขึ้นรถ", "event_date": "2026-10-15", "user_name": "ผู้ดูแล"},
            headers=_make_web_headers(owner),
        )
    assert resp.status_code == 200, resp.text
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))

    # deep DB verify
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT title, event_date, activity_id FROM activity_checkin_sheets WHERE id = $1", sheet_id)
        assert row["title"] == "เช็คขึ้นรถ"
        assert row["event_date"] == date(2026, 10, 15)
        assert row["activity_id"] == activity_id

    lst = client.get(_activity_api(room_id, f"/{activity_id}/checkins"), headers=_make_web_headers(owner))
    assert lst.status_code == 200, lst.text
    data = lst.json()
    assert len(data) == 1
    assert data[0]["title"] == "เช็คขึ้นรถ"
    assert data[0]["checked_count"] == 0
    assert data[0]["total_count"] == 3  # ผู้เข้าร่วม 3 คน


async def test_checkin_sheet_update_title_and_clear_date(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็คเข้า", "event_date": "2026-10-15", "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))

    upd = client.patch(
        _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}"),
        json={"title": "เช็คเข้าฐาน", "event_date": None, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert upd.status_code == 200, upd.text
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT title, event_date FROM activity_checkin_sheets WHERE id = $1", sheet_id)
        assert row["title"] == "เช็คเข้าฐาน"
        assert row["event_date"] is None


async def test_checkin_sheet_delete_soft_deletes_records(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็คขึ้นรถ", "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))
    pid = (await _participant_ids(db_pool, activity_id))[0]

    mark = client.put(
        _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}/records/{pid}"),
        json={"is_present": True, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert mark.status_code == 200, mark.text

    d = client.request("DELETE", _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}"),
                       json={"user_name": "ผู้ดูแล"}, headers=_make_web_headers(owner))
    assert d.status_code == 200, d.text
    async with db_pool.acquire() as conn:
        sheet = await conn.fetchval("SELECT deleted_at FROM activity_checkin_sheets WHERE id = $1", sheet_id)
        rec = await conn.fetchval("SELECT deleted_at FROM activity_checkin_records WHERE sheet_id = $1 AND participant_id = $2", sheet_id, pid)
        assert sheet is not None
        assert rec is not None


async def test_checkin_record_upsert_and_checked_count(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็คขึ้นรถ", "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))
    pids = await _participant_ids(db_pool, activity_id)

    for pid in pids[:2]:
        r = client.put(
            _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}/records/{pid}"),
            json={"is_present": True, "user_name": "ผู้ดูแล"},
            headers=_make_web_headers(owner),
        )
        assert r.status_code == 200, r.text

    # PUT ซ้ำคนเดิม → ยังเป็น 1 active row (upsert ไม่ได้เพิ่มแถว)
    again = client.put(
        _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}/records/{pids[0]}"),
        json={"is_present": True, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert again.status_code == 200, again.text

    async with db_pool.acquire() as conn:
        cnt = await conn.fetchval(
            "SELECT COUNT(*) FROM activity_checkin_records WHERE sheet_id = $1 AND deleted_at IS NULL", sheet_id)
        present = await conn.fetchval(
            "SELECT COUNT(*) FROM activity_checkin_records WHERE sheet_id = $1 AND is_present = TRUE AND deleted_at IS NULL", sheet_id)
        assert cnt == 2
        assert present == 2

    lst = client.get(_activity_api(room_id, f"/{activity_id}/checkins"), headers=_make_web_headers(owner))
    assert lst.json()[0]["checked_count"] == 2
    assert lst.json()[0]["total_count"] == 3


async def test_checkin_records_batch_atomic_rollback(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็คเข้าฐาน", "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))
    pids = await _participant_ids(db_pool, activity_id)

    # participant_id จากกิจกรรมอื่น → error + rollback ทั้งชุด
    other_room, other_act, _ = await _setup_activity(client, db_pool, student_nos=(7,))
    other_pid = (await _participant_ids(db_pool, other_act))[0]

    bad = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}/records"),
        json={"records": [
            {"participant_id": pids[0], "is_present": True},
            {"participant_id": other_pid, "is_present": True},
        ], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert bad.status_code == 404, bad.text
    # rollback — pids[0] ไม่ถูก commit
    async with db_pool.acquire() as conn:
        cnt = await conn.fetchval(
            "SELECT COUNT(*) FROM activity_checkin_records WHERE sheet_id = $1 AND deleted_at IS NULL", sheet_id)
        assert cnt == 0


async def test_checkin_sheet_wrong_activity_404(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    _, other_act, _ = await _setup_activity(client, db_pool, student_nos=(7,))
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็ค A", "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))
    # sheet ของ activity นี้ → เอาไปดูกับ activity อื่น → 404
    got = client.get(_activity_api(room_id, f"/{other_act}/checkins/{sheet_id}"), headers=_make_web_headers(owner))
    assert got.status_code == 404


async def test_checkin_record_participant_not_in_activity(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    _, other_act, _ = await _setup_activity(client, db_pool, student_nos=(7,))
    other_pid = (await _participant_ids(db_pool, other_act))[0]

    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็ค A", "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    sheet_id = int(resp.json()["message"].split("ID: ")[1].rstrip(")"))
    bad = client.put(
        _activity_api(room_id, f"/{activity_id}/checkins/{sheet_id}/records/{other_pid}"),
        json={"is_present": True, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert bad.status_code == 404


async def test_checkin_sheet_write_rbac_forbidden(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    member = await _add_plain_member(db_pool, room_id)
    # สร้างแผ่นด้วย member → 403
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/checkins"),
        json={"title": "เช็ค A", "user_name": "สมาชิก"},
        headers=_make_web_headers(member),
    )
    assert resp.status_code == 403
    # อ่าน (member-readable) → 200
    lst = client.get(_activity_api(room_id, f"/{activity_id}/checkins"), headers=_make_web_headers(member))
    assert lst.status_code == 200


async def test_http_get_checkins_route_not_shadowed(db_pool, client):
    """GET /{activity_id}/checkins ต้องไม่ไปชนกับ GET /{activity_id} หรือ /me/roles"""
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    r1 = client.get(_activity_api(room_id, f"/{activity_id}/checkins"), headers=_make_web_headers(owner))
    assert r1.status_code == 200
    r2 = client.get(_activity_api(room_id, "/me/roles"), headers=_make_web_headers(owner))
    assert r2.status_code == 200
    r3 = client.get(_activity_api(room_id, f"/{activity_id}"), headers=_make_web_headers(owner))
    assert r3.status_code == 200


# ================================================================
# ➕ Available students + batch add
# ================================================================

async def test_list_available_students_excludes_existing_and_pending(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1, 2))
    # นักเรียน 3: active ยังไม่เข้ากิจกรรม → ควรอยู่ใน available
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="ใหม่", last_name="สุด"), 3)
    # นักเรียน 4: pending (ไม่ active) → ไม่ควรอยู่
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="รอ", last_name="อนุมัติ"), 4, status="pending")

    av = client.get(_activity_api(room_id, f"/{activity_id}/participants/available"), headers=_make_web_headers(owner))
    assert av.status_code == 200, av.text
    nos = [s["student_no"] for s in av.json()]
    assert 3 in nos
    assert 1 not in nos  # เข้ากิจกรรมแล้ว
    assert 4 not in nos  # pending


async def test_http_get_available_route_not_shadowed(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    av = client.get(_activity_api(room_id, f"/{activity_id}/participants/available"), headers=_make_web_headers(owner))
    assert av.status_code == 200
    # /me/roles ยัง 200 (ไม่ถูก shadow)
    r2 = client.get(_activity_api(room_id, "/me/roles"), headers=_make_web_headers(owner))
    assert r2.status_code == 200


async def test_batch_add_participants_inserts_and_audits(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1,))
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="ใหม่", last_name="หนึ่ง"), 2)
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="ใหม่", last_name="สอง"), 3)

    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [
            {"student_no": 2, "role_type": "staff", "metadata": {"bus_number": "B1"}},
            {"student_no": 3},
        ], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200, resp.text
    assert "เพิ่มผู้เข้าร่วม 2 คนสำเร็จ" in resp.json()["message"]

    async with db_pool.acquire() as conn:
        cnt = await conn.fetchval("SELECT COUNT(*) FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL", activity_id)
        assert cnt == 3
        row = await conn.fetchrow(
            "SELECT ap.role_type, ap.metadata FROM activity_participants ap "
            "JOIN students s ON ap.student_id = s.id WHERE ap.activity_id = $1 AND s.student_no = 2 AND ap.deleted_at IS NULL",
            activity_id,
        )
        assert row["role_type"] == "staff"
        assert _parse_metadata(row["metadata"]).get("bus_number") == "B1"
        audits = await conn.fetchval(
            "SELECT COUNT(*) FROM audit_logs WHERE entity_type = 'ACTIVITY_PARTICIPANT' AND action = 'CREATE' AND endpoint_or_command = 'batch_add_participants'"
        )
        assert audits == 2


async def test_batch_add_participants_revives_soft_deleted(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1, 2))
    pids = await _participant_ids(db_pool, activity_id)
    # นำ student_no 1 ออก (soft delete)
    d = client.request("DELETE", _activity_api(room_id, f"/{activity_id}/participants/{pids[0]}"),
                       json={"user_name": "ผู้ดูแล"}, headers=_make_web_headers(owner))
    assert d.status_code == 200, d.text

    # batch add student_no 1 กลับ → revive แถวเดิม ไม่ INSERT ซ้ำ
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [{"student_no": 1}], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200, resp.text
    async with db_pool.acquire() as conn:
        cnt = await conn.fetchval("SELECT COUNT(*) FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL", activity_id)
        assert cnt == 2  # revive ไม่ใช่แถวใหม่
        revived = await conn.fetchval("SELECT id FROM activity_participants WHERE activity_id = $1 AND student_id = $2 AND deleted_at IS NULL",
                                      activity_id, await conn.fetchval("SELECT student_id FROM activity_participants WHERE id = $1", pids[0]))
        assert revived == pids[0]


async def test_batch_add_participants_duplicate_rejected(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1,))
    await _insert_student(db_pool, room_id, await _insert_user(db_pool, first_name="ใหม่", last_name="หนึ่ง"), 2)
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [{"student_no": 2}, {"student_no": 2}], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 400
    async with db_pool.acquire() as conn:
        cnt = await conn.fetchval("SELECT COUNT(*) FROM activity_participants WHERE activity_id = $1 AND deleted_at IS NULL", activity_id)
        assert cnt == 1  # rollback — ไม่มีแถวใหม่


async def test_batch_add_participants_rbac_forbidden(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1,))
    member = await _add_plain_member(db_pool, room_id)
    resp = client.post(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [{"student_no": 1}], "user_name": "สมาชิก"},
        headers=_make_web_headers(member),
    )
    assert resp.status_code == 403
    # available (member-readable) → 200
    av = client.get(_activity_api(room_id, f"/{activity_id}/participants/available"), headers=_make_web_headers(member))
    assert av.status_code == 200


# ================================================================
# ⚡ Extended batch — role_type / status / earned_hours
# ================================================================

async def test_batch_update_sets_role_type_status_earned_hours(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    pids = await _participant_ids(db_pool, activity_id)

    resp = client.patch(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [
            {"participant_id": pids[0], "role_type": "leader", "status": "attended",
             "earned_hours": 5.5, "metadata": {"bus_number": "A1"}},
            {"participant_id": pids[1], "status": "cancelled"},
        ], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200, resp.text

    async with db_pool.acquire() as conn:
        r1 = await conn.fetchrow("SELECT role_type, status, earned_hours, metadata FROM activity_participants WHERE id = $1", pids[0])
        assert r1["role_type"] == "leader"
        assert r1["status"] == "attended"
        assert float(r1["earned_hours"]) == 5.5
        assert _parse_metadata(r1["metadata"]).get("bus_number") == "A1"
        r2 = await conn.fetchrow("SELECT role_type, status, earned_hours FROM activity_participants WHERE id = $1", pids[1])
        assert r2["role_type"] == "participant"  # ไม่ส่ง → ไม่แตะ
        assert r2["status"] == "cancelled"
        assert float(r2["earned_hours"]) == 0.0


async def test_batch_update_extended_none_preserves_existing(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool)
    pids = await _participant_ids(db_pool, activity_id)
    # เซ็ต role_type = staff ก่อน
    await ActivityService.batch_update_participants(
        pool=db_pool, activity_id=activity_id,
        items=[{"participant_id": pids[0], "role_type": "staff", "role_detail": None, "metadata": {"bus_number": "X1"}}],
        user_name="ผู้ดูแล", client_source="TEST", actor_identifier="user_id:1",
        room_id=room_id, actor_user_id=owner,
    )
    # PATCH ไม่ส่ง role_type/status/hours → ไม่แตะ
    resp = client.patch(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [{"participant_id": pids[0], "metadata": {"seat_number": "A2"}}], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert resp.status_code == 200, resp.text
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT role_type, status, metadata FROM activity_participants WHERE id = $1", pids[0])
        assert row["role_type"] == "staff"
        meta = _parse_metadata(row["metadata"])
        assert meta.get("bus_number") == "X1"
        assert meta.get("seat_number") == "A2"


# ================================================================
# 🧩 Dynamic Fields — definition validation + participant values + export
# ================================================================

async def test_update_activity_dynamic_fields_valid(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1,))
    defs = [
        {"key": "df_1", "label": "หมายเลขกลุ่ม", "type": "input"},
        {"key": "df_2", "label": "ไปกับรถบัส", "type": "boolean"},
    ]
    upd = client.patch(
        _activity_api(room_id, f"/{activity_id}"),
        json={"metadata": {"dynamic_fields": defs}, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert upd.status_code == 200, upd.text
    got = client.get(_activity_api(room_id, f"/{activity_id}"), headers=_make_web_headers(owner))
    meta = got.json()["metadata"]
    assert _parse_metadata(meta if isinstance(meta, str) else json.dumps(meta)).get("dynamic_fields") == defs


async def test_update_activity_dynamic_fields_duplicate_key_rejected(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1,))
    bad = client.patch(
        _activity_api(room_id, f"/{activity_id}"),
        json={"metadata": {"dynamic_fields": [
            {"key": "df_1", "label": "กลุ่ม", "type": "input"},
            {"key": "df_1", "label": "กลุ่มซ้ำ", "type": "input"},
        ]}, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert bad.status_code == 400


async def test_update_activity_dynamic_fields_empty_label_rejected(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1,))
    bad = client.patch(
        _activity_api(room_id, f"/{activity_id}"),
        json={"metadata": {"dynamic_fields": [{"key": "df_1", "label": "  ", "type": "input"}]}, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert bad.status_code == 400


async def test_dynamic_field_value_round_trip(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1, 2))
    pids = await _participant_ids(db_pool, activity_id)
    # เพิ่ม def แล้วตั้งค่าให้ participant ผ่าน batch (df_1 key อยู่ใน metadata)
    upd = client.patch(
        _activity_api(room_id, f"/{activity_id}"),
        json={"metadata": {"dynamic_fields": [{"key": "df_1", "label": "หมายเลขกลุ่ม", "type": "input"}]}, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert upd.status_code == 200, upd.text

    batch = client.patch(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [
            {"participant_id": pids[0], "metadata": {"df_1": "กลุ่มแดง"}},
            {"participant_id": pids[1], "metadata": {"df_1": "กลุ่มน้ำเงิน"}},
        ], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert batch.status_code == 200, batch.text

    got = client.get(_activity_api(room_id, f"/{activity_id}"), headers=_make_web_headers(owner))
    parts = got.json()["participants"]
    values = {}
    for p in parts:
        meta = _parse_metadata(p["metadata"] if isinstance(p["metadata"], str) else json.dumps(p["metadata"]))
        values[p["student_no"]] = meta.get("df_1")
    assert values.get(1) == "กลุ่มแดง"
    assert values.get(2) == "กลุ่มน้ำเงิน"


async def test_export_dynamic_fields_column_and_summary(db_pool, client):
    room_id, activity_id, owner = await _setup_activity(client, db_pool, student_nos=(1, 2))
    pids = await _participant_ids(db_pool, activity_id)
    upd = client.patch(
        _activity_api(room_id, f"/{activity_id}"),
        json={"metadata": {"dynamic_fields": [{"key": "df_1", "label": "หมายเลขกลุ่ม", "type": "input"}]}, "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert upd.status_code == 200, upd.text
    batch = client.patch(
        _activity_api(room_id, f"/{activity_id}/participants/batch"),
        json={"items": [{"participant_id": pids[0], "metadata": {"df_1": "กลุ่มแดง"}}], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert batch.status_code == 200, batch.text

    exp = client.post(
        _activity_api(room_id, "/export"),
        json={"activity_id": activity_id, "metadata_keys": [], "user_name": "ผู้ดูแล"},
        headers=_make_web_headers(owner),
    )
    assert exp.status_code == 200, exp.text
    wb = openpyxl.load_workbook(io.BytesIO(exp.content))
    ws = wb["รายชื่อผู้เข้าร่วม"]
    headers = [c.value for c in ws[1]]
    assert "หมายเลขกลุ่ม" in headers  # header = label ที่ตั้ง ไม่ใช่คีย์ df_1
    assert "df_1" not in headers
    # ค่าคอลัมน์ของแถว participant 1
    col_idx = headers.index("หมายเลขกลุ่ม") + 1
    row_values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_values.append(row)
    # participant 1 (student_no=1) ควรมี "กลุ่มแดง"
    found = any(r[col_idx - 1] == "กลุ่มแดง" for r in row_values)
    assert found
    # สรุป: ไม่มีคีย์ดิบ dynamic_fields ใน summary sheet (ใช้ label "ฟิลด์เพิ่มเติมต่อคน" แทน)
    summary_text = ""
    for row in wb["สรุป"].iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                summary_text += str(cell) + " "
    assert "dynamic_fields" not in summary_text
